import json
import time
from datetime import datetime, timedelta, date
import pathlib
import pandas as pd
import pygal
import random
import sqlalchemy as sa
from StyleFrame import StyleFrame, Styler, utils
from flask import render_template, flash, redirect, url_for, request, g, \
    jsonify, current_app, send_file, Response
from flask_babel import _, get_locale
from flask_login import current_user, login_required
from pygal.style import BlueStyle
from sqlalchemy import desc
from wtforms import FieldList, FormField

from app import db
from app.admin.forms import CellsForm, FuelForm
from app.main import bp
from app.main.forms import EditProfileForm, SearchForm, MessageForm, ManualInputForm
from app.models import Close1Tank1, Close1Tank2, Close1Tank3, Close1Tank4, Close2Tank1, Close2Tank2, \
    Close2Tank3, Close2Tank4, Close3Tank1, Close3Tank2, Close3Tank3, Close3Tank4, TruckFalse, RealisationStats, \
    TempAzsTrucks3, TempAzsTrucks4, VariantNalivaForTrip, TempAzsTrucks, TempAzsTrucks2, UserLogs, GlobalSettings, \
    TempAzsTrucks2SecondTrip, TempAzsTrucks3SecondTrip, TempAzsTrucks4SecondTrip
from app.models import User, Message, Notification, FuelResidue, AzsList, Tanks, FuelRealisation, Priority, \
    PriorityList, ManualInfo, Trucks, TruckTanks, Trip, WorkType, Errors, \
    Trips, Result, TrucksForAzs, VariantSlivaForTrip, TruckTanksVariations
from app.translate import translate
from openpyxl import load_workbook


@bp.route('/stats', methods=['GET', 'POST'])
@login_required
def stats():
    # Graphs
    labels = []
    fuel_92 = list()
    fuel_95 = list()
    fuel_50 = list()
    previous_month = datetime.today() - timedelta(days=30)
    select_last_month = RealisationStats.query.filter(RealisationStats.date <= datetime.today(),
                                                      previous_month < RealisationStats.date).filter_by(azs_id=1).all()
    for row in select_last_month:
        if row.fuel_type == 92:
            fuel_92.append(row.realisation)
            labels.append(datetime.strftime(row.date, "%d/%m"))
        elif row.fuel_type == 95:
            fuel_95.append(row.realisation)
        elif row.fuel_type == 50:
            fuel_50.append(row.realisation)

    graph = pygal.StackedLine(fill=True, style=BlueStyle, height=500)
    graph.title = 'Реализация топлива в течение месяца'
    graph.x_labels = labels
    graph.add('АИ-92', fuel_92)
    graph.add('АИ-95', fuel_95)
    graph.add('Дт', fuel_50)
    graph_data = graph.render_data_uri()

    return render_template('stats.html', title='Статистика по реализации за месяц', stats=True, graph_data=graph_data)


@bp.before_app_request
def before_request():
    if current_user.is_authenticated:
        current_user.last_seen = datetime.utcnow()
        db.session.commit()
        g.search_form = SearchForm()
    g.locale = str(get_locale())


@bp.route('/', methods=['GET', 'POST'])
@bp.route('/index', methods=['GET', 'POST'])
@login_required
def index():
    def check():
        errors = 0
        azs_list = AzsList.query.all()
        error_tank_list = list()
        errors_list = list()
        for azs in azs_list:
            if azs.active:
                tanks_list = Tanks.query.filter_by(azs_id=azs.id, active=True).all()
                for tank in tanks_list:
                    if tank.active:
                        residue = FuelResidue.query.filter_by(tank_id=tank.id).first()
                        realisation = FuelRealisation.query.filter_by(tank_id=tank.id).first()
                        if residue is not None:
                            if residue.fuel_volume is None or residue.fuel_volume <= 0 \
                                    or residue.download_time < datetime.now() - timedelta(seconds=900):
                                if residue.download_time < datetime.now() - timedelta(seconds=900):
                                    error_text = "АЗС №" + str(azs.number) + ", резервуар №" + str(tank.tank_number) + \
                                                 " - возможно данные об остатках устарели"
                                else:
                                    error_text = "АЗС №" + str(azs.number) + ", резервуар №" + str(tank.tank_number) + \
                                                 " - нет данных об остатках"
                                sql = Errors(timestamp=datetime.now(), error_text=error_text, azs_id=azs.id,
                                             tank_id=tank.id, active=True, error_type="residue_error")
                                db.session.add(sql)
                                db.session.commit()
                                errors = errors + 1
                                error_tank_list.append(tank.id)

                                errors_list.append(error_text)

                        if realisation.fuel_realisation_1_days is None or realisation.fuel_realisation_1_days <= 0 \
                                or realisation.download_time < datetime.now() - timedelta(seconds=900):
                            if realisation.download_time < datetime.now() - timedelta(seconds=900):
                                error_text = "АЗС №" + str(azs.number) + ", резервуар №" + str(tank.tank_number) + \
                                             " - возможно данные о реализации устарели"
                            else:
                                error_text = "АЗС №" + str(azs.number) + ", резервуар №" + str(tank.tank_number) + \
                                             " - нет данных о реализации"
                            sql = Errors(timestamp=datetime.now(), error_text=error_text, azs_id=azs.id,
                                         tank_id=tank.id, active=True, error_type="time_error")
                            db.session.add(sql)
                            db.session.commit()
                            errors = errors + 1
                            error_tank_list.append(tank.id)
                            errors_list.append(error_text)
                        priority_list = PriorityList.query.all()
                        for priority in priority_list:
                            if priority.day_stock_from <= realisation.days_stock_min <= priority.day_stock_to:
                                this_priority = PriorityList.query.filter_by(priority=priority.priority).first_or_404()
                                if not this_priority.id:
                                    print('Резервуар №' + str(tank.tank_number) +
                                          ' не попадает в диапазон приоритетов!!!')
        return errors, errors_list

    trips = Trips.query.order_by(desc("date")).first()
    if trips.date.strftime("%d.%m.%Y") == datetime.today().strftime("%d.%m.%Y"):
        trips_today = True
    else:
        trips_today = False

    priority = Priority.query.order_by('priority').all()
    azs_list = list()
    error_list = list()
    for i in priority:
        azs_dict = {}
        if i.day_stock <= 2:
            azs = AzsList.query.filter_by(id=i.azs_id).first_or_404()
            tank = Tanks.query.filter_by(id=i.tank_id).first_or_404()
            azs_dict = {'number': azs.number,
                        'azs_id': i.azs_id,
                        'tank': tank.tank_number,
                        'day_stock': i.day_stock}
            azs_list.append(azs_dict)
    errors_num, errors = check()
    for error in errors:
        error_list.append(error)
    return render_template('index.html', title='Главная', azs_list=azs_list, error_list=error_list, index=True)


@bp.route('/user/<username>')
@login_required
def user(username):
    user = User.query.filter_by(username=username).first_or_404()
    page = request.args.get('page', 1, type=int)
    posts = user.posts.order_by(Post.timestamp.desc()).paginate(
        page, current_app.config['POSTS_PER_PAGE'], False)
    next_url = url_for('main.user', username=user.username,
                       page=posts.next_num) if posts.has_next else None
    prev_url = url_for('main.user', username=user.username,
                       page=posts.prev_num) if posts.has_prev else None
    return render_template('user.html', user=user, posts=posts.items,
                           next_url=next_url, prev_url=prev_url)


@bp.route('/user/<username>/popup')
@login_required
def user_popup(username):
    user = User.query.filter_by(username=username).first_or_404()
    return render_template('user_popup.html', user=user)


@bp.route('/edit_profile', methods=['GET', 'POST'])
@login_required
def edit_profile():
    form = EditProfileForm(current_user.username)
    if form.validate_on_submit():
        current_user.username = form.username.data
        current_user.about_me = form.about_me.data
        db.session.commit()
        flash(_('Your changes have been saved.'))
        return redirect(url_for('main.edit_profile'))
    elif request.method == 'GET':
        form.username.data = current_user.username
        form.about_me.data = current_user.about_me
    return render_template('edit_profile.html', title=_('Edit Profile'),
                           form=form)


@bp.route('/follow/<username>')
@login_required
def follow(username):
    user = User.query.filter_by(username=username).first()
    if user is None:
        flash(_('User %(username)s not found.', username=username))
        return redirect(url_for('main.index'))
    if user == current_user:
        flash(_('You cannot follow yourself!'))
        return redirect(url_for('main.user', username=username))
    current_user.follow(user)
    db.session.commit()
    flash(_('You are following %(username)s!', username=username))
    return redirect(url_for('main.user', username=username))


@bp.route('/unfollow/<username>')
@login_required
def unfollow(username):
    user = User.query.filter_by(username=username).first()
    if user is None:
        flash(_('User %(username)s not found.', username=username))
        return redirect(url_for('main.index'))
    if user == current_user:
        flash(_('You cannot unfollow yourself!'))
        return redirect(url_for('main.user', username=username))
    current_user.unfollow(user)
    db.session.commit()
    flash(_('You are not following %(username)s.', username=username))
    return redirect(url_for('main.user', username=username))


@bp.route('/translate', methods=['POST'])
@login_required
def translate_text():
    return jsonify({'text': translate(request.form['text'],
                                      request.form['source_language'],
                                      request.form['dest_language'])})


@bp.route('/search')
@login_required
def search():
    if not g.search_form.validate():
        return redirect(url_for('main.explore'))
    page = request.args.get('page', 1, type=int)
    posts, total = Post.search(g.search_form.q.data, page,
                               current_app.config['POSTS_PER_PAGE'])
    next_url = url_for('main.search', q=g.search_form.q.data, page=page + 1) \
        if total > page * current_app.config['POSTS_PER_PAGE'] else None
    prev_url = url_for('main.search', q=g.search_form.q.data, page=page - 1) \
        if page > 1 else None
    return render_template('search.html', title=_('Search'), posts=posts,
                           next_url=next_url, prev_url=prev_url)


@bp.route('/send_message/<recipient>', methods=['GET', 'POST'])
@login_required
def send_message(recipient):
    user = User.query.filter_by(username=recipient).first_or_404()
    form = MessageForm()
    if form.validate_on_submit():
        msg = Message(author=current_user, recipient=user,
                      body=form.message.data)
        db.session.add(msg)
        user.add_notification('unread_message_count', user.new_messages())
        db.session.commit()
        flash(_('Your message has been sent.'))
        return redirect(url_for('main.user', username=recipient))
    return render_template('send_message.html', title=_('Send Message'),
                           form=form, recipient=recipient)


@bp.route('/messages')
@login_required
def messages():
    current_user.last_message_read_time = datetime.utcnow()
    current_user.add_notification('unread_message_count', 0)
    db.session.commit()
    page = request.args.get('page', 1, type=int)
    messages = current_user.messages_received.order_by(
        Message.timestamp.desc()).paginate(
        page, current_app.config['POSTS_PER_PAGE'], False)
    next_url = url_for('main.messages', page=messages.next_num) \
        if messages.has_next else None
    prev_url = url_for('main.messages', page=messages.prev_num) \
        if messages.has_prev else None
    return render_template('messages.html', messages=messages.items,
                           next_url=next_url, prev_url=prev_url)


@bp.route('/export_posts')
@login_required
def export_posts():
    if current_user.get_task_in_progress('export_posts'):
        flash(_('An export task is currently in progress'))
    else:
        current_user.launch_task('export_posts', _('Exporting posts...'))
        db.session.commit()
    return redirect(url_for('main.user', username=current_user.username))


@bp.route('/notifications')
@login_required
def notifications():
    since = request.args.get('since', 0.0, type=float)
    notifications = current_user.notifications.filter(
        Notification.timestamp > since).order_by(Notification.timestamp.asc())
    return jsonify([{
        'name': n.name,
        'data': n.get_data(),
        'timestamp': n.timestamp
    } for n in notifications])


@bp.route('/residue_xlsx/остатки_<datetime>')
@login_required
def residue_xlsx(datetime):
    timenow = datetime

    online = FuelResidue.query.outerjoin(AzsList).outerjoin(Tanks).order_by(AzsList.number).all()
    online_list = list()
    for data in online:
        azs_number = AzsList.query.filter_by(id=data.azs_id).first()
        tank_number = Tanks.query.filter_by(id=data.tank_id).first()
        if data.auto:
            auto = "Автоматически"
        else:
            auto = "По книжным остаткам"
        online_dict = {
            '№': 'АЗС № ' + str(azs_number.number),
            'Резервуар №': tank_number.tank_number,
            'Вид топлива': data.product_code,
            'Процент (%)': data.percent,
            'Текущий остаток (л)': data.fuel_volume,
            'Свободная емкость (до 95%) (л)': round(tank_number.corrected_capacity - data.fuel_volume, 1),
            'Время замера АИСом': data.datetime,
            'Время получения данных': data.download_time,
            'Тип формирования': auto
        }
        online_list.append(online_dict)
    df = pd.DataFrame(online_list)
    excel_writer = StyleFrame.ExcelWriter(
        r'/home/administrator/aob-test/files/онлайн-остатки_' + str(timenow) + '.xlsx')
    sf = StyleFrame(df)

    sf.apply_style_by_indexes(indexes_to_style=sf[sf['Текущий остаток (л)'] == 0],
                              cols_to_style=['Время получения данных',
                                             'Тип формирования', 'Время замера АИСом',
                                             'Свободная емкость (до 95%) (л)',
                                             'Резервуар №',
                                             '№',
                                             'Вид топлива',
                                             'Текущий остаток (л)',
                                             'Процент (%)'], styler_obj=Styler(bg_color='F5455F'))

    sf.apply_style_by_indexes(indexes_to_style=sf[sf['Тип формирования'] == 'По книжным остаткам'],
                              cols_to_style=['Время получения данных',
                                             'Тип формирования',
                                             'Время замера АИСом',
                                             'Свободная емкость (до 95%) (л)',
                                             'Резервуар №',
                                             '№',
                                             'Вид топлива',
                                             'Текущий остаток (л)',
                                             'Процент (%)'], styler_obj=Styler(bg_color='BFEDFF'))

    sf.apply_style_by_indexes(indexes_to_style=sf[sf['Тип формирования'] == 'По книжным остаткам'],
                              cols_to_style=['Время получения данных',
                                             'Время замера АИСом'],
                              styler_obj=Styler(number_format=utils.number_formats.date_time, bg_color='BFEDFF'))

    sf.apply_style_by_indexes(indexes_to_style=sf[sf['Текущий остаток (л)'] == 0],
                              cols_to_style=['Время получения данных',
                                             'Время замера АИСом'],
                              styler_obj=Styler(number_format=utils.number_formats.date_time, bg_color='F5455F'))

    sf.to_excel(excel_writer=excel_writer, row_to_add_filters=0,
                columns_and_rows_to_freeze='A1', best_fit=['Время получения данных',
                                                           'Тип формирования', 'Время замера АИСом',
                                                           'Свободная емкость (до 95%) (л)',
                                                           'Резервуар №',
                                                           '№',
                                                           'Вид топлива',
                                                           'Процент (%)',
                                                           'Текущий остаток (л)',
                                                           'Процент (%)'
                                                           ])

    # df.to_excel(r'/home/administrator/aob-test/files/онлайн-остатки_'+str(timenow)+'.xlsx')
    excel_writer.save()
    path = '/home/administrator/aob-test/files/онлайн-остатки_' + str(timenow) + '.xlsx'
    return send_file(path)


@bp.route('/realisation_xlsx/реализация_<datetime>')
@login_required
def realisation_xlsx(datetime):
    timenow = datetime

    realisation = FuelRealisation.query.outerjoin(AzsList).outerjoin(Tanks).order_by(AzsList.number).all()
    realisation_list = list()
    for data in realisation:
        azs_number = AzsList.query.filter_by(id=data.azs_id).first()
        tank_number = Tanks.query.filter_by(id=data.tank_id).first()
        realisation_dict = {
            '№': 'АЗС № ' + str(azs_number.number),
            'Резервуар №': tank_number.tank_number,
            'Вид топлива': data.product_code,
            'Сред. за 10 дней(л)': data.average_10_days,
            'Сред. за 7 дней(л)': data.average_7_days,
            'Сред. за 3 дня(л)': data.average_3_days,
            'Реализация за 1 день(л)': data.fuel_realisation_1_days,
            'Реализация за 1 день(л) неделю назад': data.fuel_realisation_week_ago,
            'Реализация за 1 час': data.fuel_realisation_hour,
            'Мин. запас суток': data.days_stock_min,
            'Время выгрузки': data.download_time
        }
        realisation_list.append(realisation_dict)
    df = pd.DataFrame(realisation_list)
    excel_writer = StyleFrame.ExcelWriter(r'/home/administrator/aob-test/files/реализация_' + str(timenow) + '.xlsx')
    sf = StyleFrame(df)
    sf.to_excel(excel_writer=excel_writer, row_to_add_filters=0,
                columns_and_rows_to_freeze='A1', best_fit=['№',
                                                           'Резервуар №',
                                                           'Вид топлива',
                                                           'Сред. за 10 дней(л)',
                                                           'Сред. за 7 дней(л)',
                                                           'Сред. за 3 дня(л)',
                                                           'Реализация за 1 день(л)',
                                                           'Реализация за 1 день(л) неделю назад',
                                                           'Реализация за 1 час',
                                                           'Мин. запас суток',
                                                           'Время выгрузки'
                                                           ])
    excel_writer.save()
    path = '/home/administrator/aob-test/files/реализация_' + str(timenow) + '.xlsx'
    return send_file(path)


@bp.route('/online', methods=['POST', 'GET'])
@login_required
def online():
    azs_list = AzsList.query.order_by('number').all()
    online = FuelResidue.query.outerjoin(AzsList).outerjoin(Tanks).order_by(AzsList.number).all()
    tanks_list = Tanks.query.all()
    return render_template('online.html', title='Online остатки', online_active=True,
                           online=online, azs_list=azs_list, tanks_list=tanks_list,
                           datetime=datetime.now().strftime("%Y-%m-%d-%H-%M"))


@bp.route('/online.json', methods=['POST', 'GET'])
@login_required
def online_json():
    rows = list()
    online = FuelResidue.query.outerjoin(AzsList).outerjoin(Tanks).order_by(AzsList.number).all()
    for i in online:
        tank = Tanks.query.filter_by(id=i.tank_id).first()
        if tank.deactive != True:
            azs_number = AzsList.query.filter_by(id=i.azs_id).first()
            if len(str(azs_number.number)) == 1:
                azs_number = str(0) + str(azs_number.number)
            else:
                azs_number = str(azs_number.number)
            if i.auto == True:
                auto = "Автоматически"
            else:
                auto = "По книжным остаткам"
            url_name = "АЗС №" + str(azs_number)
            url = '<p ' + azs_number + '> </p> <a href="' + str(
                url_for('main.page_azs', id=i.azs_id)) + '">' + url_name + '</a>'
            tank_number = Tanks.query.filter_by(id=i.tank_id).first()
            row = {'azs_number': url,
                   'url': url_for('main.page_azs', id=i.azs_id),
                   'tank_number': tank_number.tank_number,
                   'product_code': i.product_code,
                   'percent': str(i.percent) + str("%"),
                   'fuel_volume': i.fuel_volume,
                   'free_volume': i.free_volume,
                   'datetime': i.datetime.strftime("(%H:%M) %d.%m.%Y"),
                   'download_time': i.download_time.strftime("(%H:%M) %d.%m.%Y"),
                   'auto': auto
                   }

            rows.append(row)
    return Response(json.dumps(rows), mimetype='application/json')


@bp.route('/page/azs/id<id>', methods=['POST', 'GET'])
@login_required
def page_azs(id):
    azs_list = AzsList.query.filter_by(id=id).first()
    tanks_list = Tanks.query.filter_by(azs_id=id, active=True).all()
    online = FuelResidue.query.outerjoin(Tanks).order_by(Tanks.tank_number).all()
    realisation = FuelRealisation.query.all()
    check_if_exist = FuelResidue.query.filter_by(azs_id=id).first()

    # Graphs
    labels = []
    fuel_92 = list()
    fuel_95 = list()
    fuel_50 = list()
    previous_month = datetime.today() - timedelta(days=30)
    select_last_month = RealisationStats.query.filter(RealisationStats.date <= datetime.today(),
                                                      previous_month < RealisationStats.date).filter_by(azs_id=id).all()
    for row in select_last_month:
        if row.fuel_type == 92:
            fuel_92.append(row.realisation)
            labels.append(datetime.strftime(row.date, "%d/%m"))
        elif row.fuel_type == 95:
            fuel_95.append(row.realisation)
        elif row.fuel_type == 50 or row.fuel_type:
            fuel_50.append(row.realisation)

    graph = pygal.Line(height=500)
    graph.title = 'Реализация топлива в течение месяца'
    graph.x_labels = labels
    graph.add('АИ-92', fuel_92)
    graph.add('АИ-95', fuel_95)
    graph.add('Дт', fuel_50)
    graph_data = graph.render_data_uri()
    print(fuel_92)
    print(fuel_95)
    print(fuel_50)
    return render_template('page_azs.html', title='АЗС № ' + str(azs_list.number), page_azs_active=True,
                           online=online, realisation=realisation, azs_list=azs_list, tanks_list=tanks_list,
                           azs_list_active=True, check_if_exist=check_if_exist, graph_data=graph_data)


@bp.route('/download_realisation_info')
@login_required
def download_realisation_info():
    if current_user.get_task_in_progress('download_realisation_info'):
        flash(_('Выгрузка данных уже выполняется!'))
    else:
        current_user.launch_task('download_realisation_info', _('Выгружаю данные по реализации...'))
        db.session.commit()
    return redirect(url_for('main.realisation'))


@bp.route('/realisation', methods=['POST', 'GET'])
@login_required
def realisation():
    azs_list = AzsList.query.all()
    realisation = FuelRealisation.query.order_by(FuelRealisation.shop_id).all()
    tanks_list = Tanks.query.all()
    return render_template('realisation.html', title='Реализация топлива', realisation_active=True,
                           realisation=realisation, azs_list=azs_list, tanks_list=tanks_list,
                           datetime=datetime.now().strftime("%Y-%m-%d-%H-%M"))


@bp.route('/realisation.json', methods=['POST', 'GET'])
@login_required
def realisation_json():
    rows = list()
    realisation = FuelRealisation.query.order_by(FuelRealisation.shop_id).all()
    for i in realisation:
        tank = Tanks.query.filter_by(id=i.tank_id).first()
        if tank.deactive != True:
            azs_number = AzsList.query.filter_by(id=i.azs_id).first()
            if len(str(azs_number.number)) == 1:
                azs_number = str(0) + str(azs_number.number)
            else:
                azs_number = str(azs_number.number)
            url_name = "АЗС №" + str(azs_number)
            url = '<p ' + azs_number + '> </p> <a href="' + str(
                url_for('main.page_azs', id=i.azs_id)) + '">' + url_name + '</a>'
            tank_number = Tanks.query.filter_by(id=i.tank_id).first()
            row = {'azs_number': url,
                   'tank_number': tank_number.tank_number,
                   'product_code': i.product_code,
                   'fuel_realisation_10_days': i.fuel_realisation_10_days,
                   'fuel_realisation_7_days': i.fuel_realisation_7_days,
                   'fuel_realisation_3_days': i.fuel_realisation_3_days,
                   'fuel_realisation_1_days': i.fuel_realisation_1_days,
                   'fuel_realisation_week_ago': i.fuel_realisation_week_ago,
                   'fuel_realisation_hour': i.days_stock_min,
                   'days_stock_min': i.fuel_realisation_hour,
                   'download_time': i.download_time.strftime("(%H:%M) %d.%m.%Y")
                   }

            rows.append(row)
    return Response(json.dumps(rows), mimetype='application/json')


@bp.route('/priority', methods=['GET', 'POST'])
@login_required
def priority():
    azs_list = AzsList.query.all()
    tanks_list = Tanks.query.all()
    priority = Priority.query.order_by("priority").all()
    priority_list = PriorityList.query.all()
    return render_template('priority.html', title='Список АЗС по приоритету', priority_active=True,
                           azs_list=azs_list, priority=priority, tanks_list=tanks_list, priority_list=priority_list)


@bp.route('/manual/id<id>', methods=['GET', 'POST'])
@login_required
def manual_input(id):
    tank = Tanks.query.filter_by(id=id).first_or_404()
    azs = AzsList.query.filter_by(id=tank.azs_id).first_or_404()
    form = ManualInputForm()
    if form.validate_on_submit():
        input = ManualInfo(fuel_volume=form.residue.data,
                           fuel_realisation_max=form.realisation.data,
                           tank_id=id,
                           azs_id=tank.azs_id,
                           timestamp=datetime.now())

        db.session.add(input)
        db.session.commit()
        logs = UserLogs(user_id=current_user.id,
                        action="manual_input_action",
                        timestamp=datetime.now())
        db.session.add(logs)
        db.session.commit()
        return redirect(url_for('main.index'))
    return render_template('manual_input.html', title='Ручной ввод данных',
                           manual_input=True,
                           form=form,
                           azs_number=str(azs.number),
                           tank_number=str(tank.tank_number))


@bp.route('/manual', methods=['GET', 'POST'])
@login_required
def manual():
    since = datetime.now() - timedelta(hours=3)
    manual = ManualInfo.query.filter(ManualInfo.timestamp > since).all()

    return render_template('manual.html', title='Остатки введенные в ручную', manual=manual)


@bp.route('/azs', methods=['POST', 'GET'])
@login_required
def azs_redirect():
    azs = AzsList.query.first_or_404()
    return redirect(url_for('main.azs', id=azs.id))


@bp.route('/page/azs/', methods=['POST', 'GET'])
@login_required
def azs():
    azs_list = AzsList.query.order_by("number").all()
    return render_template('azs_list.html', azs_list=azs_list, title="Список АЗС", azs_list_active=True)


@bp.route('/start', methods=['POST', 'GET'])
@login_required
def start():
    logs = UserLogs(user_id=current_user.id,
                    action="preparation_started",
                    timestamp=datetime.now())
    db.session.add(logs)
    db.session.commit()

    def check():
        logs = UserLogs(user_id=current_user.id,
                        action="error_check_started",
                        timestamp=datetime.now())
        db.session.add(logs)
        db.session.commit()
        errors = 0
        azs_list = AzsList.query.all()
        error_tank_list = list()
        for azs in azs_list:
            if azs.active:
                tanks_list = Tanks.query.filter_by(azs_id=azs.id, active=True).all()
                for tank in tanks_list:
                    if tank.active:
                        residue = FuelResidue.query.filter_by(tank_id=tank.id).first()
                        realisation = FuelRealisation.query.filter_by(tank_id=tank.id).first()
                        if residue is not None:
                            if residue.fuel_volume is None or residue.fuel_volume <= 0:
                                print('   Резервуар №' + str(tank.tank_number) + ' не содержит данных об остатках! ')
                                '''error_text = "Ошибка при подключении к базе данных АЗС №" + str(i.number)
                                sql = Errors(timestamp=datetime.now(), error_text=error_text, azs_id=i.id, active=True,
                                             error_type="connection_error")
                                db.session.add(sql)
                                db.session.commit()'''
                                errors = errors + 1
                                error_tank_list.append(tank.id)
                            if realisation.fuel_realisation_1_days is None or realisation.fuel_realisation_1_days <= 0:
                                print('   Резервуар №' + str(tank.tank_number) + ' не содержит данных о реализации! ')
                                errors = errors + 1
                                error_tank_list.append(tank.id)
                            priority_list = PriorityList.query.all()
                            for priority in priority_list:
                                if priority.day_stock_from <= realisation.days_stock_min <= priority.day_stock_to:
                                    this_priority = PriorityList.query.filter_by(
                                        priority=priority.priority).first_or_404()
                                    if not this_priority.id:
                                        print('Резервуар №' + str(tank.tank_number) +
                                              ' не попадает в диапазон приоритетов!!!')
        return errors, error_tank_list

    error, tanks = check()
    if error > 10:
        logs = UserLogs(user_id=current_user.id,
                        action="error_check_ended_with_error",
                        timestamp=datetime.now())
        db.session.add(logs)
        db.session.commit()
        flash("Number of error: " + str(error) + ", wrong tanks " + " ".join(str(x) for x in tanks))
        return redirect(url_for('main.index'))
    else:
        logs = UserLogs(user_id=current_user.id,
                        action="error_check_ended_successfully",
                        timestamp=datetime.now())
        db.session.add(logs)
        db.session.commit()
        return redirect(url_for('main.load'))


@bp.route('/prepare_tables_first', methods=['POST', 'GET'])
@login_required
def prepare_tables_first():
    def preparation():
        # функция создает таблицу всех возможных комбинаций налива топлива в бензовозы
        # для каждой азс и каждого бензовоза

        # очистка таблицы TempAzsTrucks в БД
        db.engine.execute("TRUNCATE TABLE `temp_azs_trucks`")
        # формируем массивы данных из БД
        azs_list = AzsList.query.filter_by(active=True).all()  # получаем список АКТИВНЫХ АЗС
        truck_list = Trucks.query.filter_by(active=True).all()  # получаем список АКТИВНЫХ бензовозов
        azs_tanks = Tanks.query.filter_by(active=True).all()  # получаем список АКТИЫНЫХ резервуаров всех АЗС
        truck_cells_list = TruckTanks.query.all()  # получаем список всех отсеков бензовозов
        # счетчик номера варианта налива
        variant_counter = 1
        # создаем список для записи в таблицу TempAzsTrucks
        temp_azs_truck_list = list()
        # создаем словарь для добавления в список temp_azs_trucks_list
        temp_azs_truck_dict = dict()
        for azs in azs_list:  # перебераем активные АЗС
            # создаем переменные для определения есть эти виды топлива на АЗС или нет
            is_92 = 0
            is_95 = 0
            is_50 = 0
            # перебираем таблицу из памяти со всеми АКТИЫНЫМИ резервуарами АЗС
            for row in azs_tanks:
                # проверяем есть ли у этой АЗС (которую перебираем в цикле) резервуары с 92 топливом
                if (row.azs_id == azs.id) and (row.fuel_type == 92):
                    is_92 = 1  # если есть, то помечаем соответствующий вид топлива
                # проверяем есть ли у этой АЗС резервуары с 95 топливом
                if (row.azs_id == azs.id) and (row.fuel_type == 95):
                    is_95 = 1  # если есть, то помечаем соответствующий вид топлива
                # проверяем есть ли у этой АЗС резервуары с 50 топливом
                if (row.azs_id == azs.id) and (row.fuel_type == 50):
                    is_50 = 1  # если есть, то помечаем соответствующий вид топлива

                # для ускорения проверяем, если все три вида топлив есть,
                # то можно цикл остановить, так как больше точно ничего не найдем
                if (is_92 == 1 and is_95 == 1 and is_50 == 1) and row.azs_id == azs.id:
                    break
            # В зависимости от найденных видов топлива на АЗС, формируем список
            azs_types = list()
            if (is_92 == 1) and (is_95 == 1) and (is_50 == 1):
                azs_types = [92, 95, 50]
            if (is_92 == 1) and (is_95 == 1) and (is_50 == 0):
                azs_types = [92, 95]
            if (is_92 == 1) and (is_95 == 0) and (is_50 == 1):
                azs_types = [92, 50]
            if (is_92 == 0) and (is_95 == 1) and (is_50 == 1):
                azs_types = [95, 50]
            if (is_92 == 1) and (is_95 == 0) and (is_50 == 0):
                azs_types = [92]
            if (is_92 == 0) and (is_95 == 1) and (is_50 == 0):
                azs_types = [95]
            if (is_92 == 0) and (is_95 == 0) and (is_50 == 1):
                azs_types = [50]

            # считаем количество отсеков в каждом из активных бензовозов
            # для этого:
            # перебираем список всех АКТИВНЫХ бензовозов
            for truck in truck_list:
                # создаем список для хранения видов топлива для заполнения данного бензовоза
                fuel_types = list()
                cell_counter = 0  # при смене бензовоза обнуляем счетчик отсеков
                for cell in truck_cells_list:  # перебираем все отсеки всех бензовозов
                    if cell.truck_id == truck.id:  # выбираем все отсеки у конкретного бензовоза
                        cell_counter = cell_counter + 1  # увеличиваем счетчик отсеков бензовоза
                # если у бензовоза ОДИН отсек
                if cell_counter == 1:
                    # тогда делаем 1 вложенный цикл, который перебирает все возможные виды топлива данной АЗС
                    for a in azs_types:
                        # добавляем в список один из всех возможных видов топлива данного бензовоза
                        fuel_types = [a]
                        # считаем количество отсеков с каждым видом топлива в данном варианте налива
                        cells_count_92 = 0
                        cells_count_95 = 0
                        cells_count_50 = 0
                        for i in fuel_types:
                            if i == 92:
                                cells_count_92 = cells_count_92 + 1
                            if i == 95:
                                cells_count_95 = cells_count_95 + 1
                            if i == 50:
                                cells_count_50 = cells_count_50 + 1

                        # перебираем все возможные виды топлива для данного бензовоза, с получением порядкового номера
                        # отсека и его емкости
                        for index, type in enumerate(fuel_types):
                            # перебираем список всего отсеков данного бензовоза
                            for cell in truck_cells_list:
                                # если нашли id нашего бензовоза и порядковый номер отсека совпадает с порядковым
                                # номером из списка fuel_types
                                if cell.truck_id == truck.id and cell.number == index + 1:
                                    # то формируем словарь с данными для записи в таблицу TempAzsTrucks в БД
                                    temp_azs_truck_dict = {'variant_id': variant_counter,
                                                           'azs_id': azs.id,
                                                           'truck_tank_id': cell.id,
                                                           'truck_id': truck.id,
                                                           'fuel_type': fuel_types[index],
                                                           'capacity': cell.capacity,
                                                           'cells_92': cells_count_92,
                                                           'cells_95': cells_count_95,
                                                           'cells_50': cells_count_50}
                                    # добавляем словарь в список temp_azs_truck_list, созданный ранее
                                    temp_azs_truck_list.append(temp_azs_truck_dict)
                                    # останавливаем итерацию цикла, так как искать больше нет смысла
                                    break
                        variant_counter = variant_counter + 1

                # если у бензовоза ДВА отсека
                if cell_counter == 2:
                    # тогда делаем 2 вложенных цикла которые переберают все возможные виды топлива данной азс
                    for a in azs_types:
                        for b in azs_types:
                            # добавляем в список два из всех возможных видов топлива данного бензовоза
                            fuel_types = [a, b]
                            # считаем количество отсеков с каждым видом топлива в данном варианте налива
                            cells_count_92 = 0
                            cells_count_95 = 0
                            cells_count_50 = 0
                            for i in fuel_types:
                                if i == 92:
                                    cells_count_92 = cells_count_92 + 1
                                if i == 95:
                                    cells_count_95 = cells_count_95 + 1
                                if i == 50:
                                    cells_count_50 = cells_count_50 + 1

                            # перебираем все возможные виды топлива для данного бензовоза,
                            # с получением порядкового номера отсека и его емкости
                            for index, type in enumerate(fuel_types):
                                # перебираем список всего отсеков данного бензовоза
                                for cell in truck_cells_list:
                                    # если нашли id нашего бензовоза и порядковый номер отсека совпадает с порядковым
                                    # номером из списка fuel_types
                                    if cell.truck_id == truck.id and cell.number == index + 1:
                                        # то формируем словарь с данными для записи в таблицу TempAzsTrucks в БД
                                        temp_azs_truck_dict = {'variant_id': variant_counter,
                                                               'azs_id': azs.id,
                                                               'truck_tank_id': cell.id,
                                                               'truck_id': truck.id,
                                                               'fuel_type': fuel_types[index],
                                                               'capacity': cell.capacity,
                                                               'cells_92': cells_count_92,
                                                               'cells_95': cells_count_95,
                                                               'cells_50': cells_count_50}
                                        # добавляем словарь в список temp_azs_truck_list, созданный ранее
                                        temp_azs_truck_list.append(temp_azs_truck_dict)
                                        # останавливаем итерацию цикла, так как искать больше нет смысла
                                        break
                # по аналогии для ТРЕХ отсеков бензовоза
                if cell_counter == 3:
                    for a in azs_types:
                        for b in azs_types:
                            for c in azs_types:
                                fuel_types = [a, b, c]
                                # считаем количество отсеков с каждым видом топлива в данном варианте налива
                                cells_count_92 = 0
                                cells_count_95 = 0
                                cells_count_50 = 0
                                for i in fuel_types:
                                    if i == 92:
                                        cells_count_92 = cells_count_92 + 1
                                    if i == 95:
                                        cells_count_95 = cells_count_95 + 1
                                    if i == 50:
                                        cells_count_50 = cells_count_50 + 1

                                # перебираем все возможные виды топлива для данного бензовоза,
                                # с получением порядкового номера отсека и его емкости
                                for index, type in enumerate(fuel_types):
                                    for cell in truck_cells_list:
                                        if cell.truck_id == truck.id and cell.number == index + 1:
                                            # то формируем словарь с данными для записи в таблицу TempAzsTrucks в БД
                                            temp_azs_truck_dict = {'variant_id': variant_counter,
                                                                   'azs_id': azs.id,
                                                                   'truck_tank_id': cell.id,
                                                                   'truck_id': truck.id,
                                                                   'fuel_type': fuel_types[index],
                                                                   'capacity': cell.capacity,
                                                                   'cells_92': cells_count_92,
                                                                   'cells_95': cells_count_95,
                                                                   'cells_50': cells_count_50}
                                            # добавляем словарь в список temp_azs_truck_list, созданный ранее
                                            temp_azs_truck_list.append(temp_azs_truck_dict)
                                            # останавливаем итерацию цикла, так как искать больше нет смысла
                                            break
                                variant_counter = variant_counter + 1
                # по аналогии для ЧЕТЫРЕХ отсеков бензовоза
                if cell_counter == 4:
                    for a in azs_types:
                        for b in azs_types:
                            for c in azs_types:
                                for d in azs_types:
                                    fuel_types = [a, b, c, d]
                                    cells_count_92 = 0
                                    cells_count_95 = 0
                                    cells_count_50 = 0
                                    for i in fuel_types:
                                        if i == 92:
                                            cells_count_92 = cells_count_92 + 1
                                        if i == 95:
                                            cells_count_95 = cells_count_95 + 1
                                        if i == 50:
                                            cells_count_50 = cells_count_50 + 1
                                    for index, type in enumerate(fuel_types):
                                        for cell in truck_cells_list:
                                            if cell.truck_id == truck.id and cell.number == index + 1:
                                                # то формируем словарь с данными для записи в таблицу TempAzsTrucks в БД
                                                temp_azs_truck_dict = {'variant_id': variant_counter,
                                                                       'azs_id': azs.id,
                                                                       'truck_tank_id': cell.id,
                                                                       'truck_id': truck.id,
                                                                       'fuel_type': fuel_types[index],
                                                                       'capacity': cell.capacity,
                                                                       'cells_92': cells_count_92,
                                                                       'cells_95': cells_count_95,
                                                                       'cells_50': cells_count_50}
                                                # добавляем словарь в список temp_azs_truck_list, созданный ранее
                                                temp_azs_truck_list.append(temp_azs_truck_dict)
                                                # останавливаем итерацию цикла, так как искать больше нет смысла
                                                break
                                    variant_counter = variant_counter + 1
                    # по аналогии для ПЯТИ отсеков бензовоза
                    if cell_counter == 5:
                        for a in azs_types:
                            for b in azs_types:
                                for c in azs_types:
                                    for d in azs_types:
                                        for e in azs_types:
                                            fuel_types = [a, b, c, d, e]
                                            # считаем количество отсеков с каждым видом топлива в данном варианте налива
                                            cells_count_92 = 0
                                            cells_count_95 = 0
                                            cells_count_50 = 0
                                            for i in fuel_types:
                                                if i == 92:
                                                    cells_count_92 = cells_count_92 + 1
                                                if i == 95:
                                                    cells_count_95 = cells_count_95 + 1
                                                if i == 50:
                                                    cells_count_50 = cells_count_50 + 1

                                            for index, type in enumerate(fuel_types):
                                                for cell in truck_cells_list:
                                                    if cell.truck_id == truck.id and cell.number == index + 1:
                                                        # то формируем словарь с данными для записи в таблицу
                                                        # TempAzsTrucks в БД
                                                        temp_azs_truck_dict = {'variant_id': variant_counter,
                                                                               'azs_id': azs.id,
                                                                               'truck_tank_id': cell.id,
                                                                               'truck_id': truck.id,
                                                                               'fuel_type': fuel_types[index],
                                                                               'capacity': cell.capacity,
                                                                               'cells_92': cells_count_92,
                                                                               'cells_95': cells_count_95,
                                                                               'cells_50': cells_count_50}
                                                        # добавляем словарь в список temp_azs_truck_list,
                                                        # созданный ранее
                                                        temp_azs_truck_list.append(temp_azs_truck_dict)
                                                        # останавливаем итерацию цикла, так как искать больше нет смысла
                                                        break

                                            variant_counter = variant_counter + 1
                    # по аналогии для ШЕСТИ отсеков бензовоза
                    if cell_counter == 6:
                        for a in azs_types:
                            for b in azs_types:
                                for c in azs_types:
                                    for d in azs_types:
                                        for e in azs_types:
                                            for f in azs_types:
                                                fuel_types = [a, b, c, d, e, f]
                                                # считаем количество отсеков с каждым видом топлива
                                                # в данном варианте налива
                                                cells_count_92 = 0
                                                cells_count_95 = 0
                                                cells_count_50 = 0
                                                for i in fuel_types:
                                                    if i == 92:
                                                        cells_count_92 = cells_count_92 + 1
                                                    if i == 95:
                                                        cells_count_95 = cells_count_95 + 1
                                                    if i == 50:
                                                        cells_count_50 = cells_count_50 + 1

                                                for index, type in enumerate(fuel_types):
                                                    for cell in truck_cells_list:
                                                        if cell.truck_id == truck.id and cell.number == index + 1:
                                                            # то формируем словарь с данными для записи в таблицу
                                                            # TempAzsTrucks в БД
                                                            temp_azs_truck_dict = {'variant_id': variant_counter,
                                                                                   'azs_id': azs.id,
                                                                                   'truck_tank_id': cell.id,
                                                                                   'truck_id': truck.id,
                                                                                   'fuel_type': fuel_types[index],
                                                                                   'capacity': cell.capacity,
                                                                                   'cells_92': cells_count_92,
                                                                                   'cells_95': cells_count_95,
                                                                                   'cells_50': cells_count_50}
                                                            # добавляем словарь в список temp_azs_truck_list,
                                                            # созданный ранее
                                                            temp_azs_truck_list.append(temp_azs_truck_dict)
                                                            # останавливаем итерацию цикла,
                                                            # так как искать больше нет смысла
                                                            break

                                            variant_counter = variant_counter + 1
                    # по аналогии для СЕМИ отсеков бензовоза
                    if cell_counter == 7:
                        for a in azs_types:
                            for b in azs_types:
                                for c in azs_types:
                                    for d in azs_types:
                                        for e in azs_types:
                                            for f in azs_types:
                                                for g in azs_types:
                                                    fuel_types = [a, b, c, d, e, f, g]

                                                    # считаем количество отсеков с каждым видом топлива
                                                    # в данном варианте налива
                                                    cells_count_92 = 0
                                                    cells_count_95 = 0
                                                    cells_count_50 = 0
                                                    for i in fuel_types:
                                                        if i == 92:
                                                            cells_count_92 = cells_count_92 + 1
                                                        if i == 95:
                                                            cells_count_95 = cells_count_95 + 1
                                                        if i == 50:
                                                            cells_count_50 = cells_count_50 + 1

                                                    for index, type in enumerate(fuel_types):
                                                        for cell in truck_cells_list:
                                                            if cell.truck_id == truck.id and cell.number == index + 1:
                                                                # то формируем словарь с данными для записи в таблицу
                                                                # TempAzsTrucks в БД
                                                                temp_azs_truck_dict = {'variant_id': variant_counter,
                                                                                       'azs_id': azs.id,
                                                                                       'truck_tank_id': cell.id,
                                                                                       'truck_id': truck.id,
                                                                                       'fuel_type': fuel_types[index],
                                                                                       'capacity': cell.capacity,
                                                                                       'cells_92': cells_count_92,
                                                                                       'cells_95': cells_count_95,
                                                                                       'cells_50': cells_count_50}
                                                                # добавляем словарь в список temp_azs_truck_list,
                                                                # созданный ранее
                                                                temp_azs_truck_list.append(temp_azs_truck_dict)
                                                                # останавливаем итерацию цикла, так как искать больше нет смысла
                                                                break

                                            variant_counter = variant_counter + 1
                    # по аналогии для ВОСЬМИ отсеков бензовоза
                    if cell_counter == 8:
                        for a in azs_types:
                            for b in azs_types:
                                for c in azs_types:
                                    for d in azs_types:
                                        for e in azs_types:
                                            for f in azs_types:
                                                for g in azs_types:
                                                    for h in azs_types:
                                                        fuel_types = [a, b, c, d, e, f, g, h]

                                                        # считаем количество отсеков с каждым видом топлива
                                                        # в данном варианте налива
                                                        cells_count_92 = 0
                                                        cells_count_95 = 0
                                                        cells_count_50 = 0
                                                        for i in fuel_types:
                                                            if i == 92:
                                                                cells_count_92 = cells_count_92 + 1
                                                            if i == 95:
                                                                cells_count_95 = cells_count_95 + 1
                                                            if i == 50:
                                                                cells_count_50 = cells_count_50 + 1

                                                        for index, type in enumerate(fuel_types):
                                                            for cell in truck_cells_list:
                                                                if cell.truck_id == truck.id and cell.number == index + 1:
                                                                    # то формируем словарь с данными для записи в таблицу
                                                                    # TempAzsTrucks в БД
                                                                    temp_azs_truck_dict = {
                                                                        'variant_id': variant_counter,
                                                                        'azs_id': azs.id,
                                                                        'truck_tank_id': cell.id,
                                                                        'truck_id': truck.id,
                                                                        'fuel_type': fuel_types[index],
                                                                        'capacity': cell.capacity,
                                                                        'cells_92': cells_count_92,
                                                                        'cells_95': cells_count_95,
                                                                        'cells_50': cells_count_50}
                                                                    # добавляем словарь в список temp_azs_truck_list,
                                                                    # созданный ранее
                                                                    temp_azs_truck_list.append(temp_azs_truck_dict)
                                                                    # останавливаем итерацию цикла,
                                                                    # так как искать больше нет смысла
                                                                    break

                                            variant_counter = variant_counter + 1
        # После выполнения функции записываем все полученные данные в таблицу TempAzsTrucks в базе данных
        return temp_azs_truck_list

    Close1_Tank1 = Close1Tank1.query.all()
    Close1_Tank2 = Close1Tank2.query.all()
    Close1_Tank3 = Close1Tank3.query.all()
    Close1_Tank4 = Close1Tank4.query.all()
    Close2_Tank1 = Close2Tank1.query.all()
    Close2_Tank2 = Close2Tank2.query.all()
    Close2_Tank3 = Close2Tank3.query.all()
    Close2_Tank4 = Close2Tank4.query.all()
    Close3_Tank1 = Close3Tank1.query.all()
    Close3_Tank2 = Close3Tank2.query.all()
    Close3_Tank3 = Close3Tank3.query.all()
    Close3_Tank4 = Close3Tank4.query.all()

    def select_close_tank_table(count, tanks_counter):
        if count == 1 and tanks_counter == 1:
            table_sliv_variant = Close1_Tank1
            return table_sliv_variant
        if count == 1 and tanks_counter == 2:
            table_sliv_variant = Close1_Tank2
            return table_sliv_variant
        if count == 1 and tanks_counter == 3:
            table_sliv_variant = Close1_Tank3
            return table_sliv_variant
        if count == 1 and tanks_counter == 4:
            table_sliv_variant = Close1_Tank4
            return table_sliv_variant

        if count == 2 and tanks_counter == 1:
            table_sliv_variant = Close2_Tank1
            return table_sliv_variant
        if count == 2 and tanks_counter == 2:
            table_sliv_variant = Close2_Tank2
            return table_sliv_variant
        if count == 2 and tanks_counter == 3:
            table_sliv_variant = Close2_Tank3
            return table_sliv_variant
        if count == 2 and tanks_counter == 4:
            table_sliv_variant = Close2_Tank4
            return table_sliv_variant

        if count == 3 and tanks_counter == 1:
            table_sliv_variant = Close3_Tank1
            return table_sliv_variant
        if count == 3 and tanks_counter == 2:
            table_sliv_variant = Close3_Tank2
            return table_sliv_variant
        if count == 3 and tanks_counter == 3:
            table_sliv_variant = Close3_Tank3
            return table_sliv_variant
        if count == 3 and tanks_counter == 4:
            table_sliv_variant = Close3_Tank4
            return table_sliv_variant

    # функция формирует все возмоные варианты слива топлива на АЗС
    def preparation_two():

        # очищаем таблицу для первого рейса
        db.engine.execute("TRUNCATE TABLE `temp_azs_trucks2`")
        # очищаем таблицу для второго рейса
        db.engine.execute("TRUNCATE TABLE `temp_azs_trucks2_second_trip`")
        # создаем массив для хранения сформированных итоговых данных
        # для последующей записи в таблицу TempAzsTrucks2 в БД
        temp_azs_trucks_2_list = list()
        temp_azs_trucks_2_dict = dict()
        # помещаем таблицу всех вариантов налива бензовозов TempAzsTrucks в переменную
        temp_azs_trucks = preparation()
        # помещаем таблицу всех АКТИВНЫХ АЗС в переменную
        table_azs_list = AzsList.query.filter_by(active=True).all()
        # помещаем таблицу все АКТИВНЫЕ резервуары АЗС в переменную
        table_tanks = Tanks.query.filter_by(active=True).all()

        # Получаем количество вариантов заполнения бензовоза (благодаря таблице TempAzsTrucks полю - variant_id)
        # счетчик варианта слива для таблицы TempAzsTrucks2
        variant_counter_sliv = 1
        # Для каждой АКТИВНОЙ АЗС считаем количество АКТИЫНЫХ резервуаров по каждому из видов топлива
        tanks_count = dict()
        # перебираем таблицу со списком АКТИВНЫХ АЗС

        for i in table_azs_list:

            # счетчик для хранения количества АКТИВНЫХ резервуаров данной азс
            tank_count_92 = 0
            tank_count_95 = 0
            tank_count_50 = 0
            # список с айдишниками АКТИВНЫХ резервуаров данной АЗС
            tanks_list_92 = list()
            tanks_list_95 = list()
            tanks_list_50 = list()
            # перебираем список АКТИВНЫХ резервуаров данной АЗС
            for tank in table_tanks:
                if i.id == tank.azs_id:
                    if tank.fuel_type == 92:
                        tank_count_92 = tank_count_92 + 1
                        tanks_list_92.append(tank.id)
                    if tank.fuel_type == 95:
                        tank_count_95 = tank_count_95 + 1
                        tanks_list_95.append(tank.id)
                    if tank.fuel_type == 50:
                        tank_count_50 = tank_count_50 + 1
                        tanks_list_50.append(tank.id)
                    tanks_count[i.id] = {
                        'tank_count_92': tank_count_92,
                        'tank_count_95': tank_count_95,
                        'tank_count_50': tank_count_50,
                        'tanks_list_92': tanks_list_92,
                        'tanks_list_95': tanks_list_95,
                        'tanks_list_50': tanks_list_50}
        # создаем словарь с обобщенными даннными для каждого варианта налива (основным ключем в словаре
        # является variant_id)
        slovar = dict()

        # создаем словарь с заполнеными ключами "variant_id, truck_id, cells_92, cells_95 и cells_50".
        # Остальные ячейки обнуляем
        for i in temp_azs_trucks:
            slovar[i['variant_id']] = {'azs_id': i['azs_id'],
                                       'cells_list_92': [],
                                       'cells_list_95': [],
                                       'cells_list_50': [],
                                       'capacity_92': 0,
                                       'capacity_95': 0,
                                       'capacity_50': 0,
                                       'cells_92': i['cells_92'],
                                       'cells_95': i['cells_95'],
                                       'cells_50': i['cells_50'],
                                       'truck_id': i['truck_id'],
                                       'cells_capacity_list_92': [],
                                       'cells_capacity_list_95': [],
                                       'cells_capacity_list_50': []
                                       }
        # перебираем таблицу TempAzsTrucks с вариантами налива и дополняем словарь
        for i in temp_azs_trucks:
            # обращаемся к ячейкам словаря по ключу variant_id
            temp_variant = i['variant_id']
            # если в текущей строке в таблице TempAzsTrucks вид топлива - 92,
            # то заполняем словарь согласно таблице для этого вида топлива
            if i['fuel_type'] == 92:
                # создаем пустой список для хранения айдишников отвеков бензовоза
                cells_list_92 = list()
                # создаем пустой список для хранения емкостей отсеков бензовоза
                cells_capacity_list_92 = list()
                # добавляем емкость отсека из текущей строки в список
                cells_capacity_list_92.append(i['capacity'])
                # добавляем айдишник отсека из текущей строки в список
                cells_list_92.append(i['truck_tank_id'])
                # обновляем данные в словаре
                slovar[temp_variant] = {'azs_id': i['azs_id'],
                                        'capacity_92': slovar[temp_variant]['capacity_92'] + i['capacity'],
                                        'capacity_95': slovar[temp_variant]['capacity_95'],
                                        'capacity_50': slovar[temp_variant]['capacity_50'],
                                        'cells_list_92': slovar[temp_variant]['cells_list_92'] + cells_list_92,
                                        'cells_list_95': slovar[temp_variant]['cells_list_95'],
                                        'cells_list_50': slovar[temp_variant]['cells_list_50'],
                                        'cells_92': slovar[temp_variant]['cells_92'],
                                        'cells_95': slovar[temp_variant]['cells_95'],
                                        'cells_50': slovar[temp_variant]['cells_50'],
                                        'truck_id': slovar[temp_variant]['truck_id'],
                                        'cells_capacity_list_92': slovar[temp_variant][
                                                                      'cells_capacity_list_92'] + cells_capacity_list_92,
                                        'cells_capacity_list_95': slovar[temp_variant]['cells_capacity_list_95'],
                                        'cells_capacity_list_50': slovar[temp_variant]['cells_capacity_list_50']
                                        }
            # выполняем действия для 95 вида топлива по аналогии с 92 видом топлива
            if i['fuel_type'] == 95:
                cells_list_95 = list()
                cells_capacity_list_95 = list()
                cells_capacity_list_95.append(i['capacity'])
                cells_list_95.append(i['truck_tank_id'])
                slovar[temp_variant] = {'azs_id': i['azs_id'],
                                        'capacity_92': slovar[temp_variant]['capacity_92'],
                                        'capacity_95': slovar[temp_variant]['capacity_95'] + i['capacity'],
                                        'capacity_50': slovar[temp_variant]['capacity_50'],
                                        'cells_list_92': slovar[temp_variant]['cells_list_92'],
                                        'cells_list_95': slovar[temp_variant]['cells_list_95'] + cells_list_95,
                                        'cells_list_50': slovar[temp_variant]['cells_list_50'],
                                        'cells_92': slovar[temp_variant]['cells_92'],
                                        'cells_95': slovar[temp_variant]['cells_95'],
                                        'cells_50': slovar[temp_variant]['cells_50'],
                                        'truck_id': slovar[temp_variant]['truck_id'],
                                        'cells_capacity_list_92': slovar[temp_variant]['cells_capacity_list_92'],
                                        'cells_capacity_list_95': slovar[temp_variant][
                                                                      'cells_capacity_list_95'] + cells_capacity_list_95,
                                        'cells_capacity_list_50': slovar[temp_variant]['cells_capacity_list_50']
                                        }

            # выполняем действия для дизеля вида топлива по аналогии с 92 и 95 видами топлива
            if i['fuel_type'] == 50:
                cells_list_50 = list()
                cells_capacity_list_50 = list()
                cells_capacity_list_50.append(i['capacity'])
                cells_list_50.append(i['truck_tank_id'])
                slovar[temp_variant] = {'azs_id': i['azs_id'],
                                        'capacity_92': slovar[temp_variant]['capacity_92'],
                                        'capacity_95': slovar[temp_variant]['capacity_95'],
                                        'capacity_50': slovar[temp_variant]['capacity_50'] + i['capacity'],
                                        'cells_list_92': slovar[temp_variant]['cells_list_92'],
                                        'cells_list_95': slovar[temp_variant]['cells_list_95'],
                                        'cells_list_50': slovar[temp_variant]['cells_list_50'] + cells_list_50,
                                        'cells_92': slovar[temp_variant]['cells_92'],
                                        'cells_95': slovar[temp_variant]['cells_95'],
                                        'cells_50': slovar[temp_variant]['cells_50'],
                                        'truck_id': slovar[temp_variant]['truck_id'],
                                        'cells_capacity_list_92': slovar[temp_variant]['cells_capacity_list_92'],
                                        'cells_capacity_list_95': slovar[temp_variant]['cells_capacity_list_95'],
                                        'cells_capacity_list_50': slovar[temp_variant][
                                                                      'cells_capacity_list_50'] + cells_capacity_list_50
                                        }

        # перебираем варианты налива от первого до последнего
        for variant in range(1, temp_azs_trucks[-1]['variant_id']):
            # получаем айдишник текущей АЗС
            azs_id = slovar[variant]['azs_id']
            # Получаем количество резервуаров АЗС по каждому виду топлива
            count_92 = tanks_count[azs_id]['tank_count_92']
            count_95 = tanks_count[azs_id]['tank_count_95']
            count_50 = tanks_count[azs_id]['tank_count_50']
            # создаем переменные для хранения таблиц необходимых для перебора всех возможных вариантов слива
            table_sliv_variant_92 = None
            table_sliv_variant_50 = None
            table_sliv_variant_95 = None

            # Благодаря тому, что мы знаем количество резервуаров АЗС и количество отсеков бензовоза с этим видом
            # топлива получаем нужную константную таблицу с вариантами слива бензовоза
            # для 92 вида топлива
            if count_92 == 1 and slovar[variant]['cells_92'] == 1:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 1 and slovar[variant]['cells_92'] == 2:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 1 and slovar[variant]['cells_92'] == 3:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 1 and slovar[variant]['cells_92'] == 4:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 2 and slovar[variant]['cells_92'] == 1:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 2 and slovar[variant]['cells_92'] == 2:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 2 and slovar[variant]['cells_92'] == 3:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 2 and slovar[variant]['cells_92'] == 4:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 3 and slovar[variant]['cells_92'] == 1:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 3 and slovar[variant]['cells_92'] == 2:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 3 and slovar[variant]['cells_92'] == 3:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            if count_92 == 3 and slovar[variant]['cells_92'] == 4:
                table_sliv_variant_92 = select_close_tank_table(count_92, slovar[variant]['cells_92'])

            # для 95 вида топлива
            if count_95 == 1 and slovar[variant]['cells_95'] == 1:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 1 and slovar[variant]['cells_95'] == 2:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 1 and slovar[variant]['cells_95'] == 3:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 1 and slovar[variant]['cells_95'] == 4:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 2 and slovar[variant]['cells_95'] == 1:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 2 and slovar[variant]['cells_95'] == 2:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 2 and slovar[variant]['cells_95'] == 3:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 2 and slovar[variant]['cells_95'] == 4:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 3 and slovar[variant]['cells_95'] == 1:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 3 and slovar[variant]['cells_95'] == 2:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 3 and slovar[variant]['cells_95'] == 3:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            if count_95 == 3 and slovar[variant]['cells_95'] == 4:
                table_sliv_variant_95 = select_close_tank_table(count_95, slovar[variant]['cells_95'])

            # # для 50 вида топлива
            if count_50 == 1 and slovar[variant]['cells_50'] == 1:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 1 and slovar[variant]['cells_50'] == 2:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 1 and slovar[variant]['cells_50'] == 3:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 1 and slovar[variant]['cells_50'] == 4:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 2 and slovar[variant]['cells_50'] == 1:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 2 and slovar[variant]['cells_50'] == 2:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 2 and slovar[variant]['cells_50'] == 3:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 2 and slovar[variant]['cells_50'] == 4:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 3 and slovar[variant]['cells_50'] == 1:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 3 and slovar[variant]['cells_50'] == 2:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 3 and slovar[variant]['cells_50'] == 3:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])

            if count_50 == 3 and slovar[variant]['cells_50'] == 4:
                table_sliv_variant_50 = select_close_tank_table(count_50, slovar[variant]['cells_50'])
            # если у данной АЗС есть 92 топливо, то перебираем таблицу, возвращенную функцией select_close_tank_table
            if table_sliv_variant_92 is not None:
                # если резервуар с таким видом топлива один, то
                if count_92 == 1:
                    # перебираем таблицу
                    for variant_sliv in table_sliv_variant_92:
                        # если первая ячейка таблицы не пуста или не равна NULL, то
                        if variant_sliv.tank1 is not None:
                            # обнуляем счетчик суммарного слива
                            sum_sliv = 0
                            # создаем список, в котором будем хранить айди отсеков бензовоза в виде строковых параметров
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            # перебираем варианты слива, записанные в виде строки, разделенные знаком "+"
                            # из таблицы table_sliv_variant_92
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                # по индексу получаем емкость каждого отсека, суммируем емкости
                                # и получаем суммарный слив
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                # по индексу получаем айди отсеков бензовоза, и формирвем из них строку,
                                # где айди разделены знаком "+"
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            # формируем целую строку
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            # заполняем словарь полученными даннми
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            # добавляем полученынй словарь в списокдля последующей записи в базу
                            # в таблицу TempAzsTrucks2
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        # увеличиваем счетчик варианта слива на единицу
                        variant_counter_sliv = variant_counter_sliv + 1
                # для двух резервуаров выполняем те же действия как с одним
                if count_92 == 2:
                    for variant_sliv in table_sliv_variant_92:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()

                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)

                        if variant_sliv.tank2 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()

                            for index, number in enumerate(variant_sliv.tank2.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank2,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][1],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)

                        variant_counter_sliv = variant_counter_sliv + 1
                # для трех резервуаров, по аналогии с одним
                if count_92 == 3:
                    for variant_sliv in table_sliv_variant_92:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)

                        if variant_sliv.tank2 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank2.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank2,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][1],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)

                        if variant_sliv.tank3 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank3.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank3,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][2],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)

                        variant_counter_sliv = variant_counter_sliv + 1
                # по аналогии с 4 резервуарами
                if count_92 == 4:
                    for variant_sliv in table_sliv_variant_92:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)

                        if variant_sliv.tank2 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank2.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank2,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][1],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)

                        if variant_sliv.tank3 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank3.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank3,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][2],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)

                        if variant_sliv.tank4 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank4.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_92'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_92'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 92,
                                                      'str_sliv': variant_sliv.tank4,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_92'][3],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)

                        variant_counter_sliv = variant_counter_sliv + 1

            # --- Для 95 вида топлива выполняем по аналогии с 92 видом топлива
            if table_sliv_variant_95 is not None:
                if count_95 == 1:
                    for variant_sliv in table_sliv_variant_95:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        variant_counter_sliv = variant_counter_sliv + 1
                if count_95 == 2:
                    for variant_sliv in table_sliv_variant_95:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank2 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank2.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank2,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][1],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        variant_counter_sliv = variant_counter_sliv + 1
                if count_95 == 3:
                    for variant_sliv in table_sliv_variant_95:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank2 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank2.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank2,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][1],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank3 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank3.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank3,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][2],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        variant_counter_sliv = variant_counter_sliv + 1
                if count_95 == 4:
                    for variant_sliv in table_sliv_variant_95:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank2 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank2.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank2,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][1],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank3 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank3.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank3,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][2],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank4 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank4.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_95'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_95'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 95,
                                                      'str_sliv': variant_sliv.tank4,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_95'][3],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        variant_counter_sliv = variant_counter_sliv + 1
            # --- Для 50 вида топлива выполняем по аналогии с 92 и 95 видом топлива
            if table_sliv_variant_50 is not None:
                if count_50 == 1:
                    for variant_sliv in table_sliv_variant_50:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        variant_counter_sliv = variant_counter_sliv + 1
                if count_50 == 2:
                    for variant_sliv in table_sliv_variant_50:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank2 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank2.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank2,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][1],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        variant_counter_sliv = variant_counter_sliv + 1
                if count_50 == 3:
                    for variant_sliv in table_sliv_variant_50:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank2 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank2.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank2,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][1],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank3 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""

                            for index, number in enumerate(variant_sliv.tank3.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank3,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][2],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        variant_counter_sliv = variant_counter_sliv + 1
                if count_50 == 4:
                    for variant_sliv in table_sliv_variant_50:
                        if variant_sliv.tank1 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank1.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank1,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][0],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank2 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank2.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank2,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][1],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank3 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank3.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank3,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][2],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        if variant_sliv.tank4 is not None:
                            sum_sliv = 0
                            str_sliv_cells_list = list()
                            str_sliv_cells = ""
                            for index, number in enumerate(variant_sliv.tank4.split('+')):
                                sum_sliv = sum_sliv + slovar[variant]['cells_capacity_list_50'][int(number) - 1]
                                str_sliv_cells_list.append(str(slovar[variant]['cells_list_50'][int(number) - 1]))
                            str_sliv_cells = '+'.join(str_sliv_cells_list)
                            temp_azs_trucks_2_dict = {'variant': variant,
                                                      'azs_id': azs_id,
                                                      'truck_id': slovar[variant]['truck_id'],
                                                      'variant_sliv': variant_counter_sliv,
                                                      'fuel_type': 50,
                                                      'str_sliv': variant_sliv.tank4,
                                                      'tank_id': tanks_count[azs_id]['tanks_list_50'][3],
                                                      'truck_tank_id_string': str_sliv_cells,
                                                      'sum_sliv': sum_sliv
                                                      }
                            temp_azs_trucks_2_list.append(temp_azs_trucks_2_dict)
                        variant_counter_sliv = variant_counter_sliv + 1
        db.engine.execute(TempAzsTrucks.__table__.insert(), temp_azs_trucks)

        global global_temp_azs_trucks_2_list
        global global_temp_azs_trucks
        global_temp_azs_trucks = temp_azs_trucks
        global_temp_azs_trucks_2_list = temp_azs_trucks_2_list
        return temp_azs_trucks_2_list, temp_azs_trucks

    ''' 
    # функция определяет может ли вариант слива в данный момент слиться на АЗС,
    #     # сможет ли вариант слива слиться после времени бензовоза в пути
    #     # определяет новый запас суток
    #     # определяет остатки топлива после слива
    #     # определяет, может ли бензовой зайти на АЗС
    # может ли этот варинат налива пройти по дороге с весами
    '''
    "------------------------------------------------------------------------------------------------------------"
    "---------------------------- определяем, сможет ли бензовоз пройти весы ------------------------------------"
    "------------------------------------------------------------------------------------------------------------"

    def is_it_fit():
        temp_azs_trucks_2_list, temp_azs_trucks = preparation_two()
        # получаем остатки из базы
        residue = FuelResidue.query.all()
        # получаем реализацию из базы
        realisation = FuelRealisation.query.all()
        # получаем информацию о времени до азс
        trip = Trip.query.all()
        # получаем список отсеков всех бензовозов
        truck_tanks = TruckTanks.query.all()
        # получаем информанию об азс, на которые не могут заехать определенные бензовозы
        trucks_false = TruckFalse.query.all()
        # создаем пустой словарь для хранеия в нем данных о реализации и остатков
        realisation_n_residue = dict()
        # создаем пустой словарь для хранеия в нем данных времени пути до АЗС
        azs_trip_access = dict()
        azs_trip_time = dict()

        # заполняем словарь с данными об остатках топлива
        for i in residue:
            realisation_n_residue[i.tank_id] = {'azs_id': i.azs_id,  # id_АЗС
                                                'fuel_volume': i.fuel_volume,  # остаток топлива в резервуаре
                                                'free_volume': i.free_volume,  # свободная емкость в резервуаре
                                                # реализация в резервуаре (нулевая, так как заполняем ее потом)
                                                'fuel_realisation': 0,
                                                # максимальная реализация топлива в резервуаре (среди всех периодов)
                                                'fuel_realisation_max': 0
                                                }
        # заполняем словарь данными о реализации топлива
        for i in realisation:
            realisation_n_residue[i.tank_id] = {'azs_id': i.azs_id,  # id_АЗС
                                                # Оставляем прошлое значение остатка топлива в резервуаре
                                                'fuel_volume': realisation_n_residue[i.tank_id]['fuel_volume'],
                                                # свободная емкость в резервуаре
                                                'free_volume': realisation_n_residue[i.tank_id]['free_volume'],
                                                # Добавляем в словарь реализацию из этого резервуара
                                                # (среднюю в час за шестичасовой период)
                                                'fuel_realisation': i.fuel_realisation_hour / 6,
                                                # максимальная реализация топлива в резервуаре (среди всех периодов)
                                                'fuel_realisation_max': i.fuel_realisation_max
                                                }

        # формируем словарь для хранения данных о растоянии до АЗС и времени пути
        for i in trip:
            azs_trip_time[i.azs_id] = {'time_to_before_lunch': i.time_to_before_lunch,
                                       'time_to': i.time_to,
                                       'weigher': i.weigher}
        # формируем словарь для хранения данных о бензовозах и азс на которые они не могут заезжать
        for i in trucks_false:
            azs_trip_access[str(i.azs_id) + '-' + str(i.truck_id)] = {'access': False}

        is_it_fit_list = list()
        for i in temp_azs_trucks_2_list:
            # проверяем, сольется ли бензовоз в данный момент
            # из свободной емкости резервуара вычитаем сумму слива бензовоза
            sliv = realisation_n_residue[i['tank_id']]['free_volume'] - i['sum_sliv']
            # получаем время до азс до обеда (в минутах)
            time_to = azs_trip_time[i['azs_id']]['time_to_before_lunch']
            # считаем примерное количество топлива, которое будет реализовано за время в пути бензовоза
            realis_time = realisation_n_residue[i["tank_id"]]['fuel_realisation'] * (time_to / 60)
            # проверяем сольется ли бензовоз, с учетом реализации за время его пути к АЗС
            # из свободной емкости резервуара вычитаем сумму слива бензовоза, и прибавляем количество топлива,
            # которое реализуется у данного резервуара за время пути бензовоза к ней
            sliv_later = realisation_n_residue[i['tank_id']]['free_volume'] - i['sum_sliv'] + realis_time

            # если бензовоз не сливается в данный момент (то есть переменная sliv - меньше нуля)
            if sliv < 0:
                # записываем в базу, что бензовоз в данный момент слиться не сможет
                i['is_it_fit'] = False
                # новый запас суток и новые остатки не считаем
                i['new_fuel_volume'] = 0
                i['new_days_stock'] = 0
            # если бензовоз сможет слиться нат екущий момент (то есть переменная sliv - больше нуля)
            else:
                # записываем в базу, что бензовоз в данный момент сольется
                i['is_it_fit'] = True
                # расчитываем количество отстатков в резервуаре после слива
                i['new_fuel_volume'] = realisation_n_residue[i['tank_id']]['fuel_volume'] + i['sum_sliv']
                # расчитываем новый запас суток
                i['new_days_stock'] = i['new_fuel_volume'] / realisation_n_residue[i['tank_id']]['fuel_realisation_max']
            # если бензовоз не сливается после времени затраченного на дорогу (то есть переменная sliv_later
            # - меньше нуля)
            if sliv_later < 0:
                # записываем в базу, что бензовоз слиться не сможет
                i['is_it_fit_later'] = False
                # новый запас суток и новые остатки не считаем
                i['new_fuel_volume'] = 0
                i['new_days_stock'] = 0
            else:
                # записываем в базу, что бензовоз сольется спустя время затраченное на дорогу
                i['is_it_fit_later'] = True
                # расчитываем количество отстатков в резервуаре после слива
                i['new_fuel_volume'] = realisation_n_residue[i['tank_id']]['fuel_volume'] + i['sum_sliv']
                # расчитываем новый запас суток
                i['new_days_stock'] = i['new_fuel_volume'] / realisation_n_residue[i['tank_id']]['fuel_realisation_max']
            # проверяем, сможет ли бензовоз заехать на АЗС (нет ли для него никаких ограничений)
            # т.е. проверяем наличие ключа в словаре azs_trip_access, в котором содержатся ограничения для бензовозов
            # ключ АЗС-АйдиБензовоза
            if str(i['azs_id']) + '-' + str(i['truck_id']) in azs_trip_access:
                # если ограничения есть, то ставим False
                i['is_it_able_to_enter'] = False
            else:
                # если ограничений нет, то ставим True
                i['is_it_able_to_enter'] = True

            # добавляем словарь в список для записи в базу данных, в таблицу TempAzsTrucks2
            is_it_fit_list.append(i)
        # создаем словарь для хранения переменных с информацией о том, влезет ли определенный вид топлива
        # в резервуар на АЗС (3 переменные по всем видам топлива)
        fuel_types_dict = dict()
        # создаем список  для хранеия обновленной таблицы TempAzsTrucks2
        is_variant_sliv_good_list = list()
        # заполняем переменные словаря fuel_types_dict единицами
        # ключем в словаре является связка variant:variant_sliv
        for i in is_it_fit_list:
            fuel_types_dict[str(i['variant']) + ':' + str(i['variant_sliv'])] = {'is_it_fit_92': 1,
                                                                                 'is_it_fit_95': 1,
                                                                                 'is_it_fit_50': 1}
        # заново перебираем таблицу TempAzsTrucks2(которая хранится в виде списка словарей)
        for i in is_it_fit_list:
            # в переменную key заносим ключ итого словаря (связка variant:variant_sliv)
            key = str(i['variant']) + ':' + str(i['variant_sliv'])
            # если находим вид топлива которое не сливается, то помечаем его нулем
            if i['fuel_type'] == 92 and i['is_it_fit_later'] == 0:
                fuel_types_dict[key] = {'is_it_fit_92': 0,
                                        'is_it_fit_95': fuel_types_dict[key]['is_it_fit_95'],
                                        'is_it_fit_50': fuel_types_dict[key]['is_it_fit_50']}
            if i['fuel_type'] == 95 and i['is_it_fit_later'] == 0:
                fuel_types_dict[key] = {'is_it_fit_92': fuel_types_dict[key]['is_it_fit_92'],
                                        'is_it_fit_95': 0,
                                        'is_it_fit_50': fuel_types_dict[key]['is_it_fit_50']}
            if i['fuel_type'] == 50 and i['is_it_fit_later'] == 0:
                fuel_types_dict[key] = {'is_it_fit_92': fuel_types_dict[key]['is_it_fit_92'],
                                        'is_it_fit_95': fuel_types_dict[key]['is_it_fit_95'],
                                        'is_it_fit_50': 0}
        # снова беребираем список словарей с данными из таблицы TempAzsTrucks2
        for i in is_it_fit_list:
            # в переменную key заносим ключ итого словаря (связка variant:variant_sliv)
            key = str(i['variant']) + ':' + str(i['variant_sliv'])
            # если все три вида топлива (или все виды топлива которые мы везем на азс) сливаются, то помечаем столбец
            # в котором хранится информация о том, сливается ли данный вариант (is_variant_sliv_good) единицей
            if fuel_types_dict[key]['is_it_fit_92'] == 1 and fuel_types_dict[key]['is_it_fit_95'] == 1 and \
                    fuel_types_dict[key]['is_it_fit_50'] == 1:
                i['is_variant_sliv_good'] = 1
                # добавляем обновленный словарь в список
                is_variant_sliv_good_list.append(i)
            # если не все топливо из данного варианта слива сливается, то ставим в ячейке ноль (False)
            else:
                i['is_variant_sliv_good'] = 0
                # добавляем обновленный словарь в список
                is_variant_sliv_good_list.append(i)
        # создаем словарь для хранения обновленных данных (по факту - в словаре появится ячейка is_variant_good)
        is_variant_good_list = dict()
        # перебираем список словарей с данными из таблицы TempAzsTrucks2 (из предыдущего цикла)
        for i in is_variant_sliv_good_list:
            # в созданные ранее словарь добавляем ячейки с нулями
            is_variant_good_list[str(i['variant'])] = {'is_it_92': 0,
                                                       'is_it_95': 0,
                                                       'is_it_50': 0}
        # снова перебираем список словарей
        for i in is_variant_sliv_good_list:
            # если находим определенный вид топлива, и вариант слива относящийся к этой строке таблицы отмечен True,
            # то помечаем го цифрой 2
            # а если нет, то единицей
            if i['fuel_type'] == 92 and i['is_variant_sliv_good'] == 1:
                is_variant_good_list[str(i['variant'])] = {'is_it_92': 2,
                                                           'is_it_95': is_variant_good_list[str(i['variant'])][
                                                               'is_it_95'],
                                                           'is_it_50': is_variant_good_list[str(i['variant'])][
                                                               'is_it_50']}
            if i['fuel_type'] == 92 and i['is_variant_sliv_good'] == 0 and is_variant_good_list[str(i['variant'])][
                'is_it_92'] != 2:
                is_variant_good_list[str(i['variant'])] = {'is_it_92': 1,
                                                           'is_it_95': is_variant_good_list[str(i['variant'])][
                                                               'is_it_95'],
                                                           'is_it_50': is_variant_good_list[str(i['variant'])][
                                                               'is_it_50']}
            if i['fuel_type'] == 95 and i['is_variant_sliv_good'] == 1:
                is_variant_good_list[str(i['variant'])] = {
                    'is_it_92': is_variant_good_list[str(i['variant'])]['is_it_92'],
                    'is_it_95': 2,
                    'is_it_50': is_variant_good_list[str(i['variant'])]['is_it_50']}
            if i['fuel_type'] == 95 and i['is_variant_sliv_good'] == 0 and is_variant_good_list[str(i['variant'])][
                'is_it_95'] != 2:
                is_variant_good_list[str(i['variant'])] = {
                    'is_it_92': is_variant_good_list[str(i['variant'])]['is_it_92'],
                    'is_it_95': 1,
                    'is_it_50': is_variant_good_list[str(i['variant'])]['is_it_50']}
            if i['fuel_type'] == 50 and i['is_variant_sliv_good'] == 1:
                is_variant_good_list[str(i['variant'])] = {
                    'is_it_92': is_variant_good_list[str(i['variant'])]['is_it_92'],
                    'is_it_95': is_variant_good_list[str(i['variant'])]['is_it_95'],
                    'is_it_50': 2}
            if i['fuel_type'] == 50 and i['is_variant_sliv_good'] == 0 and is_variant_good_list[str(i['variant'])][
                'is_it_50'] != 2:
                is_variant_good_list[str(i['variant'])] = {
                    'is_it_92': is_variant_good_list[str(i['variant'])]['is_it_92'],
                    'is_it_95': is_variant_good_list[str(i['variant'])]['is_it_95'],
                    'is_it_50': 1}

        # создаем финальный список для данной функции, который будет записан в таблицу TempAzsTrucks2 в БД
        final_list2 = list()
        final_list = list()
        test_dict = dict()
        weigher_dict = dict()  # словарь с наполнением при весах
        weigher_variant_good_dict = dict()
        trip = Trip.query.filter_by(weigher=1).all()
        cells = TruckTanks.query.all()
        truck_tanks_variations = TruckTanksVariations.query.all()
        azs_list_weigher_variant_id = list()
        azs_list_weigher_truck_id = list()
        azs_list_weigher_fuel_type = list()
        azs_list_weigher_truck_tank_id = list()
        for i in trip:
            for azs in temp_azs_trucks:
                if azs['azs_id'] == i.azs_id:
                    azs_list_weigher_variant_id.append(azs['variant_id'])
                    azs_list_weigher_truck_id.append(azs['truck_id'])
            for index, x in enumerate(azs_list_weigher_variant_id):
                key = str(azs_list_weigher_variant_id[index]) + ':' + str(azs_list_weigher_truck_id[index])
                test_dict[key] = {'1': None,
                                  '2': None,
                                  '3': None,
                                  '4': None,
                                  '5': None,
                                  '6': None,
                                  '7': None,
                                  '8': None
                                  }

        azs_list_weigher_variant_id = list()
        azs_list_weigher_truck_id = list()
        azs_list_weigher_fuel_type = list()
        azs_list_weigher_truck_tank_id = list()
        for i in trip:
            for azs in temp_azs_trucks:
                if azs['azs_id'] == i.azs_id:
                    azs_list_weigher_variant_id.append(azs['variant_id'])
                    azs_list_weigher_truck_id.append(azs['truck_id'])
                    azs_list_weigher_fuel_type.append(azs['fuel_type'])
                    azs_list_weigher_truck_tank_id.append(azs['truck_tank_id'])
            for index, x in enumerate(azs_list_weigher_variant_id):
                key = str(azs_list_weigher_variant_id[index]) + ':' + str(azs_list_weigher_truck_id[index])

                fuel_type_boolean = azs_list_weigher_fuel_type[index]
                for cell in cells:
                    if cell.id == azs_list_weigher_truck_tank_id[index]:
                        if cell.number == 1:
                            test_dict[key] = {'1': fuel_type_boolean,
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 2:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': fuel_type_boolean,
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 3:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': fuel_type_boolean,
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 4:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': fuel_type_boolean,
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 5:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': fuel_type_boolean,
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 6:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': fuel_type_boolean,
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 7:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['7'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': fuel_type_boolean,
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 8:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': fuel_type_boolean
                                              }

        for cell in truck_tanks_variations:
            weigher_dict[str(cell.truck_id) + ":" + str(cell.variant_good)] = {'1': None,
                                                                               '2': None,
                                                                               '3': None,
                                                                               '4': None,
                                                                               '5': None,
                                                                               '6': None,
                                                                               '7': None,
                                                                               '8': None
                                                                               }

            weigher_variant_good_dict[cell.truck_id] = {'variant_good': []}

        for cell in truck_tanks_variations:
            if cell.variant_good not in weigher_variant_good_dict[cell.truck_id]['variant_good']:
                temp_list = list()
                temp_list.append(cell.variant_good)
                weigher_variant_good_dict[cell.truck_id] = {'variant_good':
                                                                weigher_variant_good_dict[cell.truck_id][
                                                                    'variant_good'] + temp_list
                                                            }

            for truck_cell in cells:
                if truck_cell.id == cell.truck_tank_id:
                    number = truck_cell.number

            key = str(cell.truck_id) + ':' + str(cell.variant_good)
            if number == 1:
                weigher_dict[key] = {'1': cell.diesel,
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 2:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': cell.diesel,
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 3:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': cell.diesel,
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 4:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': cell.diesel,
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 5:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': cell.diesel,
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 6:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': cell.diesel,
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 7:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': cell.diesel,
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 8:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': cell.diesel
                                     }

        # перебираем список словарей
        for i in is_variant_sliv_good_list:
            # если все виды топлива данного варианта сливаются, то ячейку is_variant_good в таблице записываем True,
            # если же нет, то в ячейку is_variant_good пишем False
            if is_variant_good_list[str(i['variant'])]['is_it_92'] != 1 \
                    and is_variant_good_list[str(i['variant'])]['is_it_95'] != 1 \
                    and is_variant_good_list[str(i['variant'])]['is_it_50'] != 1:
                i['is_variant_good'] = True
                # добавляем получившийся словарь в список
                final_list.append(i)
            else:
                i['is_variant_good'] = False
                # добавляем получившийся словарь в список
                final_list.append(i)

        is_variant_weighter_good = list()
        is_variant_weighter_not_good = list()
        for i in final_list:
            key = str(i['variant']) + ':' + str(i['truck_id'])

            if key in test_dict:
                trig_final = 0  # Изначально считаем, что бензовоз нельзя везти если есть весы
                cells_list = list()

                cell_1 = test_dict[key]['1']
                cells_list.append(cell_1)
                cell_2 = test_dict[key]['2']
                cells_list.append(cell_2)
                cell_3 = test_dict[key]['3']
                cells_list.append(cell_3)
                cell_4 = test_dict[key]['4']
                cells_list.append(cell_4)
                cell_5 = test_dict[key]['5']
                cells_list.append(cell_5)
                cell_6 = test_dict[key]['6']
                cells_list.append(cell_6)
                cell_7 = test_dict[key]['7']
                cells_list.append(cell_7)
                cell_8 = test_dict[key]['8']
                cells_list.append(cell_8)

                for variant_good in weigher_variant_good_dict[i['truck_id']]['variant_good']:
                    variant_key = str(i['truck_id']) + ":" + str(variant_good)
                    weigher_list = list()

                    weigher_list.append(weigher_dict[variant_key]['1'])
                    weigher_list.append(weigher_dict[variant_key]['2'])
                    weigher_list.append(weigher_dict[variant_key]['3'])
                    weigher_list.append(weigher_dict[variant_key]['4'])
                    weigher_list.append(weigher_dict[variant_key]['5'])
                    weigher_list.append(weigher_dict[variant_key]['6'])
                    weigher_list.append(weigher_dict[variant_key]['7'])
                    weigher_list.append(weigher_dict[variant_key]['8'])

                    trig = 1  # Можно завозить дизель для этой комбинации налива
                    for index, cell in enumerate(cells_list):
                        if cells_list[index] == 50 and (weigher_list[index] == 0 or weigher_list[index] == None):
                            trig = 0  # Нельзя завозить дизель для этой комбинации налива
                            break

                    if trig == 1:  # Если можнно завозить дизель для одной их всевозможных комбинаций налива,
                        trig_final = 1  # то значит  можно завозить

                if trig_final == 1:  # Значит вариант подходит (дизель можно везти)!

                    if int(i['variant']) == 412:
                        print("zbs -----------------------------------------------")
                    is_variant_weighter_good.append(i['variant'])
                else:
                    is_variant_weighter_not_good.append(i['variant'])

        for i in final_list:
            if i['variant'] in is_variant_weighter_not_good:
                i['is_variant_weigher_good'] = False
            else:
                i['is_variant_weigher_good'] = True
            final_list2.append(i)
        # записываем данные из списка в базу
        db.engine.execute(TempAzsTrucks2.__table__.insert(), final_list2)
        return final_list2

    ''' функция отсеивает все варианты из таблицы TempAzsTrucks2 и дает им оценку'''

    def preparation_four():
        final_list = is_it_fit()
        # очищаем таблицу
        db.engine.execute("TRUNCATE TABLE `temp_azs_trucks3`")
        db.engine.execute("TRUNCATE TABLE `temp_azs_trucks4`")
        temp_azs_trucks3_list = list()
        fuel_realisation = FuelRealisation.query.all()
        days_stock_dict = dict()

        # перебираем список из предыдущей функции
        for i in final_list:
            # если вариант сливается, бензовоз может заехать на АЗС и бензовоз сливается полностью,
            # и бензовоз проходит через весы
            if i['is_it_fit_later'] == True and i['is_it_able_to_enter'] == True and i['is_variant_good'] == True \
                    and i['is_variant_sliv_good'] == True and i['is_variant_weigher_good'] == True:
                # добавляем словарь в список
                temp_azs_trucks3_list.append(i)
        new_days_stock_dict = dict()
        temp_azs_trucks4_dict = dict()

        for i in temp_azs_trucks3_list:
            temp_azs_trucks4_dict[str(i['variant'])] = {'variant_sliv_92': [],
                                                        'variant_sliv_95': [],
                                                        'variant_sliv_50': [],
                                                        'volume_92': [],
                                                        'volume_95': [],
                                                        'volume_50': [],
                                                        'azs_id': i['azs_id'],
                                                        'truck_id': i['truck_id']
                                                        }
        for i in temp_azs_trucks3_list:
            variant_sliv_92 = list()
            variant_sliv_95 = list()
            variant_sliv_50 = list()
            volume_92 = list()
            volume_95 = list()
            volume_50 = list()

            if i['fuel_type'] == 92:
                variant_sliv_92.append(i['variant_sliv'])
                volume_92.append(i['sum_sliv'])
                if i['variant_sliv'] in temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_92']:
                    index = temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_92'].index(i['variant_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_92'].insert(index, temp_azs_trucks4_dict[
                        str(i['variant'])]['volume_92'][index] + i['sum_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_92'].pop(index + 1)
                else:
                    temp_azs_trucks4_dict[str(i['variant'])] = {
                        'variant_sliv_92': temp_azs_trucks4_dict[str(i['variant'])][
                                               'variant_sliv_92'] + variant_sliv_92,
                        'variant_sliv_95': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_95'],
                        'variant_sliv_50': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_50'],
                        'volume_92': temp_azs_trucks4_dict[str(i['variant'])]['volume_92'] + volume_92,
                        'volume_95': temp_azs_trucks4_dict[str(i['variant'])]['volume_95'],
                        'volume_50': temp_azs_trucks4_dict[str(i['variant'])]['volume_50'],
                        'azs_id': temp_azs_trucks4_dict[str(i['variant'])]['azs_id'],
                        'truck_id': temp_azs_trucks4_dict[str(i['variant'])]['truck_id']
                        }
            if i['fuel_type'] == 95:
                variant_sliv_95.append(i['variant_sliv'])
                volume_95.append(i['sum_sliv'])
                if i['variant_sliv'] in temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_95']:
                    index = temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_95'].index(i['variant_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_95'].insert(index, temp_azs_trucks4_dict[
                        str(i['variant'])]['volume_95'][index] + i['sum_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_95'].pop(index + 1)
                else:
                    temp_azs_trucks4_dict[str(i['variant'])] = {
                        'variant_sliv_92': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_92'],
                        'variant_sliv_95': temp_azs_trucks4_dict[str(i['variant'])][
                                               'variant_sliv_95'] + variant_sliv_95,
                        'variant_sliv_50': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_50'],
                        'volume_92': temp_azs_trucks4_dict[str(i['variant'])]['volume_92'],
                        'volume_95': temp_azs_trucks4_dict[str(i['variant'])]['volume_95'] + volume_95,
                        'volume_50': temp_azs_trucks4_dict[str(i['variant'])]['volume_50'],
                        'azs_id': temp_azs_trucks4_dict[str(i['variant'])][
                            'azs_id'],
                        'truck_id': temp_azs_trucks4_dict[str(i['variant'])][
                            'truck_id']

                        }
            if i['fuel_type'] == 50:
                variant_sliv_50.append(i['variant_sliv'])
                volume_50.append(i['sum_sliv'])
                if i['variant_sliv'] in temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_50']:
                    index = temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_50'].index(i['variant_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_50'].insert(index, temp_azs_trucks4_dict[
                        str(i['variant'])]['volume_50'][index] + i['sum_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_50'].pop(index + 1)
                else:
                    temp_azs_trucks4_dict[str(i['variant'])] = {
                        'variant_sliv_92': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_92'],
                        'variant_sliv_95': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_95'],
                        'variant_sliv_50': temp_azs_trucks4_dict[str(i['variant'])][
                                               'variant_sliv_50'] + variant_sliv_50,
                        'volume_92': temp_azs_trucks4_dict[str(i['variant'])]['volume_92'],
                        'volume_95': temp_azs_trucks4_dict[str(i['variant'])]['volume_95'],
                        'volume_50': temp_azs_trucks4_dict[str(i['variant'])]['volume_50'] + volume_50,
                        'azs_id': temp_azs_trucks4_dict[str(i['variant'])][
                            'azs_id'],
                        'truck_id': temp_azs_trucks4_dict[str(i['variant'])][
                            'truck_id']
                        }
        azs_trucks_4_list_final = list()
        azs_trucks_4_dict_final = {'variant': 0,
                                   'sum_92': 0,
                                   'sum_95': 0,
                                   'sum_50': 0,
                                   'min_rez1': 0,
                                   'min_rez2': 0,
                                   'min_rez3': 0,
                                   'variant_sliv_50': 0,
                                   'variant_sliv_92': 0,
                                   'variant_sliv_95': 0,
                                   'azs_id': 0,
                                   'truck_id': 0
                                   }

        for i in temp_azs_trucks4_dict:
            variant_sliv_92 = temp_azs_trucks4_dict[i]['variant_sliv_92']
            variant_sliv_95 = temp_azs_trucks4_dict[i]['variant_sliv_95']
            variant_sliv_50 = temp_azs_trucks4_dict[i]['variant_sliv_50']
            volume_92 = temp_azs_trucks4_dict[i]['volume_92']
            volume_95 = temp_azs_trucks4_dict[i]['volume_95']
            volume_50 = temp_azs_trucks4_dict[i]['volume_50']
            if len(variant_sliv_92) != 0 and len(variant_sliv_95) != 0 and len(variant_sliv_50) != 0:
                for index_a, a in enumerate(variant_sliv_92):
                    for index_b, b in enumerate(variant_sliv_95):
                        for index_c, c in enumerate(variant_sliv_50):
                            azs_trucks_4_dict_final = {'variant': i,
                                                       'sum_92': volume_92[index_a],
                                                       'sum_95': volume_95[index_b],
                                                       'sum_50': volume_50[index_c],
                                                       'min_rez1': 0,
                                                       'min_rez2': 0,
                                                       'min_rez3': 0,
                                                       'variant_sliv_50': c,
                                                       'variant_sliv_92': a,
                                                       'variant_sliv_95': b,
                                                       'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                                       'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                                       }
                            azs_trucks_4_list_final.append(azs_trucks_4_dict_final)

            if len(variant_sliv_92) != 0 and len(variant_sliv_95) != 0 and len(variant_sliv_50) == 0:
                for index_a, a in enumerate(variant_sliv_92):
                    for index_b, b in enumerate(variant_sliv_95):
                        azs_trucks_4_dict_final = {'variant': i,
                                                   'sum_92': volume_92[index_a],
                                                   'sum_95': volume_95[index_b],
                                                   'sum_50': 0,
                                                   'min_rez1': 0,
                                                   'min_rez2': 0,
                                                   'min_rez3': 0,
                                                   'variant_sliv_50': 0,
                                                   'variant_sliv_92': a,
                                                   'variant_sliv_95': b,
                                                   'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                                   'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                                   }
                        azs_trucks_4_list_final.append(azs_trucks_4_dict_final)

            if len(variant_sliv_92) != 0 and len(variant_sliv_95) == 0 and len(variant_sliv_50) != 0:
                for index_a, a in enumerate(variant_sliv_92):
                    for index_c, c in enumerate(variant_sliv_50):
                        azs_trucks_4_dict_final = {'variant': i,
                                                   'sum_92': volume_92[index_a],
                                                   'sum_95': 0,
                                                   'sum_50': volume_50[index_c],
                                                   'min_rez1': 0,
                                                   'min_rez2': 0,
                                                   'min_rez3': 0,
                                                   'variant_sliv_50': c,
                                                   'variant_sliv_92': a,
                                                   'variant_sliv_95': 0,
                                                   'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                                   'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                                   }
                        azs_trucks_4_list_final.append(azs_trucks_4_dict_final)
            if len(variant_sliv_92) == 0 and len(variant_sliv_95) != 0 and len(variant_sliv_50) != 0:
                for index_b, b in enumerate(variant_sliv_95):
                    for index_c, c in enumerate(variant_sliv_50):
                        azs_trucks_4_dict_final = {'variant': i,
                                                   'sum_92': 0,
                                                   'sum_95': volume_95[index_b],
                                                   'sum_50': volume_50[index_c],
                                                   'min_rez1': 0,
                                                   'min_rez2': 0,
                                                   'min_rez3': 0,
                                                   'variant_sliv_50': c,
                                                   'variant_sliv_92': 0,
                                                   'variant_sliv_95': b,
                                                   'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                                   'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                                   }
                        azs_trucks_4_list_final.append(azs_trucks_4_dict_final)
            if len(variant_sliv_92) == 0 and len(variant_sliv_95) == 0 and len(variant_sliv_50) != 0:
                for index_c, c in enumerate(variant_sliv_50):
                    azs_trucks_4_dict_final = {'variant': i,
                                               'sum_92': 0,
                                               'sum_95': 0,
                                               'sum_50': volume_50[index_c],
                                               'min_rez1': 0,
                                               'min_rez2': 0,
                                               'min_rez3': 0,
                                               'variant_sliv_50': c,
                                               'variant_sliv_92': 0,
                                               'variant_sliv_95': 0,
                                               'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                               'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                               }
                    azs_trucks_4_list_final.append(azs_trucks_4_dict_final)
            if len(variant_sliv_92) == 0 and len(variant_sliv_95) != 0 and len(variant_sliv_50) == 0:
                for index_b, b in enumerate(variant_sliv_95):
                    azs_trucks_4_dict_final = {'variant': i,
                                               'sum_92': 0,
                                               'sum_95': volume_95[index_b],
                                               'sum_50': 0,
                                               'min_rez1': 0,
                                               'min_rez2': 0,
                                               'min_rez3': 0,
                                               'variant_sliv_50': 0,
                                               'variant_sliv_92': 0,
                                               'variant_sliv_95': b,
                                               'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                               'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                               }
                    azs_trucks_4_list_final.append(azs_trucks_4_dict_final)
            if len(variant_sliv_92) != 0 and len(variant_sliv_95) == 0 and len(variant_sliv_50) == 0:
                for index_a, a in enumerate(variant_sliv_92):
                    azs_trucks_4_dict_final = {'variant': i,
                                               'sum_92': volume_92[index_a],
                                               'sum_95': 0,
                                               'sum_50': 0,
                                               'min_rez1': 0,
                                               'min_rez2': 0,
                                               'min_rez3': 0,
                                               'variant_sliv_50': 0,
                                               'variant_sliv_92': a,
                                               'variant_sliv_95': 0,
                                               'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                               'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                               }
                    azs_trucks_4_list_final.append(azs_trucks_4_dict_final)

        db.engine.execute(TempAzsTrucks3.__table__.insert(), temp_azs_trucks3_list)
        return azs_trucks_4_list_final, temp_azs_trucks3_list

    # определение худшего запаса суток среди всех резервуаров АЗС
    def preparation_six():
        # берем таблицу 4
        table_azs_trucks_4, table_azs_trucks_3 = preparation_four()
        table_azs_trucks_4_list = list()
        fuel_realisation = FuelRealisation.query.all()
        days_stock_old_dict = dict()
        for i in fuel_realisation:
            days_stock_old_dict[i.azs_id] = {'days_stock': [],
                                             'tank_id': []}

        for i in fuel_realisation:
            days_stock_min_old_list = [i.days_stock_min]
            tank_ids_list = [i.tank_id]
            days_stock_old_dict[i.azs_id] = {
                'days_stock': days_stock_old_dict[i.azs_id]['days_stock'] + days_stock_min_old_list,
                'tank_id': days_stock_old_dict[i.azs_id]['tank_id'] + tank_ids_list}
        variants_dict = dict()
        for i in table_azs_trucks_3:
            variants_dict[(i['variant'], i['variant_sliv'])] = {'days_stock': [],
                                                                'tank_id': []}
        for i in table_azs_trucks_3:
            days_stock_min_new_list = [i['new_days_stock']]
            tank_ids_list = [i['tank_id']]
            variants_dict[(i['variant'], i['variant_sliv'])] = {
                'days_stock': variants_dict[(i['variant'], i['variant_sliv'])]['days_stock'] + days_stock_min_new_list,
                'tank_id': variants_dict[(i['variant'], i['variant_sliv'])]['tank_id'] + tank_ids_list
                }
        for row in table_azs_trucks_4:

            variant = row['variant']
            variants_list = list()
            if row['variant_sliv_92']:
                variants_list.append(row['variant_sliv_92'])
            if row['variant_sliv_95']:
                variants_list.append(row['variant_sliv_95'])
            if row['variant_sliv_50']:
                variants_list.append(row['variant_sliv_50'])
            new_days_stock_dict = dict()
            tank_list = days_stock_old_dict[row['azs_id']]['tank_id']
            stock_list = days_stock_old_dict[row['azs_id']]['days_stock']
            for index, tank in enumerate(tank_list):
                new_days_stock_dict[tank] = stock_list[index]
            for variant_sliv in variants_list:
                tank_list_new = variants_dict[(int(row['variant']), variant_sliv)]['tank_id']
                stock_list_new = variants_dict[(int(row['variant']), variant_sliv)]['days_stock']
                for index, tank in enumerate(tank_list_new):
                    new_days_stock_dict[tank] = stock_list_new[index]

            sorted_new_days_stock_dict = sorted(new_days_stock_dict.items(), key=lambda x: x[1])
            if len(sorted_new_days_stock_dict) < 3:
                temp_azs_trucks_4_dict = {'truck_id': row['truck_id'],
                                          'azs_id': row['azs_id'],
                                          'variant': row['variant'],
                                          'sum_92': row['sum_92'],
                                          'sum_95': row['sum_95'],
                                          'sum_50': row['sum_50'],
                                          'min_rez1': round(sorted_new_days_stock_dict[0][1], 1),
                                          'min_rez2': round(sorted_new_days_stock_dict[1][1], 1),
                                          'min_rez3': 1000,
                                          'variant_sliv_92': row['variant_sliv_92'],
                                          'variant_sliv_95': row['variant_sliv_95'],
                                          'variant_sliv_50': row['variant_sliv_50']
                                          }

            else:
                temp_azs_trucks_4_dict = {'truck_id': row['truck_id'],
                                          'azs_id': row['azs_id'],
                                          'variant': row['variant'],
                                          'sum_92': row['sum_92'],
                                          'sum_95': row['sum_95'],
                                          'sum_50': row['sum_50'],
                                          'min_rez1': round(sorted_new_days_stock_dict[0][1], 1),
                                          'min_rez2': round(sorted_new_days_stock_dict[1][1], 1),
                                          'min_rez3': round(sorted_new_days_stock_dict[2][1], 1),
                                          'variant_sliv_92': row['variant_sliv_92'],
                                          'variant_sliv_95': row['variant_sliv_95'],
                                          'variant_sliv_50': row['variant_sliv_50']
                                          }
            table_azs_trucks_4_list.append(temp_azs_trucks_4_dict)

        db.engine.execute(TempAzsTrucks4.__table__.insert(), table_azs_trucks_4_list)

    preparation_six()
    return redirect(url_for('main.wait_first'))


@bp.route('/wait_first', methods=['POST', 'GET'])
@login_required
def wait_first():
    time.sleep(3)
    return redirect(url_for('main.start_first_trip'))


@bp.route('/start_first_trip', methods=['POST', 'GET'])
@login_required
def start_first_trip():
    def create_trip():
        logs = UserLogs(user_id=current_user.id,
                        action="trip_creation_started",
                        timestamp=datetime.now())
        db.session.add(logs)
        db.session.commit()
        work_type = WorkType.query.filter_by(active=True).first_or_404()
        if work_type.id == 2 or work_type.id == 3:
            fuel_type = int(work_type.fuel_type)
            min_days_stock_global = float(work_type.days_stock_limit)
        else:
            fuel_type = 0

        table_azs_trucks_4 = TempAzsTrucks4.query.all()
        trucks_for_azs_dict = dict()
        azs_trucks_best_days_stock = dict()
        azs_trucks_max_92 = dict()
        azs_trucks_max_95 = dict()
        azs_trucks_max_50 = dict()
        azs_trucks_min_92 = dict()
        azs_trucks_min_95 = dict()
        azs_trucks_min_50 = dict()
        azs_for_trucks_dict = dict()
        for i in table_azs_trucks_4:
            # словарь для хранения информации о том, на какие АЗС может отправиться бензовоз
            azs_for_trucks_dict[i.truck_id] = {'azs_ids': [],
                                               'priority': []}

            # словарь для хранения информации о том, какие бензовозы могут отправиться на данную АЗС
            trucks_for_azs_dict[i.azs_id] = {'azs_trucks': [0]}
            # Словарь для первого режима работы
            # Информация о лучщем заполнении для пары АЗС:БЕНЗОВОЗ
            azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_rez1': -1,
                                                                                 'min_rez2': -1,
                                                                                 'min_rez3': -1,
                                                                                 'variant': 0,
                                                                                 'variant_sliv_92': 0,
                                                                                 'variant_sliv_95': 0,
                                                                                 'variant_sliv_50': 0}
            # Словари для второго режима работы (вывоз максимального количества определенного топлива)
            azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_92': -1,
                                                                        'max_volume_95': -1,
                                                                        'max_volume_50': -1,
                                                                        'min_rez1': -1,
                                                                        'variant': 0,
                                                                        'variant_sliv_92': 0,
                                                                        'variant_sliv_95': 0,
                                                                        'variant_sliv_50': 0
                                                                        }

            azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_92': -1,
                                                                        'max_volume_95': -1,
                                                                        'max_volume_50': -1,
                                                                        'min_rez1': -1,
                                                                        'variant': 0,
                                                                        'variant_sliv_92': 0,
                                                                        'variant_sliv_95': 0,
                                                                        'variant_sliv_50': 0
                                                                        }

            azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_50': -1,
                                                                        'min_rez1': -1,
                                                                        'variant': 0,
                                                                        'variant_sliv_92': 0,
                                                                        'variant_sliv_95': 0,
                                                                        'variant_sliv_50': 0
                                                                        }
            # Словари для третьего режима работы (вывоз минимального количества топлива определенного вида)
            azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_92': -1,
                                                                        'min_rez1': -1,
                                                                        'min_rez2': -1,
                                                                        'min_rez3': -1,
                                                                        'variant': 0,
                                                                        'variant_sliv_92': 0,
                                                                        'variant_sliv_95': 0,
                                                                        'variant_sliv_50': 0
                                                                        }

            azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_95': -1,
                                                                        'min_rez1': -1,
                                                                        'min_rez2': -1,
                                                                        'min_rez3': -1,
                                                                        'variant': 0,
                                                                        'variant_sliv_92': 0,
                                                                        'variant_sliv_95': 0,
                                                                        'variant_sliv_50': 0
                                                                        }

            azs_trucks_min_50[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_50': -1,
                                                                        'min_rez1': -1,
                                                                        'min_rez2': -1,
                                                                        'min_rez3': -1,
                                                                        'variant': 0,
                                                                        'variant_sliv_92': 0,
                                                                        'variant_sliv_95': 0,
                                                                        'variant_sliv_50': 0
                                                                        }
        # Перебираем таблицу table_azs_truck_4 для заполнения словарей для всех трех режимов работы
        for i in table_azs_trucks_4:
            # заполняем словарь для хранения информации о том, на какие азс может отправиться бензовоз
            azs_list = list()
            azs_list.append(i.azs_id)
            if i.azs_id not in azs_for_trucks_dict[i.truck_id]['azs_ids']:
                azs_for_trucks_dict[i.truck_id] = {'azs_ids': azs_for_trucks_dict[i.truck_id]['azs_ids'] + azs_list,
                                                   'priority': 0}

            # заполняем словарь для хранения информации о том, какие бензовозы могут отправиться на данную АЗС
            trucks_list = list()
            trucks_list.append(i.truck_id)
            if i.truck_id not in trucks_for_azs_dict[i.azs_id]['azs_trucks']:
                trucks_for_azs_dict[i.azs_id] = {'azs_trucks': trucks_for_azs_dict[i.azs_id]['azs_trucks'] + trucks_list
                                                 }

            # Заполняем словари информацией о лучщем заполнении для пары АЗС:БЕНЗОВОЗ (для первого режима работы)
            min_rez1 = azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
            min_rez2 = azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez2']
            min_rez3 = azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez3']

            if i.min_rez1 > min_rez1 \
                    or (i.min_rez1 == min_rez1 and i.min_rez2 > min_rez2) \
                    or (i.min_rez1 == min_rez1 and i.min_rez2 == min_rez2 and i.min_rez3 > min_rez3):
                azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_rez1': i.min_rez1,
                                                                                     'min_rez2': i.min_rez2,
                                                                                     'min_rez3': i.min_rez3,
                                                                                     'variant': i.variant,
                                                                                     'variant_sliv_92': i.variant_sliv_92,
                                                                                     'variant_sliv_95': i.variant_sliv_95,
                                                                                     'variant_sliv_50': i.variant_sliv_50}

        if work_type.id == 2:
            if min_days_stock_global == -1:
                for i in table_azs_trucks_4:
                    # Заполняем словари для второго режима работы(вывоз максимального количества топлива определенного вида)
                    min_rez1_92 = azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
                    min_rez1_95 = azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
                    min_rez1_50 = azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']

                    max_volume_92 = azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)]['max_volume_92']
                    max_volume_95 = azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)]['max_volume_95']
                    max_volume_50 = azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)]['max_volume_50']

                    if i.min_rez1 > min_rez1_92 or (i.min_rez1 == min_rez1_92 and i.sum_92 > max_volume_92):
                        azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_92': i.sum_92,
                                                                                    'max_volume_95': i.sum_95,
                                                                                    'max_volume_50': i.sum_50,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }
                    if i.min_rez1 > min_rez1_95 or (i.min_rez1 == min_rez1_95 and i.sum_95 > max_volume_95):
                        azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_95': i.sum_95,
                                                                                    'max_volume_92': i.sum_92,
                                                                                    'max_volume_50': i.sum_50,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }
                    if i.min_rez1 > min_rez1_50 or (i.min_rez1 == min_rez1_50 and i.sum_50 > max_volume_50):
                        azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_50': i.sum_50,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }
            else:
                for i in table_azs_trucks_4:
                    max_volume_92 = azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)]['max_volume_92']
                    max_volume_95 = azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)]['max_volume_95']
                    max_volume_50 = azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)]['max_volume_50']

                    if i.min_rez1 >= min_days_stock_global and i.sum_92 > max_volume_92:
                        azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_92': i.sum_92,
                                                                                    'max_volume_95': i.sum_95,
                                                                                    'max_volume_50': i.sum_50,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }

                    if (i.min_rez1 >= min_days_stock_global and i.sum_95 > max_volume_95):
                        azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_95': i.sum_95,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }

                    if (i.min_rez1 >= min_days_stock_global and i.sum_50 > max_volume_50):
                        azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_50': i.sum_50,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }
                for i in table_azs_trucks_4:
                    min_rez1_92 = azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
                    min_rez1_95 = azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
                    min_rez1_50 = azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']

                    max_volume_92 = azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)]['max_volume_92']
                    max_volume_95 = azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)]['max_volume_95']
                    max_volume_50 = azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)]['max_volume_50']

                    if azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1'] == -1:
                        if i.min_rez1 > min_rez1_92 or (i.min_rez1 == min_rez1_92 and i.sum_92 > max_volume_92):
                            azs_trucks_max_92[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_92': i.sum_92,
                                                                                        'max_volume_95': i.sum_95,
                                                                                        'max_volume_50': i.sum_50,
                                                                                        'min_rez1': i.min_rez1,
                                                                                        'variant': i.variant,
                                                                                        'variant_sliv_92': i.variant_sliv_92,
                                                                                        'variant_sliv_95': i.variant_sliv_95,
                                                                                        'variant_sliv_50': i.variant_sliv_50
                                                                                        }
                    if azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1'] == -1:
                        if i.min_rez1 > min_rez1_95 or (i.min_rez1 == min_rez1_95 and i.sum_95 > max_volume_95):
                            azs_trucks_max_95[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_95': i.sum_95,
                                                                                        'min_rez1': i.min_rez1,
                                                                                        'variant': i.variant,
                                                                                        'variant_sliv_92': i.variant_sliv_92,
                                                                                        'variant_sliv_95': i.variant_sliv_95,
                                                                                        'variant_sliv_50': i.variant_sliv_50
                                                                                        }
                    if azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1'] == -1:
                        if i.min_rez1 > min_rez1_50 or (i.min_rez1 == min_rez1_50 and i.sum_50 > max_volume_50):
                            azs_trucks_max_50[str(i.azs_id) + ':' + str(i.truck_id)] = {'max_volume_50': i.sum_50,
                                                                                        'min_rez1': i.min_rez1,
                                                                                        'variant': i.variant,
                                                                                        'variant_sliv_92': i.variant_sliv_92,
                                                                                        'variant_sliv_95': i.variant_sliv_95,
                                                                                        'variant_sliv_50': i.variant_sliv_50
                                                                                        }
        elif work_type.id == 3:
            if min_days_stock_global == -1:
                for i in table_azs_trucks_4:
                    # Заполняем словари для второго режима работы
                    # (вывоз минимального количества топлива определенного вида)
                    min_rez1_92 = azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
                    min_rez1_95 = azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
                    min_rez1_50 = azs_trucks_min_50[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']

                    min_volume_92 = azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)]['min_volume_92']
                    min_volume_95 = azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)]['min_volume_95']
                    min_volume_50 = azs_trucks_min_50[str(i.azs_id) + ':' + str(i.truck_id)]['min_volume_50']

                    if i.min_rez1 > min_rez1_92 or (i.min_rez1 == min_rez1_92 and i.sum_92 < min_volume_92):
                        azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_92': i.sum_92,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }
                    if i.min_rez1 > min_rez1_95 or (i.min_rez1 == min_rez1_95 and i.sum_95 < min_volume_95):
                        azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_95': i.sum_95,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }
                    if i.min_rez1 > min_rez1_50 or (i.min_rez1 == min_rez1_50 and i.sum_50 < min_volume_50):
                        azs_trucks_min_50[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_50': i.sum_50,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }
            else:
                for i in table_azs_trucks_4:
                    min_volume_92 = azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)]['min_volume_92']
                    min_volume_95 = azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)]['min_volume_95']
                    min_volume_50 = azs_trucks_min_50[str(i.azs_id) + ':' + str(i.truck_id)]['min_volume_50']
                    if (i.min_rez1 >= min_days_stock_global and i.sum_92 < min_volume_92):
                        azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_92': i.sum_92,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }

                    if (i.min_rez1 >= min_days_stock_global and i.sum_95 < min_volume_95):
                        azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_95': i.sum_95,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }

                    if (i.min_rez1 >= min_days_stock_global and i.sum_50 < min_volume_50):
                        azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_50': i.sum_50,
                                                                                    'min_rez1': i.min_rez1,
                                                                                    'variant': i.variant,
                                                                                    'variant_sliv_92': i.variant_sliv_92,
                                                                                    'variant_sliv_95': i.variant_sliv_95,
                                                                                    'variant_sliv_50': i.variant_sliv_50
                                                                                    }
                for i in table_azs_trucks_4:
                    min_rez1_92 = azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
                    min_rez1_95 = azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
                    min_rez1_50 = azs_trucks_min_50[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']

                    min_volume_92 = azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)]['min_volume_92']
                    min_volume_95 = azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)]['min_volume_95']
                    min_volume_50 = azs_trucks_min_50[str(i.azs_id) + ':' + str(i.truck_id)]['min_volume_50']
                    if azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1'] == -1:
                        if i.min_rez1 > min_rez1_92 or (i.min_rez1 == min_rez1_92 and i.sum_92 < min_volume_92):
                            azs_trucks_min_92[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_92': i.sum_92,
                                                                                        'min_rez1': i.min_rez1,
                                                                                        'variant': i.variant,
                                                                                        'variant_sliv_92': i.variant_sliv_92,
                                                                                        'variant_sliv_95': i.variant_sliv_95,
                                                                                        'variant_sliv_50': i.variant_sliv_50
                                                                                        }
                    if azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1'] == -1:
                        if i.min_rez1 > min_rez1_95 or (i.min_rez1 == min_rez1_95 and i.sum_95 < min_volume_95):
                            azs_trucks_min_95[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_95': i.sum_95,
                                                                                        'min_rez1': i.min_rez1,
                                                                                        'variant': i.variant,
                                                                                        'variant_sliv_92': i.variant_sliv_92,
                                                                                        'variant_sliv_95': i.variant_sliv_95,
                                                                                        'variant_sliv_50': i.variant_sliv_50
                                                                                        }
                    if azs_trucks_min_50[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1'] == -1:
                        if i.min_rez1 > min_rez1_50 or (i.min_rez1 == min_rez1_50 and i.sum_50 < min_volume_50):
                            azs_trucks_min_50[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_volume_50': i.sum_50,
                                                                                        'min_rez1': i.min_rez1,
                                                                                        'variant': i.variant,
                                                                                        'variant_sliv_92': i.variant_sliv_92,
                                                                                        'variant_sliv_95': i.variant_sliv_95,
                                                                                        'variant_sliv_50': i.variant_sliv_50
                                                                                        }
        # Заполняем словари для третьего режима работы (вывоз минимального количества топлива определенного вида)

        # ПОДГОТОВКА К РАССТАНОВКЕ БЕНЗОВОЗОВ

        global_settings = GlobalSettings.query.filter_by(name="algorithm").first()
        variant_send_truck = global_settings.algorithm_id
        choise_good = 0
        active_azs_count = Priority.query.count()
        active_trucks = Trucks.query.filter_by(active=True).count()  # получаем количество активных бензовозов
        active_azs = Priority.query.order_by("priority").all()  # получаем список активных АЗС из таблицы Priority
        # с сортировкой по важности (чем меньше число (стобец priority), тем важнее отправить бензовоз на эту АЗС
        choices_dict_work_type_1 = dict()  # храним итоговые варианты расстановки с оценкой каждого для 1 режима работы
        choices_dict_work_type_2 = dict()  # храним итоговые варианты расстановки с оценкой каждого для 2 режима работы
        choices_dict_work_type_3 = dict()  # храним итоговые варианты расстановки с оценкой каждого для 3 режима работы
        azs_queue_dict = dict()  # создаем словарь для хранения id АЗС в порядке важности отправки бензовоза на АЗС
        # (нужна для последующего анализа итоговых расстановок)

        for i in active_azs:  # заполняем словарь
            azs_queue_dict[i.azs_id] = {'queue': i.priority}

        # таймаут для принудительной остановки расстановки бензовозов через
        # указанное количество времени (сейчас минута)
        timeout = time.time() + 60 * 1

        # создаем словарь для хранения всех удачных расстановок
        good_choices_dict = dict()

        # количество успешных расстановок
        number_of_success_loops = 0
        # вводим переменную на случай, если расстановка бензовозов не удастся.
        # Изначально переменной присвоена 1, что означает неудачную расстановку.
        alarm = 1

        # Главный цикл расстановки бензовозов (количество попыток от 0 до 10 млн)

        for x in trucks_for_azs_dict:
            print("АЗС:", x, "БЕЗОВОЗЫ:", trucks_for_azs_dict[x])

        # АЛГОРИТМ №1 - СЛУЧАНАЯ РАССТАНОВКА С ОГРОМНЫМ КОЛИЧЕСТВОМ ВАРИАНТОВ
        if variant_send_truck == 1:
            if active_trucks <= active_azs_count:
                trucks_for_azs_dict_sorted = dict()
                for i in active_azs:
                    azs_id = i.azs_id
                    if azs_id in trucks_for_azs_dict:
                        trucks_for_azs_dict_sorted[azs_id] = {'azs_trucks': trucks_for_azs_dict[azs_id]['azs_trucks']}

                trucks_for_azs_list_sorted = sorted(trucks_for_azs_dict_sorted,
                                                    key=lambda k: len(trucks_for_azs_dict_sorted[k]['azs_trucks']))
                print('Sorted table')
                for k in sorted(trucks_for_azs_dict_sorted,
                                key=lambda k: len(trucks_for_azs_dict_sorted[k]['azs_trucks'])):
                    print(k, trucks_for_azs_dict_sorted[k]['azs_trucks'])

                print('Start!')

                choise_good = 0  # переменная-триггер, по которой определяем, что расстановка удалась (или нет!)
                for choice in range(0, 1000000000):
                    choice_azs_truck_dict = dict()
                    used_trucks = list()
                    temp_truck_count = 0
                    for i in trucks_for_azs_list_sorted:  # перебираем все активные АЗС
                        azs_id = i
                        if azs_id in trucks_for_azs_dict:  # если есть хотябы один бензовоз, который можно отправить на АЗС
                            azs_trucks = trucks_for_azs_dict[azs_id]['azs_trucks']  # получаем список всех бензовозов,
                            # которые можно отправить на эту АЗС (включая 0 - т.е. АЗС на которую не будет отправлен бензовоз)
                            truck_id = random.choice(azs_trucks)  # функцией RANDOM из списка azs_trucks
                            # выбираем бензовоз для этой АЗС
                            if truck_id in used_trucks and truck_id != 0:  # если данный безовоз уже был в данном варианте
                                # расстановки и он не равен 0, то считаем вариант не удачным, и досрочно прерыываем цикл
                                # good = 0
                                break
                            # если все хорошо, то
                            else:
                                # добавляем этот бензовоз к списку использованных
                                # в данном варианте бензовозов
                                used_trucks.append(truck_id)

                                # добавляем параметр azs_id-truck_id
                                # в словарь с расстановкой
                                choice_azs_truck_dict[azs_id] = {'truck_id': truck_id}
                                # если безовоз не нулевой, то уменьшаем количество бензовозов которые
                                if truck_id != 0:
                                    # требуется расставить на 1
                                    temp_truck_count = temp_truck_count + 1
                                    # если все бензовозы расставлены (счетчик равен 0)

                    if temp_truck_count == active_trucks:
                        # то помечаем вариант хорошим
                        choise_good = 1
                        number_of_success_loops = number_of_success_loops + 1
                        good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict}

                    if time.time() > timeout:
                        # то цикл принудительно прерывается
                        break

                if choise_good == 1:
                    print("OK! Нашли ", number_of_success_loops, "вариантов")
                else:
                    print("NOT OK!!!")
            else:
                print("Расстановка бензовозов невозможна! Количество активных бензовозов больше числа активных АЗС!")

        # АЛГОРИТМ №2 - СЛУЧАНАЯ РАССТАНОВКА ОТ ВЕРХА К НИЗУ
        if variant_send_truck == 2:
            if active_trucks <= active_azs_count:
                choise_good = 0
                for choice in range(0, 1000000000):
                    choice_azs_truck_dict = dict()
                    used_trucks = list()
                    temp_truck_count = 0
                    good = 0
                    for i in active_azs:  # перебираем все активные АЗС
                        azs_id = i.azs_id
                        if azs_id in trucks_for_azs_dict:  # если есть хотябы один бензовоз, который можно отправить на АЗС
                            trig = 0
                            for x in range(1, 1000):
                                azs_trucks = trucks_for_azs_dict[azs_id][
                                    'azs_trucks']  # получаем список всех бензовозов,
                                # которые можно отправить на эту АЗС (включая 0 - т.е. АЗС на которую не будет отправлен бензовоз)
                                truck_id = random.choice(azs_trucks)  # функцией RANDOM из списка azs_trucks
                                # выбираем бензовоз для этой АЗС
                                if truck_id in used_trucks and truck_id != 0:  # если данный безовоз уже был в данном варианте
                                    d = 1  # заглушка
                                else:  # если все хорошо, то
                                    # добавляем этот бензовоз к списку использованных
                                    # в данном варианте бензовозов
                                    used_trucks.append(truck_id)
                                    # добавляем параметр azs_id-truck_id
                                    # в словарь с расстановкой
                                    choice_azs_truck_dict[azs_id] = {'truck_id': truck_id}
                                    trig = 1
                                    if truck_id != 0:
                                        temp_truck_count = temp_truck_count + 1
                                    break
                            if trig == 0:
                                used_trucks.append(0)
                                choice_azs_truck_dict[azs_id] = {'truck_id': 0}

                            if temp_truck_count == active_trucks:
                                good = 1
                                break

                    if good == 1:
                        number_of_success_loops = number_of_success_loops + 1
                        good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict}
                        choise_good = 1

                    if time.time() > timeout:
                        # то цикл принудительно прерывается
                        break

                if choise_good == 1:
                    print("OK! Нашли ", number_of_success_loops, "вариантов")
                    '''for rex in range(1, 10):
                        print('Вариант №', rex)
                        for xer in good_choices_dict[rex]['variants']:
                            print(xer, choice_azs_truck_dict[xer]['truck_id'])
                        print("*------*------*------*------*")'''
                else:
                    print("NOT OK!!!")
            else:
                print("Расстановка бензовозов невозможна! Количество активных бензовозов больше числа активных АЗС!")

        # АЛГОРИТМ №3 - ПОЛНЫЙ ПЕРЕБОР

        if variant_send_truck == 3:
            choice_azs_truck_dict = dict()
            azs_for_trucks_with_priority = dict()
            for i in azs_for_trucks_dict:
                temp_azs_list = azs_for_trucks_dict[i]['azs_ids']
                temp_priority = list()
                for azs in temp_azs_list:
                    # if azs in azs_queue_dict:
                    temp_priority.append(azs_queue_dict[azs]['queue'])
                azs_for_trucks_with_priority[i] = {'azs_ids': azs_for_trucks_dict[i]['azs_ids'],
                                                   'priority': temp_priority}

            print('!!!!!!!!!!!!!!!АЛГОРИТМ № 3!!!!!!!!!!!!!!')
            bubble_azs_for_trucks = dict()
            # делаем сортировку пузырьком

            for i in azs_for_trucks_with_priority:

                temp_azs_list = azs_for_trucks_with_priority[i]['azs_ids']
                temp_priority = azs_for_trucks_with_priority[i]['priority']
                for n2 in range(1, len(temp_priority)):
                    for n in range(1, len(temp_priority)):
                        if temp_priority[n] < temp_priority[n - 1]:
                            temp = temp_priority[n - 1]
                            temp_priority[n - 1] = temp_priority[n]
                            temp_priority[n] = temp

                            temp2 = temp_azs_list[n - 1]
                            temp_azs_list[n - 1] = temp_azs_list[n]
                            temp_azs_list[n] = temp2
                bubble_azs_for_trucks[i] = {'azs_ids': temp_azs_list[:active_trucks - 1],
                                            'priority:': temp_priority[:active_trucks - 1]}

            k = 1
            final_azs_for_trucks = dict()
            truck_id_number = dict()
            for i in bubble_azs_for_trucks:
                truck_id_number[k] = {'truck_id': i}
                final_azs_for_trucks[k] = {"azs_ids": bubble_azs_for_trucks[i]['azs_ids']}
                k = k + 1

            if active_trucks == 11:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0

                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            for n3 in final_azs_for_trucks[3]["azs_ids"]:
                                if n3 != n2 and n3 != n1:
                                    for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                        if n4 != n3 and n4 != n2 and n4 != n1:
                                            for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                                if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                    for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                        if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                            for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                                if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                    for n8 in final_azs_for_trucks[8]["azs_ids"]:
                                                                        if n8 != n7 and n8 != n6 and n8 != n5 and n8 != n4 and n8 != n3 and n8 != n2 and n8 != n1:
                                                                            for n9 in final_azs_for_trucks[9][
                                                                                "azs_ids"]:
                                                                                if n9 != n8 and n9 != n7 and n9 != n6 and n9 != n5 and n9 != n4 and n9 != n3 and n9 != n2 and n9 != n1:
                                                                                    for n10 in final_azs_for_trucks[10][
                                                                                        "azs_ids"]:
                                                                                        if n10 != n9 and n10 != n8 and n10 != n7 and n10 != n6 and n10 != n5 and n10 != n4 and n10 != n3 and n10 != n2 and n10 != n1:
                                                                                            for n11 in \
                                                                                            final_azs_for_trucks[11][
                                                                                                "azs_ids"]:
                                                                                                if n11 != n10 and n11 != n9 and n11 != n8 and n11 != n7 and n11 != n6 and n11 != n5 and n11 != n4 and n11 != n3 and n11 != n2 and n11 != n1:

                                                                                                    choice_azs_truck_dict = dict()
                                                                                                    choice_azs_truck_dict[
                                                                                                        n1] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                1][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n2] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                2][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n3] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                3][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n4] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                4][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n5] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                5][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n6] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                6][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n7] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                7][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n8] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                8][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n9] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                9][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n10] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                10][
                                                                                                                'truck_id']}
                                                                                                    choice_azs_truck_dict[
                                                                                                        n11] = {
                                                                                                        'truck_id':
                                                                                                            truck_id_number[
                                                                                                                11][
                                                                                                                'truck_id']}
                                                                                                    number_variant = number_variant + 1
                                                                                                    if number_variant & 1000 == 0:
                                                                                                        number_of_success_loops = number_of_success_loops + 1
                                                                                                        good_choices_dict[
                                                                                                            number_of_success_loops] = {
                                                                                                            'variants': choice_azs_truck_dict}
            if active_trucks == 10:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0

                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            for n3 in final_azs_for_trucks[3]["azs_ids"]:
                                if n3 != n2 and n3 != n1:
                                    for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                        if n4 != n3 and n4 != n2 and n4 != n1:
                                            for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                                if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                    for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                        if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                            for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                                if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                    for n8 in final_azs_for_trucks[8]["azs_ids"]:
                                                                        if n8 != n7 and n8 != n6 and n8 != n5 and n8 != n4 and n8 != n3 and n8 != n2 and n8 != n1:
                                                                            for n9 in final_azs_for_trucks[9][
                                                                                "azs_ids"]:
                                                                                if n9 != n8 and n9 != n7 and n9 != n6 and n9 != n5 and n9 != n4 and n9 != n3 and n9 != n2 and n9 != n1:
                                                                                    for n10 in final_azs_for_trucks[10][
                                                                                        "azs_ids"]:
                                                                                        if n10 != n9 and n10 != n8 and n10 != n7 and n10 != n6 and n10 != n5 and n10 != n4 and n10 != n3 and n10 != n2 and n10 != n1:
                                                                                            choice_azs_truck_dict = dict()
                                                                                            choice_azs_truck_dict[
                                                                                                n1] = {'truck_id':
                                                                                                           truck_id_number[
                                                                                                               1][
                                                                                                               'truck_id']}
                                                                                            choice_azs_truck_dict[
                                                                                                n2] = {'truck_id':
                                                                                                           truck_id_number[
                                                                                                               2][
                                                                                                               'truck_id']}
                                                                                            choice_azs_truck_dict[
                                                                                                n3] = {'truck_id':
                                                                                                           truck_id_number[
                                                                                                               3][
                                                                                                               'truck_id']}
                                                                                            choice_azs_truck_dict[
                                                                                                n4] = {'truck_id':
                                                                                                           truck_id_number[
                                                                                                               4][
                                                                                                               'truck_id']}
                                                                                            choice_azs_truck_dict[
                                                                                                n5] = {'truck_id':
                                                                                                           truck_id_number[
                                                                                                               5][
                                                                                                               'truck_id']}
                                                                                            choice_azs_truck_dict[
                                                                                                n6] = {'truck_id':
                                                                                                           truck_id_number[
                                                                                                               6][
                                                                                                               'truck_id']}
                                                                                            choice_azs_truck_dict[
                                                                                                n7] = {'truck_id':
                                                                                                           truck_id_number[
                                                                                                               7][
                                                                                                               'truck_id']}
                                                                                            choice_azs_truck_dict[
                                                                                                n8] = {'truck_id':
                                                                                                           truck_id_number[
                                                                                                               8][
                                                                                                               'truck_id']}
                                                                                            choice_azs_truck_dict[
                                                                                                n9] = {'truck_id':
                                                                                                           truck_id_number[
                                                                                                               9][
                                                                                                               'truck_id']}
                                                                                            choice_azs_truck_dict[
                                                                                                n10] = {'truck_id':
                                                                                                            truck_id_number[
                                                                                                                10][
                                                                                                                'truck_id']}
                                                                                            number_variant = number_variant + 1
                                                                                            if number_variant & 1000 == 0:
                                                                                                number_of_success_loops = number_of_success_loops + 1
                                                                                                good_choices_dict[
                                                                                                    number_of_success_loops] = {
                                                                                                    'variants': choice_azs_truck_dict}
            if active_trucks == 9:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0

                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            for n3 in final_azs_for_trucks[3]["azs_ids"]:
                                if n3 != n2 and n3 != n1:
                                    for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                        if n4 != n3 and n4 != n2 and n4 != n1:
                                            for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                                if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                    for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                        if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                            for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                                if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                    for n8 in final_azs_for_trucks[8]["azs_ids"]:
                                                                        if n8 != n7 and n8 != n6 and n8 != n5 and n8 != n4 and n8 != n3 and n8 != n2 and n8 != n1:
                                                                            for n9 in final_azs_for_trucks[9][
                                                                                "azs_ids"]:
                                                                                if n9 != n8 and n9 != n7 and n9 != n6 and n9 != n5 and n9 != n4 and n9 != n3 and n9 != n2 and n9 != n1:
                                                                                    choice_azs_truck_dict = dict()
                                                                                    choice_azs_truck_dict[n1] = {
                                                                                        'truck_id': truck_id_number[1][
                                                                                            'truck_id']}
                                                                                    choice_azs_truck_dict[n2] = {
                                                                                        'truck_id': truck_id_number[2][
                                                                                            'truck_id']}
                                                                                    choice_azs_truck_dict[n3] = {
                                                                                        'truck_id': truck_id_number[3][
                                                                                            'truck_id']}
                                                                                    choice_azs_truck_dict[n4] = {
                                                                                        'truck_id': truck_id_number[4][
                                                                                            'truck_id']}
                                                                                    choice_azs_truck_dict[n5] = {
                                                                                        'truck_id': truck_id_number[5][
                                                                                            'truck_id']}
                                                                                    choice_azs_truck_dict[n6] = {
                                                                                        'truck_id': truck_id_number[6][
                                                                                            'truck_id']}
                                                                                    choice_azs_truck_dict[n7] = {
                                                                                        'truck_id': truck_id_number[7][
                                                                                            'truck_id']}
                                                                                    choice_azs_truck_dict[n8] = {
                                                                                        'truck_id': truck_id_number[8][
                                                                                            'truck_id']}
                                                                                    choice_azs_truck_dict[n9] = {
                                                                                        'truck_id': truck_id_number[9][
                                                                                            'truck_id']}

                                                                                    number_variant = number_variant + 1
                                                                                    if number_variant & 1000 == 0:
                                                                                        number_of_success_loops = number_of_success_loops + 1
                                                                                        good_choices_dict[
                                                                                            number_of_success_loops] = {
                                                                                            'variants': choice_azs_truck_dict}

            if active_trucks == 8:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0

                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            for n3 in final_azs_for_trucks[3]["azs_ids"]:
                                if n3 != n2 and n3 != n1:
                                    for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                        if n4 != n3 and n4 != n2 and n4 != n1:
                                            for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                                if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                    for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                        if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                            for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                                if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                    for n8 in final_azs_for_trucks[8]["azs_ids"]:
                                                                        if n8 != n7 and n8 != n6 and n8 != n5 and n8 != n4 and n8 != n3 and n8 != n2 and n8 != n1:
                                                                            choice_azs_truck_dict = dict()
                                                                            choice_azs_truck_dict[n1] = {
                                                                                'truck_id': truck_id_number[1][
                                                                                    'truck_id']}
                                                                            choice_azs_truck_dict[n2] = {
                                                                                'truck_id': truck_id_number[2][
                                                                                    'truck_id']}
                                                                            choice_azs_truck_dict[n3] = {
                                                                                'truck_id': truck_id_number[3][
                                                                                    'truck_id']}
                                                                            choice_azs_truck_dict[n4] = {
                                                                                'truck_id': truck_id_number[4][
                                                                                    'truck_id']}
                                                                            choice_azs_truck_dict[n5] = {
                                                                                'truck_id': truck_id_number[5][
                                                                                    'truck_id']}
                                                                            choice_azs_truck_dict[n6] = {
                                                                                'truck_id': truck_id_number[6][
                                                                                    'truck_id']}
                                                                            choice_azs_truck_dict[n7] = {
                                                                                'truck_id': truck_id_number[7][
                                                                                    'truck_id']}
                                                                            choice_azs_truck_dict[n8] = {
                                                                                'truck_id': truck_id_number[8][
                                                                                    'truck_id']}

                                                                            number_variant = number_variant + 1
                                                                            number_of_success_loops = number_of_success_loops + 1
                                                                            good_choices_dict[
                                                                                number_of_success_loops] = {
                                                                                'variants': choice_azs_truck_dict}
            if active_trucks == 7:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0
                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            for n3 in final_azs_for_trucks[3]["azs_ids"]:
                                if n3 != n2 and n3 != n1:
                                    for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                        if n4 != n3 and n4 != n2 and n4 != n1:
                                            for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                                if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                    for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                        if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                            for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                                if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                    choice_azs_truck_dict = dict()
                                                                    choice_azs_truck_dict[n1] = {
                                                                        'truck_id': truck_id_number[1]['truck_id']}
                                                                    choice_azs_truck_dict[n2] = {
                                                                        'truck_id': truck_id_number[2]['truck_id']}
                                                                    choice_azs_truck_dict[n3] = {
                                                                        'truck_id': truck_id_number[3]['truck_id']}
                                                                    choice_azs_truck_dict[n4] = {
                                                                        'truck_id': truck_id_number[4]['truck_id']}
                                                                    choice_azs_truck_dict[n5] = {
                                                                        'truck_id': truck_id_number[5]['truck_id']}
                                                                    choice_azs_truck_dict[n6] = {
                                                                        'truck_id': truck_id_number[6]['truck_id']}
                                                                    choice_azs_truck_dict[n7] = {
                                                                        'truck_id': truck_id_number[7]['truck_id']}

                                                                    number_variant = number_variant + 1

                                                                    number_of_success_loops = number_of_success_loops + 1
                                                                    good_choices_dict[number_of_success_loops] = {
                                                                        'variants': choice_azs_truck_dict}

            if active_trucks == 6:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0

                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            for n3 in final_azs_for_trucks[3]["azs_ids"]:
                                if n3 != n2 and n3 != n1:
                                    for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                        if n4 != n3 and n4 != n2 and n4 != n1:
                                            for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                                if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                    for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                        if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                            choice_azs_truck_dict = dict()
                                                            choice_azs_truck_dict[n1] = {
                                                                'truck_id': truck_id_number[1]['truck_id']}
                                                            choice_azs_truck_dict[n2] = {
                                                                'truck_id': truck_id_number[2]['truck_id']}
                                                            choice_azs_truck_dict[n3] = {
                                                                'truck_id': truck_id_number[3]['truck_id']}
                                                            choice_azs_truck_dict[n4] = {
                                                                'truck_id': truck_id_number[4]['truck_id']}
                                                            choice_azs_truck_dict[n5] = {
                                                                'truck_id': truck_id_number[5]['truck_id']}
                                                            choice_azs_truck_dict[n6] = {
                                                                'truck_id': truck_id_number[6]['truck_id']}

                                                            number_variant = number_variant + 1
                                                            number_of_success_loops = number_of_success_loops + 1
                                                            good_choices_dict[number_of_success_loops] = {
                                                                'variants': choice_azs_truck_dict}
            if active_trucks == 5:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0

                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            for n3 in final_azs_for_trucks[3]["azs_ids"]:
                                if n3 != n2 and n3 != n1:
                                    for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                        if n4 != n3 and n4 != n2 and n4 != n1:
                                            for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                                if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                    choice_azs_truck_dict = dict()
                                                    choice_azs_truck_dict[n1] = {
                                                        'truck_id': truck_id_number[1][
                                                            'truck_id']}
                                                    choice_azs_truck_dict[n2] = {
                                                        'truck_id': truck_id_number[2][
                                                            'truck_id']}
                                                    choice_azs_truck_dict[n3] = {
                                                        'truck_id': truck_id_number[3][
                                                            'truck_id']}
                                                    choice_azs_truck_dict[n4] = {
                                                        'truck_id': truck_id_number[4][
                                                            'truck_id']}
                                                    choice_azs_truck_dict[n5] = {
                                                        'truck_id': truck_id_number[5][
                                                            'truck_id']}

                                                    number_variant = number_variant + 1
                                                    number_of_success_loops = number_of_success_loops + 1
                                                    good_choices_dict[number_of_success_loops] = {
                                                        'variants': choice_azs_truck_dict}
            if active_trucks == 4:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0
                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            for n3 in final_azs_for_trucks[3]["azs_ids"]:
                                if n3 != n2 and n3 != n1:
                                    for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                        if n4 != n3 and n4 != n2 and n4 != n1:
                                            choice_azs_truck_dict = dict()
                                            choice_azs_truck_dict[n1] = {'truck_id': truck_id_number[1]['truck_id']}
                                            choice_azs_truck_dict[n2] = {'truck_id': truck_id_number[2]['truck_id']}
                                            choice_azs_truck_dict[n3] = {'truck_id': truck_id_number[3]['truck_id']}
                                            choice_azs_truck_dict[n4] = {'truck_id': truck_id_number[4]['truck_id']}
                                            number_variant = number_variant + 1
                                            number_of_success_loops = number_of_success_loops + 1
                                            good_choices_dict[number_of_success_loops] = {
                                                'variants': choice_azs_truck_dict}
            if active_trucks == 3:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0
                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            for n3 in final_azs_for_trucks[3]["azs_ids"]:
                                if n3 != n2 and n3 != n1:
                                    choice_azs_truck_dict = dict()
                                    choice_azs_truck_dict[n1] = {'truck_id': truck_id_number[1]['truck_id']}
                                    choice_azs_truck_dict[n2] = {'truck_id': truck_id_number[2]['truck_id']}
                                    choice_azs_truck_dict[n3] = {'truck_id': truck_id_number[3]['truck_id']}
                                    number_variant = number_variant + 1
                                    number_of_success_loops = number_of_success_loops + 1
                                    good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict}

            if active_trucks == 2:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0
                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    for n2 in final_azs_for_trucks[2]["azs_ids"]:
                        if n2 != n1:
                            choice_azs_truck_dict = dict()
                            choice_azs_truck_dict[n1] = {'truck_id': truck_id_number[1]['truck_id']}
                            choice_azs_truck_dict[n2] = {'truck_id': truck_id_number[2]['truck_id']}
                            number_variant = number_variant + 1
                            number_of_success_loops = number_of_success_loops + 1
                            good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict}

            if active_trucks == 1:
                # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
                choise_good = 1
                number_of_success_loops = 0
                number_variant = 0
                for n1 in final_azs_for_trucks[1]["azs_ids"]:
                    choice_azs_truck_dict = dict()
                    choice_azs_truck_dict[n1] = {'truck_id': truck_id_number[1]['truck_id']}
                    number_variant = number_variant + 1
                    number_of_success_loops = number_of_success_loops + 1
                    good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict}

        if choise_good == 1:
            for choice in good_choices_dict:
                choice_azs_truck_dict = good_choices_dict[choice]['variants']
                '''**************************************************************************************************'''
                # Оцениваем вариант расстановки на предмет не отправки бензовоза на критичные АЗС
                # переменная для хранения оценки текущей расстановки бензовозов (т.е. чем большее количество
                points = 0
                # критичных АЗС пропущено, тем меньше оценка (расстановка хуже)
                for i in choice_azs_truck_dict:  #
                    if choice_azs_truck_dict[i]['truck_id'] != 0:
                        points = points + (1 / (azs_queue_dict[i]['queue'])) * 1000
                # округляем оценку до целого числа
                points = int(points)
                '''**************************************************************************************************'''

                '''**************************************************************************************************'''
                # минимальный запас суток среди всех АЗС
                min_days_stock1_work_type_1 = 1234
                min_days_stock2_work_type_1 = 1234
                # перебираем список расстановки
                list_azs_min_days_stock = list()
                for i in choice_azs_truck_dict:
                    if choice_azs_truck_dict[i]['truck_id'] != 0:
                        list_azs_min_days_stock.append(
                            azs_trucks_best_days_stock[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                'min_rez1'])

                list_azs_min_days_stock.sort()

                min_days_stock1_work_type_1 = list_azs_min_days_stock[0]
                min_days_stock2_work_type_1 = list_azs_min_days_stock[1]
                '''**************************************************************************************************'''

                '''**************************************************************************************************'''
                # оценка количества вывозимого топлива при текущей расстановке
                min_days_stock1_work_type_2 = 1000
                min_days_stock1_work_type_3 = 1000
                sum_max_volume_92 = 0
                sum_max_volume_95 = 0
                sum_max_volume_50 = 0
                sum_min_volume_92 = 0
                sum_min_volume_95 = 0
                sum_min_volume_50 = 0
                for i in choice_azs_truck_dict:
                    if work_type.id == 2 or work_type.id == 1:
                        if choice_azs_truck_dict[i]['truck_id'] != 0:
                            key = str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])
                            # для 92 топлива
                            if fuel_type == 92:
                                sum_max_volume_92 = sum_max_volume_92 + azs_trucks_max_92[key]['max_volume_92']
                                sum_max_volume_95 = sum_max_volume_95 + azs_trucks_max_95[key]['max_volume_95']
                                sum_max_volume_50 = sum_max_volume_50 + azs_trucks_max_50[key]['max_volume_50']

                                if (azs_trucks_max_92[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])]['min_rez1']) \
                                        < min_days_stock1_work_type_2:
                                    min_days_stock1_work_type_2 = \
                                        azs_trucks_max_92[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                            'min_rez1']
                            # для 95 топлива
                            elif fuel_type == 95:
                                sum_max_volume_92 = sum_max_volume_92 + azs_trucks_max_92[key]['max_volume_92']
                                sum_max_volume_95 = sum_max_volume_95 + azs_trucks_max_95[key]['max_volume_95']
                                sum_max_volume_50 = sum_max_volume_50 + azs_trucks_max_50[key]['max_volume_50']

                                if (azs_trucks_max_95[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])]['min_rez1']
                                        < min_days_stock1_work_type_2):
                                    min_days_stock1_work_type_2 = \
                                        azs_trucks_max_95[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                            'min_rez1']
                            # для 50 топлива
                            elif fuel_type == 50:
                                sum_max_volume_92 = sum_max_volume_92 + azs_trucks_max_92[key]['max_volume_92']
                                sum_max_volume_95 = sum_max_volume_95 + azs_trucks_max_95[key]['max_volume_95']
                                sum_max_volume_50 = sum_max_volume_50 + azs_trucks_max_50[key]['max_volume_50']

                                if (azs_trucks_max_50[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])]['min_rez1']
                                        < min_days_stock1_work_type_2):
                                    min_days_stock1_work_type_2 = \
                                        azs_trucks_max_50[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                            'min_rez1']
                        else:
                            min_days_stock1_work_type_2 = 0
                    elif work_type.id == 3:
                        if choice_azs_truck_dict[i]['truck_id'] != 0:
                            key = str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])
                            # для 92 топлива
                            if fuel_type == 92:
                                sum_min_volume_92 = sum_min_volume_92 + azs_trucks_min_92[key]['min_volume_92']
                                sum_min_volume_95 = sum_min_volume_95 + azs_trucks_min_95[key]['min_volume_95']
                                sum_min_volume_50 = sum_min_volume_50 + azs_trucks_min_50[key]['min_volume_50']

                                if (
                                azs_trucks_min_92[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])]['min_rez1']) \
                                        < min_days_stock1_work_type_3:
                                    min_days_stock1_work_type_3 = \
                                        azs_trucks_min_92[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                            'min_rez1']
                            # для 95 топлива
                            elif fuel_type == 95:
                                sum_min_volume_92 = sum_min_volume_92 + azs_trucks_min_92[key]['min_volume_92']
                                sum_min_volume_95 = sum_min_volume_95 + azs_trucks_min_95[key]['min_volume_95']
                                sum_min_volume_50 = sum_min_volume_50 + azs_trucks_min_50[key]['min_volume_50']

                                if (azs_trucks_min_95[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                    'min_rez1']
                                        < min_days_stock1_work_type_3):
                                    min_days_stock1_work_type_3 = \
                                        azs_trucks_min_95[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                            'min_rez1']
                            # для 50 топлива
                            elif fuel_type == 50:
                                sum_min_volume_92 = sum_min_volume_92 + azs_trucks_min_92[key]['min_volume_92']
                                sum_min_volume_95 = sum_min_volume_95 + azs_trucks_min_95[key]['min_volume_95']
                                sum_min_volume_50 = sum_min_volume_50 + azs_trucks_min_50[key]['min_volume_50']

                                if (azs_trucks_min_50[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                    'min_rez1']
                                        < min_days_stock1_work_type_3):
                                    min_days_stock1_work_type_3 = \
                                        azs_trucks_min_50[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                            'min_rez1']
                        else:
                            min_days_stock1_work_type_3 = 0
                '''**************************************************************************************************'''

                '''**************************************************************************************************'''
                # собираем все оцененные варианты расстановки в словарь
                choices_dict_work_type_1[choice] = {'variants': choice_azs_truck_dict,
                                                    'points': points,
                                                    'days_stock_min1': min_days_stock1_work_type_1,
                                                    'days_stock_min2': min_days_stock2_work_type_1
                                                    }
                '''**************************************************************************************************'''
                '''**************************************************************************************************'''
                # собираем все оцененные варианты расстановки в словарь для второго режима работы
                choices_dict_work_type_2[number_of_success_loops] = {'variants': choice_azs_truck_dict,
                                                                     'points': points,
                                                                     'days_stock_min1': min_days_stock1_work_type_2,
                                                                     'max_volume_92': sum_max_volume_92,
                                                                     'max_volume_95': sum_max_volume_95,
                                                                     'max_volume_50': sum_max_volume_50,
                                                                     }
                '''**************************************************************************************************'''
                '''**************************************************************************************************'''
                # собираем все оцененные варианты расстановки в словарь для третьего режима работы
                choices_dict_work_type_3[number_of_success_loops] = {'variants': choice_azs_truck_dict,
                                                                     'points': points,
                                                                     'days_stock_min1': min_days_stock1_work_type_2,
                                                                     'min_volume_92': sum_min_volume_92,
                                                                     'min_volume_95': sum_min_volume_95,
                                                                     'min_volume_50': sum_min_volume_50,
                                                                     }
                '''**************************************************************************************************'''
            # сортируем полученные результаты по трем параметрам
            # На выходе получим отсортированный список ключей словаря choices_dict
            if work_type.id == 1:
                print('Режим работы № 1 . Сортировка')
                sort_choices_dict = sorted(choices_dict_work_type_1,
                                           key=lambda k: (choices_dict_work_type_1[k]['points'],
                                                          choices_dict_work_type_1[k]['days_stock_min1'],
                                                          choices_dict_work_type_1[k]['days_stock_min2']
                                                          ))
            elif work_type.id == 2:
                print('Режим работы № 2 . Сортировка')
                if fuel_type == 92:
                    sort_choices_dict = sorted(choices_dict_work_type_2,
                                               key=lambda k: (choices_dict_work_type_2[k]['points'],
                                                              choices_dict_work_type_2[k]['max_volume_92'],
                                                              choices_dict_work_type_2[k]['days_stock_min1']
                                                              ))
                elif fuel_type == 95:
                    sort_choices_dict = sorted(choices_dict_work_type_2,
                                               key=lambda k: (choices_dict_work_type_2[k]['points'],
                                                              choices_dict_work_type_2[k]['max_volume_95'],
                                                              choices_dict_work_type_2[k]['days_stock_min1']
                                                              ))
                elif fuel_type == 50:
                    sort_choices_dict = sorted(choices_dict_work_type_2,
                                               key=lambda k: (choices_dict_work_type_2[k]['points'],
                                                              choices_dict_work_type_2[k]['max_volume_50'],
                                                              choices_dict_work_type_2[k]['days_stock_min1']
                                                              ))

            elif work_type.id == 3:
                print('Режим работы № 3 . Сортировка')
                if fuel_type == 92:
                    sort_choices_dict = sorted(choices_dict_work_type_3,
                                               key=lambda k: (choices_dict_work_type_3[k]['points'],
                                                              choices_dict_work_type_3[k]['min_volume_92'],
                                                              choices_dict_work_type_3[k]['days_stock_min1']
                                                              ))
                elif fuel_type == 95:
                    sort_choices_dict = sorted(choices_dict_work_type_3,
                                               key=lambda k: (choices_dict_work_type_3[k]['points'],
                                                              choices_dict_work_type_3[k]['min_volume_95'],
                                                              choices_dict_work_type_3[k]['days_stock_min1']
                                                              ))
                elif fuel_type == 50:
                    sort_choices_dict = sorted(choices_dict_work_type_3,
                                               key=lambda k: (choices_dict_work_type_3[k]['points'],
                                                              choices_dict_work_type_3[k]['min_volume_50'],
                                                              choices_dict_work_type_3[k]['days_stock_min1']
                                                              ))
            else:
                alarm = 1

            if work_type.id == 1:
                # создаем список, в котором будем хранить лучшие варианты расстановки
                best_choices = list()
                # берем айди предыдущего варианта расстановки
                trips_last = Trips.query.order_by(desc("calculate_id")).first()

                if not trips_last:
                    previous_variant_id = 0
                else:
                    previous_variant_id = trips_last.calculate_id

                # перебираем отсортированные от худшего к лучшему варианты расстановки
                for i in sort_choices_dict:
                    # каждый вариант добавляем в список
                    best_choices.append(i)
                    # сокращаем список до 1
                    best_choices = sort_choices_dict[-1:]
                # перебираем список из 10 лучших вариантов
                for i in best_choices:
                    trips = Trips(trip_number=1, day=date.today(), date=datetime.today(), work_type_id=work_type.id,
                                  calculate_id=previous_variant_id + 1)
                    db.session.add(trips)
                    print(i, choices_dict_work_type_1[i]['points'], choices_dict_work_type_1[i]['days_stock_min1'],
                          choices_dict_work_type_1[i]['days_stock_min2'])

                    for z in trucks_for_azs_dict:
                        trucks_for_azs = TrucksForAzs(azs_id=z,
                                                      number_of_trucks=len(trucks_for_azs_dict[z]['azs_trucks']) - 1,
                                                      calculate_id=previous_variant_id + 1, trip_number=1)

                        db.session.add(trucks_for_azs)
                    variants_sliva_for_trip = list()
                    for x in choices_dict_work_type_1[i]['variants']:
                        if choices_dict_work_type_1[i]['variants'][x]['truck_id'] != 0:
                            variant_sliv_92 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])][
                                'variant_sliv_92']
                            variant = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])]['variant']
                            truck_id = choices_dict_work_type_1[i]['variants'][x]['truck_id']
                            azs_id = x
                            variant_sliv_95 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])][
                                'variant_sliv_95']
                            variant_sliv_50 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])][
                                'variant_sliv_50']
                            min_rez1 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])]['min_rez1']
                            min_rez2 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])]['min_rez2']
                            min_rez3 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])]['min_rez3']
                            query_variant = TempAzsTrucks4.query.filter_by(azs_id=azs_id, truck_id=truck_id,
                                                                           variant=variant,
                                                                           variant_sliv_92=variant_sliv_92).first()
                            calculate_id = previous_variant_id + 1
                            result = Result(azs_id=azs_id, truck_id=truck_id,
                                            variant=variant,
                                            variant_sliv_92=variant_sliv_92,
                                            variant_sliv_95=variant_sliv_95,
                                            variant_sliv_50=variant_sliv_50,
                                            min_rez1=min_rez1,
                                            min_rez2=min_rez2,
                                            min_rez3=min_rez3,
                                            volume_92=query_variant.sum_92,
                                            volume_95=query_variant.sum_95,
                                            volume_50=query_variant.sum_50,
                                            calculate_id=calculate_id,
                                            trip_number=1)
                            db.session.add(result)

                            print("АЗС:", azs_id,
                                  "Бензовоз:", truck_id,
                                  "Вариант налива:", variant,
                                  "Вариант слива 92:", variant_sliv_92,
                                  "Вариант слива 95:", variant_sliv_95,
                                  "Вариант слива 50:", variant_sliv_50)

                            variant_sliva = dict()

                            table_azs_trucks1 = TempAzsTrucks.query.filter_by(variant_id=int(variant)).all()
                            table_azs_trucks3_92 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                                  variant_sliv=variant_sliv_92).all()
                            table_azs_trucks3_95 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                                  variant_sliv=variant_sliv_95).all()
                            table_azs_trucks3_50 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                                  variant_sliv=variant_sliv_50).all()

                            for table_variant in table_azs_trucks1:
                                variant_naliva = VariantNalivaForTrip(variant_from_table=int(variant),
                                                                      calculate_id=calculate_id,
                                                                      truck_tank_id=table_variant.truck_tank_id,
                                                                      truck_id=truck_id,
                                                                      azs_id=azs_id,
                                                                      fuel_type=table_variant.fuel_type,
                                                                      capacity=table_variant.capacity,
                                                                      trip_number=1)
                                db.session.add(variant_naliva)

                            for row in table_azs_trucks3_92:
                                variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                    calculate_id=calculate_id,
                                                                    azs_id=row.azs_id,
                                                                    tank_id=row.tank_id,
                                                                    truck_id=row.truck_id,
                                                                    truck_tank_id=row.truck_tank_id_string,
                                                                    fuel_type=row.fuel_type,
                                                                    capacity=row.sum_sliv,
                                                                    trip_number=1)
                                db.session.add(variant_sliva)

                            for row in table_azs_trucks3_95:
                                variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                    calculate_id=calculate_id,
                                                                    azs_id=row.azs_id,
                                                                    tank_id=row.tank_id,
                                                                    truck_id=row.truck_id,
                                                                    truck_tank_id=row.truck_tank_id_string,
                                                                    fuel_type=row.fuel_type,
                                                                    capacity=row.sum_sliv,
                                                                    trip_number=1)
                                db.session.add(variant_sliva)

                            for row in table_azs_trucks3_50:
                                variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                    calculate_id=calculate_id,
                                                                    azs_id=row.azs_id,
                                                                    tank_id=row.tank_id,
                                                                    truck_id=row.truck_id,
                                                                    truck_tank_id=row.truck_tank_id_string,
                                                                    fuel_type=row.fuel_type,
                                                                    capacity=row.sum_sliv,
                                                                    trip_number=1)
                                db.session.add(variant_sliva)
                db.session.commit()

            elif work_type.id == 2:
                # создаем список, в котором будем хранить лучшие варианты расстановки
                best_choices = list()
                # перебираем отсортированные от худшего к лучшему варианты расстановки
                for i in sort_choices_dict:
                    # каждый вариант добавляем в список
                    best_choices.append(i)
                    # сокращаем список до 1
                    best_choices = sort_choices_dict[-1:]

                # берем айди предыдущего варианта расстановки
                trips_last = Trips.query.order_by(desc("calculate_id")).first()
                previous_variant_id = trips_last.calculate_id
                if not trips_last:
                    previous_variant_id = 0
                else:
                    previous_variant_id = trips_last.calculate_id
                results = Result.query.all()
                # перебираем список из 10 лучших вариантов
                for i in best_choices:
                    trips = Trips(trip_number=1, day=date.today(), date=datetime.today(), work_type_id=work_type.id,
                                  calculate_id=previous_variant_id + 1)

                    for x in choices_dict_work_type_2[i]['variants']:
                        if choices_dict_work_type_1[i]['variants'][x]['truck_id'] != 0:
                            if fuel_type == 92:
                                key = str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])
                                print("АЗС:", x, "бензовоз:",
                                      choices_dict_work_type_2[i]['variants'][x]['truck_id'],
                                      'variants:', azs_trucks_max_92[key]['variant'],
                                      'sliv_92:', azs_trucks_max_92[key]['variant_sliv_92'],
                                      'sliv_95:', azs_trucks_max_92[key]['variant_sliv_95'],
                                      'sliv_50:', azs_trucks_max_92[key]['variant_sliv_50'],
                                      'max_volume_92:', azs_trucks_max_92[key]['max_volume_92'],
                                      'max_volume_95:', azs_trucks_max_92[key]['max_volume_95'],
                                      'max_volume_50:', azs_trucks_max_92[key]['max_volume_50'],
                                      )
                            elif fuel_type == 95:
                                print("АЗС:", x, "бензовоз:",
                                      choices_dict_work_type_2[i]['variants'][x]['truck_id'],
                                      'variants:', azs_trucks_max_95[
                                          str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                                          'variant'],
                                      'sliv_92:', azs_trucks_max_95[
                                          str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                                          'variant_sliv_92'],
                                      'sliv_95:', azs_trucks_max_95[
                                          str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                                          'variant_sliv_95'],
                                      'sliv_50:', azs_trucks_max_95[
                                          str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                                          'variant_sliv_50'])
                            elif fuel_type == 50:
                                print("АЗС:", x, "бензовоз:",
                                      choices_dict_work_type_2[i]['variants'][x]['truck_id'],
                                      'variants:', azs_trucks_max_50[
                                          str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                                          'variant'],
                                      'sliv_92:', azs_trucks_max_50[
                                          str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                                          'variant_sliv_92'],
                                      'sliv_95:', azs_trucks_max_50[
                                          str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                                          'variant_sliv_95'],
                                      'sliv_50:', azs_trucks_max_50[
                                          str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                                          'variant_sliv_50'])

            elif work_type.id == 3:
                # создаем список, в котором будем хранить лучшие варианты расстановки
                best_choices = list()
                # берем айди предыдущего варианта расстановки
                trips_last = Trips.query.order_by(desc("calculate_id")).first()

                if not trips_last:
                    previous_variant_id = 0
                else:
                    previous_variant_id = trips_last.calculate_id

                # перебираем отсортированные от худшего к лучшему варианты расстановки
                for i in sort_choices_dict:
                    # каждый вариант добавляем в список
                    best_choices.append(i)
                    # сокращаем список до 1
                    best_choices = sort_choices_dict[-1:]
                # перебираем список из 10 лучших вариантов
                for i in best_choices:
                    trips = Trips(trip_number=1, day=date.today(), date=datetime.today(), work_type_id=work_type.id,
                                  calculate_id=previous_variant_id + 1)
                    db.session.add(trips)
                    print(i, choices_dict_work_type_1[i]['points'], choices_dict_work_type_1[i]['days_stock_min1'],
                          choices_dict_work_type_1[i]['days_stock_min2'])

                    for z in trucks_for_azs_dict:
                        trucks_for_azs = TrucksForAzs(azs_id=z,
                                                      number_of_trucks=len(trucks_for_azs_dict[z]['azs_trucks']) - 1,
                                                      calculate_id=previous_variant_id + 1, trip_number=1)

                        db.session.add(trucks_for_azs)
                    variants_sliva_for_trip = list()
                    for x in choices_dict_work_type_3[i]['variants']:
                        if choices_dict_work_type_3[i]['variants'][x]['truck_id'] != 0:
                            variant_sliv_92 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])][
                                'variant_sliv_92']
                            variant = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])]['variant']
                            truck_id = choices_dict_work_type_3[i]['variants'][x]['truck_id']
                            azs_id = x
                            variant_sliv_95 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])][
                                'variant_sliv_95']
                            variant_sliv_50 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])][
                                'variant_sliv_50']
                            min_rez1 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])]['min_rez1']
                            min_rez2 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])]['min_rez2']
                            min_rez3 = azs_trucks_best_days_stock[
                                str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])]['min_rez3']
                            query_variant = TempAzsTrucks4.query.filter_by(azs_id=azs_id, truck_id=truck_id,
                                                                           variant=variant,
                                                                           variant_sliv_92=variant_sliv_92).first()
                            calculate_id = previous_variant_id + 1
                            result = Result(azs_id=azs_id, truck_id=truck_id,
                                            variant=variant,
                                            variant_sliv_92=variant_sliv_92,
                                            variant_sliv_95=variant_sliv_95,
                                            variant_sliv_50=variant_sliv_50,
                                            min_rez1=min_rez1,
                                            min_rez2=min_rez2,
                                            min_rez3=min_rez3,
                                            volume_92=query_variant.sum_92,
                                            volume_95=query_variant.sum_95,
                                            volume_50=query_variant.sum_50,
                                            calculate_id=calculate_id,
                                            trip_number=1)
                            db.session.add(result)

                            print("АЗС:", azs_id,
                                  "Бензовоз:", truck_id,
                                  "Вариант налива:", variant,
                                  "Вариант слива 92:", variant_sliv_92,
                                  "Вариант слива 95:", variant_sliv_95,
                                  "Вариант слива 50:", variant_sliv_50)

                            variant_sliva = dict()

                            table_azs_trucks1 = TempAzsTrucks.query.filter_by(variant_id=int(variant)).all()
                            table_azs_trucks3_92 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                                  variant_sliv=variant_sliv_92).all()
                            table_azs_trucks3_95 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                                  variant_sliv=variant_sliv_95).all()
                            table_azs_trucks3_50 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                                  variant_sliv=variant_sliv_50).all()

                            for table_variant in table_azs_trucks1:
                                variant_naliva = VariantNalivaForTrip(variant_from_table=int(variant),
                                                                      calculate_id=calculate_id,
                                                                      truck_tank_id=table_variant.truck_tank_id,
                                                                      truck_id=truck_id,
                                                                      azs_id=azs_id,
                                                                      fuel_type=table_variant.fuel_type,
                                                                      capacity=table_variant.capacity,
                                                                      trip_number=1)
                                db.session.add(variant_naliva)

                            for row in table_azs_trucks3_92:
                                variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                    calculate_id=calculate_id,
                                                                    azs_id=row.azs_id,
                                                                    tank_id=row.tank_id,
                                                                    truck_id=row.truck_id,
                                                                    truck_tank_id=row.truck_tank_id_string,
                                                                    fuel_type=row.fuel_type,
                                                                    capacity=row.sum_sliv,
                                                                    trip_number=1)
                                db.session.add(variant_sliva)

                            for row in table_azs_trucks3_95:
                                variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                    calculate_id=calculate_id,
                                                                    azs_id=row.azs_id,
                                                                    tank_id=row.tank_id,
                                                                    truck_id=row.truck_id,
                                                                    truck_tank_id=row.truck_tank_id_string,
                                                                    fuel_type=row.fuel_type,
                                                                    capacity=row.sum_sliv,
                                                                    trip_number=1)
                                db.session.add(variant_sliva)

                            for row in table_azs_trucks3_50:
                                variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                    calculate_id=calculate_id,
                                                                    azs_id=row.azs_id,
                                                                    tank_id=row.tank_id,
                                                                    truck_id=row.truck_id,
                                                                    truck_tank_id=row.truck_tank_id_string,
                                                                    fuel_type=row.fuel_type,
                                                                    capacity=row.sum_sliv,
                                                                    trip_number=1)
                                db.session.add(variant_sliva)
                db.session.commit()
                print('ASFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFFF')
            logs = UserLogs(user_id=current_user.id,
                            action="first_trip_creation_success",
                            timestamp=datetime.now())
            db.session.add(logs)
            db.session.commit()
            result = True
        else:
            result = False
            logs = UserLogs(user_id=current_user.id,
                            action="trip_creation_failed",
                            timestamp=datetime.now())
            db.session.add(logs)
            db.session.commit()
            return result

    if create_trip() == False:
        return redirect(url_for('main.creation_failed'))

    def time_to_return():
        # создаем словарь в котором будут сопоставлены БЕНЗОВОЗ-АЗС первого рейса
        trucks_for_azs_first_trip = dict()
        # получаем информацию о первом рейсе
        first_trip = Trips.query.filter_by(trip_number=1).order_by(desc("calculate_id")).first()
        # получаем информацию о расстановке первого рейса
        first_trip_list = Result.query.filter_by(calculate_id=first_trip.calculate_id).all()
        trip = Trip.query.all()
        trip_dict = dict()
        for i in trip:
            trip_dict[i.azs_id] = {'time_to_before_lunch': i.time_to_before_lunch,
                                   'time_from_before_lunch': i.time_from_before_lunch}

        # заполняем словарь с сопоставлением БЕНЗОВОЗ-АЗС из первого рейса
        for i in first_trip_list:
            full_time = trip_dict[i.azs_id]['time_to_before_lunch'] + trip_dict[i.azs_id][
                'time_from_before_lunch'] + 60 + 120
            trucks_for_azs_first_trip[i.truck_id] = {'full_time_first_trip': full_time}
            trip_start_time = Trucks.query.filter_by(id=i.truck_id).first_or_404()
            t = trip_start_time.day_start
            delta = timedelta(minutes=full_time)
            trip_end = (datetime.combine(date(1, 1, 1), t) + delta).time()
            result = Result.query.filter_by(calculate_id=first_trip.calculate_id, azs_id=i.azs_id,
                                            truck_id=i.truck_id).first()
            result.time_to_return = full_time
            result.trip_end_time = trip_end
            db.session.commit()
        return trucks_for_azs_first_trip

    def is_it_fit_second():
        # получаем информацию о первом рейсе
        first_trip = Trips.query.filter_by(trip_number=1).order_by(desc("calculate_id")).first()
        # получаем информацию о расстановке первого рейса
        first_trip_list = Result.query.filter_by(calculate_id=first_trip.calculate_id).all()
        # получаем вторую таблицу
        temp_azs_trucks_2_list = global_temp_azs_trucks_2_list
        temp_azs_trucks = global_temp_azs_trucks
        first_trip = time_to_return()
        # получаем остатки из базы
        residue = FuelResidue.query.all()
        # получаем реализацию из базы
        realisation = FuelRealisation.query.all()
        # получаем информацию о времени до азс
        trip = Trip.query.all()
        # получаем список отсеков всех бензовозов
        truck_tanks = TruckTanks.query.all()
        # получаем информанию об азс, на которые не могут заехать определенные бензовозы
        trucks_false = TruckFalse.query.all()
        # создаем пустой словарь для хранеия в нем данных о реализации и остатков
        realisation_n_residue = dict()
        # создаем пустой словарь для хранеия в нем данных времени пути до АЗС
        azs_trip_access = dict()
        azs_trip_time = dict()
        # словарь со временеи окончания рабочего дня водителей бензовозов
        driver_work_time = dict()
        # заполняем словарь с данными об остатках топлива
        for i in residue:
            realisation_n_residue[i.tank_id] = {'azs_id': i.azs_id,  # id_АЗС
                                                'fuel_volume': i.fuel_volume,  # остаток топлива в резервуаре
                                                'free_volume': i.free_volume,  # свободная емкость в резервуаре
                                                # реализация в резервуаре (нулевая, так как заполняем ее потом)
                                                'fuel_realisation': 0,
                                                # максимальная реализация топлива в резервуаре (среди всех периодов)
                                                'fuel_realisation_max': 0
                                                }
        # заполняем словарь данными о реализации топлива
        for i in realisation:
            realisation_n_residue[i.tank_id] = {'azs_id': i.azs_id,  # id_АЗС
                                                # Оставляем прошлое значение остатка топлива в резервуаре
                                                'fuel_volume': realisation_n_residue[i.tank_id]['fuel_volume'],
                                                # свободная емкость в резервуаре
                                                'free_volume': realisation_n_residue[i.tank_id]['free_volume'],
                                                # Добавляем в словарь реализацию из этого резервуара
                                                # (среднюю в час за шестичасовой период)
                                                'fuel_realisation': i.fuel_realisation_hour / 6,
                                                # максимальная реализация топлива в резервуаре (среди всех периодов)
                                                'fuel_realisation_max': i.fuel_realisation_max
                                                }

        # формируем словарь для хранения данных о растоянии до АЗС и времени пути
        for i in Trucks.query.all():
            driver_work_time[i.id] = {'work_time_end': i.day_end}
        for i in trip:
            azs_trip_time[i.azs_id] = {'time_to_before_lunch': i.time_to_before_lunch,
                                       'time_to': i.time_to,
                                       'time_from': i.time_from,
                                       'time_from_before_lunch': i.time_from_before_lunch,
                                       'weigher': i.weigher
                                       }
        # формируем словарь для хранения данных о бензовозах и азс на которые они не могут заезжать
        for i in trucks_false:
            azs_trip_access[str(i.azs_id) + '-' + str(i.truck_id)] = {'access': False}

        first_trip_end_time = dict()
        # заполняем словарь с ключем truck_id данными с временем окончания первого рейса
        for i in first_trip_list:
            first_trip_end_time[i.truck_id] = {'trip_end_time': i.trip_end_time}

        is_it_fit_list = list()

        for i in temp_azs_trucks_2_list:
            # проверяем, сольется ли бензовоз в данный момент
            # из свободной емкости резервуара вычитаем сумму слива бензовоза
            sliv = realisation_n_residue[i['tank_id']]['free_volume'] - i['sum_sliv']
            # получаем время до азс до обеда (в минутах)
            time_to = azs_trip_time[i['azs_id']]['time_to_before_lunch']
            # получаем время, которое бензовоз затратил на первый рейс
            full_time_first_trip = first_trip[i['truck_id']]['full_time_first_trip']
            # считаем примерное количество топлива, которое будет реализовано (за время первого рейса + пусть до азс
            # во время второго рейса)
            realis_time = realisation_n_residue[i["tank_id"]]['fuel_realisation'] * (
                        (time_to + full_time_first_trip) / 60)
            # проверяем сольется ли бензовоз, с учетом реализации за время его пути к АЗС
            # из свободной емкости резервуара вычитаем сумму слива бензовоза, и прибавляем количество топлива,
            # которое реализуется у данного резервуара за время пути бензовоза к ней
            sliv_later = realisation_n_residue[i['tank_id']]['free_volume'] - i['sum_sliv'] + realis_time

            # считаем время возвращения бензовоза со второго рейса
            full_time = azs_trip_time[i['azs_id']]['time_to'] + azs_trip_time[i['azs_id']]['time_from'] + 60
            t = first_trip_end_time[i['truck_id']]['trip_end_time']
            delta = timedelta(minutes=full_time)
            trip_end = (datetime.combine(date(1, 1, 1), t) + delta).time()

            if driver_work_time[i['truck_id']]['work_time_end'] < trip_end:
                i['is_trip_end_time_good'] = False
            else:
                i['is_trip_end_time_good'] = True
            # если бензовоз не сливается в данный момент (то есть переменная sliv - меньше нуля)
            if sliv < 0:
                # записываем в базу, что бензовоз в данный момент слиться не сможет
                i['is_it_fit'] = False
                # новый запас суток и новые остатки не считаем
                i['second_new_fuel_volume'] = 0
                i['second_new_days_stock'] = 0
            # если бензовоз сможет слиться нат екущий момент (то есть переменная sliv - больше нуля)
            else:
                # записываем в базу, что бензовоз в данный момент сольется
                i['is_it_fit'] = True
                # расчитываем количество отстатков в резервуаре после слива
                i['second_new_fuel_volume'] = realisation_n_residue[i['tank_id']]['fuel_volume'] - realis_time + i[
                    'sum_sliv']
                # расчитываем новый запас суток
                i['second_new_days_stock'] = i['second_new_fuel_volume'] / realisation_n_residue[i['tank_id']][
                    'fuel_realisation_max']
            # если бензовоз не сливается после времени затраченного на дорогу (то есть переменная sliv_later
            # - меньше нуля)
            if sliv_later < 0:
                # записываем в базу, что бензовоз слиться не сможет
                i['is_it_fit_on_second_trip'] = False
                # новый запас суток и новые остатки не считаем
                i['second_new_fuel_volume'] = 0
                i['second_new_days_stock'] = 0
            else:
                # записываем в базу, что бензовоз сольется спустя время затраченное на дорогу
                i['is_it_fit_on_second_trip'] = True
                # расчитываем количество отстатков в резервуаре после слива
                i['second_new_fuel_volume'] = realisation_n_residue[i['tank_id']]['fuel_volume'] - realis_time + i[
                    'sum_sliv']
                # расчитываем новый запас суток
                i['second_new_days_stock'] = i['second_new_fuel_volume'] / realisation_n_residue[i['tank_id']][
                    'fuel_realisation_max']
            # проверяем, сможет ли бензовоз заехать на АЗС (нет ли для него никаких ограничений)
            # т.е. проверяем наличие ключа в словаре azs_trip_access, в котором содержатся ограничения для бензовозов
            # ключ АЗС-АйдиБензовоза
            if str(i['azs_id']) + '-' + str(i['truck_id']) in azs_trip_access:
                # если ограничения есть, то ставим False
                i['is_it_able_to_enter'] = False
            else:
                # если ограничений нет, то ставим True
                i['is_it_able_to_enter'] = True

            # добавляем словарь в список для записи в базу данных, в таблицу TempAzsTrucks2
            is_it_fit_list.append(i)
        # создаем словарь для хранения переменных с информацией о том, влезет ли определенный вид топлива
        # в резервуар на АЗС (3 переменные по всем видам топлива)
        fuel_types_dict = dict()
        # создаем список  для хранеия обновленной таблицы TempAzsTrucks2
        is_variant_sliv_good_list = list()
        # заполняем переменные словаря fuel_types_dict единицами
        # ключем в словаре является связка variant:variant_sliv
        for i in is_it_fit_list:
            fuel_types_dict[str(i['variant']) + ':' + str(i['variant_sliv'])] = {'is_it_fit_92': 1,
                                                                                 'is_it_fit_95': 1,
                                                                                 'is_it_fit_50': 1}
        # заново перебираем таблицу TempAzsTrucks2(которая хранится в виде списка словарей)
        for i in is_it_fit_list:
            # в переменную key заносим ключ итого словаря (связка variant:variant_sliv)
            key = str(i['variant']) + ':' + str(i['variant_sliv'])
            # если находим вид топлива которое не сливается, то помечаем его нулем
            if i['fuel_type'] == 92 and i['is_it_fit_later'] == 0:
                fuel_types_dict[key] = {'is_it_fit_92': 0,
                                        'is_it_fit_95': fuel_types_dict[key]['is_it_fit_95'],
                                        'is_it_fit_50': fuel_types_dict[key]['is_it_fit_50']}
            if i['fuel_type'] == 95 and i['is_it_fit_later'] == 0:
                fuel_types_dict[key] = {'is_it_fit_92': fuel_types_dict[key]['is_it_fit_92'],
                                        'is_it_fit_95': 0,
                                        'is_it_fit_50': fuel_types_dict[key]['is_it_fit_50']}
            if i['fuel_type'] == 50 and i['is_it_fit_later'] == 0:
                fuel_types_dict[key] = {'is_it_fit_92': fuel_types_dict[key]['is_it_fit_92'],
                                        'is_it_fit_95': fuel_types_dict[key]['is_it_fit_95'],
                                        'is_it_fit_50': 0}
        # снова перебираем список словарей с данными из таблицы TempAzsTrucks2
        for i in is_it_fit_list:
            # в переменную key заносим ключ итого словаря (связка variant:variant_sliv)
            key = str(i['variant']) + ':' + str(i['variant_sliv'])
            # если все три вида топлива (или все виды топлива которые мы везем на азс) сливаются, то помечаем столбец
            # в котором хранится информация о том, сливается ли данный вариант (is_variant_sliv_good) единицей
            if fuel_types_dict[key]['is_it_fit_92'] == 1 and fuel_types_dict[key]['is_it_fit_95'] == 1 and \
                    fuel_types_dict[key]['is_it_fit_50'] == 1:
                i['is_variant_sliv_good'] = 1
                # добавляем обновленный словарь в список
                is_variant_sliv_good_list.append(i)
            # если не все топливо из данного варианта слива сливается, то ставим в ячейке ноль (False)
            else:
                i['is_variant_sliv_good'] = 0
                # добавляем обновленный словарь в список
                is_variant_sliv_good_list.append(i)
        # создаем словарь для хранения обновленных данных (по факту - в словаре появится ячейка is_variant_good)
        is_variant_good_list = dict()
        # перебираем список словарей с данными из таблицы TempAzsTrucks2 (из предыдущего цикла)
        for i in is_variant_sliv_good_list:
            # в созданные ранее словарь добавляем ячейки с нулями
            is_variant_good_list[str(i['variant'])] = {'is_it_92': 0,
                                                       'is_it_95': 0,
                                                       'is_it_50': 0}
        # снова перебираем список словарей
        for i in is_variant_sliv_good_list:
            # если находим определенный вид топлива, и вариант слива относящийся к этой строке таблицы отмечен True,
            # то помечаем го цифрой 2
            # а если нет, то единицей
            if i['fuel_type'] == 92 and i['is_variant_sliv_good'] == 1:
                is_variant_good_list[str(i['variant'])] = {'is_it_92': 2,
                                                           'is_it_95': is_variant_good_list[str(i['variant'])][
                                                               'is_it_95'],
                                                           'is_it_50': is_variant_good_list[str(i['variant'])][
                                                               'is_it_50']}
            if i['fuel_type'] == 92 and i['is_variant_sliv_good'] == 0 and is_variant_good_list[str(i['variant'])][
                'is_it_92'] != 2:
                is_variant_good_list[str(i['variant'])] = {'is_it_92': 1,
                                                           'is_it_95': is_variant_good_list[str(i['variant'])][
                                                               'is_it_95'],
                                                           'is_it_50': is_variant_good_list[str(i['variant'])][
                                                               'is_it_50']}
            if i['fuel_type'] == 95 and i['is_variant_sliv_good'] == 1:
                is_variant_good_list[str(i['variant'])] = {
                    'is_it_92': is_variant_good_list[str(i['variant'])]['is_it_92'],
                    'is_it_95': 2,
                    'is_it_50': is_variant_good_list[str(i['variant'])]['is_it_50']}
            if i['fuel_type'] == 95 and i['is_variant_sliv_good'] == 0 and is_variant_good_list[str(i['variant'])][
                'is_it_95'] != 2:
                is_variant_good_list[str(i['variant'])] = {
                    'is_it_92': is_variant_good_list[str(i['variant'])]['is_it_92'],
                    'is_it_95': 1,
                    'is_it_50': is_variant_good_list[str(i['variant'])]['is_it_50']}
            if i['fuel_type'] == 50 and i['is_variant_sliv_good'] == 1:
                is_variant_good_list[str(i['variant'])] = {
                    'is_it_92': is_variant_good_list[str(i['variant'])]['is_it_92'],
                    'is_it_95': is_variant_good_list[str(i['variant'])]['is_it_95'],
                    'is_it_50': 2}
            if i['fuel_type'] == 50 and i['is_variant_sliv_good'] == 0 and is_variant_good_list[str(i['variant'])][
                'is_it_50'] != 2:
                is_variant_good_list[str(i['variant'])] = {
                    'is_it_92': is_variant_good_list[str(i['variant'])]['is_it_92'],
                    'is_it_95': is_variant_good_list[str(i['variant'])]['is_it_95'],
                    'is_it_50': 1}
        # создаем финальный список для данной функции, который будет записан в таблицу TempAzsTrucks2 в БД
        final_list2 = list()
        final_list = list()
        test_dict = dict()
        weigher_dict = dict()  # словарь с наполнением при весах
        weigher_variant_good_dict = dict()
        trip = Trip.query.filter_by(weigher=1).all()
        cells = TruckTanks.query.all()
        truck_tanks_variations = TruckTanksVariations.query.all()
        azs_list_weigher_variant_id = list()
        azs_list_weigher_truck_id = list()
        azs_list_weigher_fuel_type = list()
        azs_list_weigher_truck_tank_id = list()
        for i in trip:
            for azs in temp_azs_trucks:
                if azs['azs_id'] == i.azs_id:
                    azs_list_weigher_variant_id.append(azs['variant_id'])
                    azs_list_weigher_truck_id.append(azs['truck_id'])
            for index, x in enumerate(azs_list_weigher_variant_id):
                key = str(azs_list_weigher_variant_id[index]) + ':' + str(azs_list_weigher_truck_id[index])
                test_dict[key] = {'1': None,
                                  '2': None,
                                  '3': None,
                                  '4': None,
                                  '5': None,
                                  '6': None,
                                  '7': None,
                                  '8': None
                                  }

        azs_list_weigher_variant_id = list()
        azs_list_weigher_truck_id = list()
        azs_list_weigher_fuel_type = list()
        azs_list_weigher_truck_tank_id = list()
        for i in trip:
            for azs in temp_azs_trucks:
                if azs['azs_id'] == i.azs_id:
                    azs_list_weigher_variant_id.append(azs['variant_id'])
                    azs_list_weigher_truck_id.append(azs['truck_id'])
                    azs_list_weigher_fuel_type.append(azs['fuel_type'])
                    azs_list_weigher_truck_tank_id.append(azs['truck_tank_id'])
            for index, x in enumerate(azs_list_weigher_variant_id):
                key = str(azs_list_weigher_variant_id[index]) + ':' + str(azs_list_weigher_truck_id[index])

                fuel_type_boolean = azs_list_weigher_fuel_type[index]
                for cell in cells:
                    if cell.id == azs_list_weigher_truck_tank_id[index]:
                        if cell.number == 1:
                            test_dict[key] = {'1': fuel_type_boolean,
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 2:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': fuel_type_boolean,
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 3:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': fuel_type_boolean,
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 4:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': fuel_type_boolean,
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 5:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': fuel_type_boolean,
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 6:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': fuel_type_boolean,
                                              '7': test_dict[key]['7'],
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 7:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['7'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': fuel_type_boolean,
                                              '8': test_dict[key]['8']
                                              }
                        if cell.number == 8:
                            test_dict[key] = {'1': test_dict[key]['1'],
                                              '2': test_dict[key]['2'],
                                              '3': test_dict[key]['3'],
                                              '4': test_dict[key]['4'],
                                              '5': test_dict[key]['5'],
                                              '6': test_dict[key]['6'],
                                              '7': test_dict[key]['7'],
                                              '8': fuel_type_boolean
                                              }

        for cell in truck_tanks_variations:
            weigher_dict[str(cell.truck_id) + ":" + str(cell.variant_good)] = {'1': None,
                                                                               '2': None,
                                                                               '3': None,
                                                                               '4': None,
                                                                               '5': None,
                                                                               '6': None,
                                                                               '7': None,
                                                                               '8': None
                                                                               }

            weigher_variant_good_dict[cell.truck_id] = {'variant_good': []}

        for cell in truck_tanks_variations:
            if cell.variant_good not in weigher_variant_good_dict[cell.truck_id]['variant_good']:
                temp_list = list()
                temp_list.append(cell.variant_good)
                weigher_variant_good_dict[cell.truck_id] = {'variant_good':
                                                                weigher_variant_good_dict[cell.truck_id][
                                                                    'variant_good'] + temp_list
                                                            }

            for truck_cell in cells:
                if truck_cell.id == cell.truck_tank_id:
                    number = truck_cell.number

            key = str(cell.truck_id) + ':' + str(cell.variant_good)
            if number == 1:
                weigher_dict[key] = {'1': cell.diesel,
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 2:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': cell.diesel,
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 3:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': cell.diesel,
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 4:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': cell.diesel,
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 5:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': cell.diesel,
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 6:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': cell.diesel,
                                     '7': weigher_dict[key]['7'],
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 7:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': cell.diesel,
                                     '8': weigher_dict[key]['8']
                                     }

            if number == 8:
                weigher_dict[key] = {'1': weigher_dict[key]['1'],
                                     '2': weigher_dict[key]['2'],
                                     '3': weigher_dict[key]['3'],
                                     '4': weigher_dict[key]['4'],
                                     '5': weigher_dict[key]['5'],
                                     '6': weigher_dict[key]['6'],
                                     '7': weigher_dict[key]['7'],
                                     '8': cell.diesel
                                     }

        # перебираем список словарей
        for i in is_variant_sliv_good_list:
            # если все виды топлива данного варианта сливаются, то ячейку is_variant_good в таблице записываем True,
            # если же нет, то в ячейку is_variant_good пишем False
            if is_variant_good_list[str(i['variant'])]['is_it_92'] != 1 \
                    and is_variant_good_list[str(i['variant'])]['is_it_95'] != 1 \
                    and is_variant_good_list[str(i['variant'])]['is_it_50'] != 1:
                i['is_variant_good'] = True
                # добавляем получившийся словарь в список
                final_list.append(i)
            else:
                i['is_variant_good'] = False
                # добавляем получившийся словарь в список
                final_list.append(i)

        is_variant_weighter_good = list()
        is_variant_weighter_not_good = list()
        for i in final_list:
            key = str(i['variant']) + ':' + str(i['truck_id'])
            if key in test_dict:
                trig_final = 0  # Изначально считаем, что бензовоз нельзя везти если есть весы
                cells_list = list()

                cell_1 = test_dict[key]['1']
                cells_list.append(cell_1)
                cell_2 = test_dict[key]['2']
                cells_list.append(cell_2)
                cell_3 = test_dict[key]['3']
                cells_list.append(cell_3)
                cell_4 = test_dict[key]['4']
                cells_list.append(cell_4)
                cell_5 = test_dict[key]['5']
                cells_list.append(cell_5)
                cell_6 = test_dict[key]['6']
                cells_list.append(cell_6)
                cell_7 = test_dict[key]['7']
                cells_list.append(cell_7)
                cell_8 = test_dict[key]['8']
                cells_list.append(cell_8)

                for variant_good in weigher_variant_good_dict[i['truck_id']]['variant_good']:
                    variant_key = str(i['truck_id']) + ":" + str(variant_good)
                    weigher_list = list()

                    weigher_list.append(weigher_dict[variant_key]['1'])
                    weigher_list.append(weigher_dict[variant_key]['2'])
                    weigher_list.append(weigher_dict[variant_key]['3'])
                    weigher_list.append(weigher_dict[variant_key]['4'])
                    weigher_list.append(weigher_dict[variant_key]['5'])
                    weigher_list.append(weigher_dict[variant_key]['6'])
                    weigher_list.append(weigher_dict[variant_key]['7'])
                    weigher_list.append(weigher_dict[variant_key]['8'])

                    trig = 1  # Можно завозить дизель для этой комбинации налива
                    for index, cell in enumerate(cells_list):
                        if cells_list[index] == 50 and (weigher_list[index] == 0 or weigher_list[index] == None):
                            trig = 0  # Нельзя завозить дизель для этой комбинации налива
                            break

                    if trig == 1:  # Если можнно завозить дизель для одной их всевозможных комбинаций налива,
                        trig_final = 1  # то значит  можно завозить

                if trig_final == 1:  # Значит вариант подходит (дизель можно везти)!
                    is_variant_weighter_good.append(i['variant'])
                else:
                    is_variant_weighter_not_good.append(i['variant'])

        for i in final_list:
            if i['variant'] in is_variant_weighter_not_good:
                i['is_variant_weigher_good'] = False
            else:
                i['is_variant_weigher_good'] = True
            final_list2.append(i)
        # записываем данные из списка в базу
        db.engine.execute(TempAzsTrucks2SecondTrip.__table__.insert(), final_list2)
        return final_list2

    ''' функция отсеивает все варианты из таблицы TempAzsTrucks2 и дает им оценку ДЛЯ ВТОРОГО РЕЙСА'''

    def preparation_four_second_trip():
        final_list = is_it_fit_second()
        # очищаем таблицу
        db.engine.execute("TRUNCATE TABLE `temp_azs_trucks3_second_trip`")
        db.engine.execute("TRUNCATE TABLE `temp_azs_trucks4_second_trip`")
        temp_azs_trucks3_list = list()
        fuel_realisation = FuelRealisation.query.all()
        days_stock_dict = dict()

        # перебираем список из предыдущей функции
        for i in final_list:
            # если вариант сливается, бензовоз может заехать на АЗС и бензовоз сливается полностью,
            # и бензовоз проходит через весы
            if i['is_it_fit_on_second_trip'] == True and i['is_it_able_to_enter'] == True and i[
                'is_variant_good'] == True \
                    and i['is_variant_sliv_good'] == True and i['is_variant_weigher_good'] == True and i[
                'is_trip_end_time_good'] == True:
                # добавляем словарь в список
                temp_azs_trucks3_list.append(i)
        new_days_stock_dict = dict()
        temp_azs_trucks4_dict = dict()

        for i in temp_azs_trucks3_list:
            temp_azs_trucks4_dict[str(i['variant'])] = {'variant_sliv_92': [],
                                                        'variant_sliv_95': [],
                                                        'variant_sliv_50': [],
                                                        'volume_92': [],
                                                        'volume_95': [],
                                                        'volume_50': [],
                                                        'azs_id': i['azs_id'],
                                                        'truck_id': i['truck_id']
                                                        }
        for i in temp_azs_trucks3_list:
            variant_sliv_92 = list()
            variant_sliv_95 = list()
            variant_sliv_50 = list()
            volume_92 = list()
            volume_95 = list()
            volume_50 = list()

            if i['fuel_type'] == 92:
                variant_sliv_92.append(i['variant_sliv'])
                volume_92.append(i['sum_sliv'])
                if i['variant_sliv'] in temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_92']:
                    index = temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_92'].index(i['variant_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_92'].insert(index, temp_azs_trucks4_dict[
                        str(i['variant'])]['volume_92'][index] + i['sum_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_92'].pop(index + 1)
                else:
                    temp_azs_trucks4_dict[str(i['variant'])] = {
                        'variant_sliv_92': temp_azs_trucks4_dict[str(i['variant'])][
                                               'variant_sliv_92'] + variant_sliv_92,
                        'variant_sliv_95': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_95'],
                        'variant_sliv_50': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_50'],
                        'volume_92': temp_azs_trucks4_dict[str(i['variant'])]['volume_92'] + volume_92,
                        'volume_95': temp_azs_trucks4_dict[str(i['variant'])]['volume_95'],
                        'volume_50': temp_azs_trucks4_dict[str(i['variant'])]['volume_50'],
                        'azs_id': temp_azs_trucks4_dict[str(i['variant'])]['azs_id'],
                        'truck_id': temp_azs_trucks4_dict[str(i['variant'])]['truck_id']
                    }
            if i['fuel_type'] == 95:
                variant_sliv_95.append(i['variant_sliv'])
                volume_95.append(i['sum_sliv'])
                if i['variant_sliv'] in temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_95']:
                    index = temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_95'].index(i['variant_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_95'].insert(index, temp_azs_trucks4_dict[
                        str(i['variant'])]['volume_95'][index] + i['sum_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_95'].pop(index + 1)
                else:
                    temp_azs_trucks4_dict[str(i['variant'])] = {
                        'variant_sliv_92': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_92'],
                        'variant_sliv_95': temp_azs_trucks4_dict[str(i['variant'])][
                                               'variant_sliv_95'] + variant_sliv_95,
                        'variant_sliv_50': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_50'],
                        'volume_92': temp_azs_trucks4_dict[str(i['variant'])]['volume_92'],
                        'volume_95': temp_azs_trucks4_dict[str(i['variant'])]['volume_95'] + volume_95,
                        'volume_50': temp_azs_trucks4_dict[str(i['variant'])]['volume_50'],
                        'azs_id': temp_azs_trucks4_dict[str(i['variant'])][
                            'azs_id'],
                        'truck_id': temp_azs_trucks4_dict[str(i['variant'])][
                            'truck_id']

                    }
            if i['fuel_type'] == 50:
                variant_sliv_50.append(i['variant_sliv'])
                volume_50.append(i['sum_sliv'])
                if i['variant_sliv'] in temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_50']:
                    index = temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_50'].index(i['variant_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_50'].insert(index, temp_azs_trucks4_dict[
                        str(i['variant'])]['volume_50'][index] + i['sum_sliv'])
                    temp_azs_trucks4_dict[str(i['variant'])]['volume_50'].pop(index + 1)
                else:
                    temp_azs_trucks4_dict[str(i['variant'])] = {
                        'variant_sliv_92': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_92'],
                        'variant_sliv_95': temp_azs_trucks4_dict[str(i['variant'])]['variant_sliv_95'],
                        'variant_sliv_50': temp_azs_trucks4_dict[str(i['variant'])][
                                               'variant_sliv_50'] + variant_sliv_50,
                        'volume_92': temp_azs_trucks4_dict[str(i['variant'])]['volume_92'],
                        'volume_95': temp_azs_trucks4_dict[str(i['variant'])]['volume_95'],
                        'volume_50': temp_azs_trucks4_dict[str(i['variant'])]['volume_50'] + volume_50,
                        'azs_id': temp_azs_trucks4_dict[str(i['variant'])][
                            'azs_id'],
                        'truck_id': temp_azs_trucks4_dict[str(i['variant'])][
                            'truck_id']
                    }
        azs_trucks_4_list_final = list()
        azs_trucks_4_dict_final = {'variant': 0,
                                   'sum_92': 0,
                                   'sum_95': 0,
                                   'sum_50': 0,
                                   'min_rez1': 0,
                                   'min_rez2': 0,
                                   'min_rez3': 0,
                                   'variant_sliv_50': 0,
                                   'variant_sliv_92': 0,
                                   'variant_sliv_95': 0,
                                   'azs_id': 0,
                                   'truck_id': 0
                                   }

        for i in temp_azs_trucks4_dict:
            variant_sliv_92 = temp_azs_trucks4_dict[i]['variant_sliv_92']
            variant_sliv_95 = temp_azs_trucks4_dict[i]['variant_sliv_95']
            variant_sliv_50 = temp_azs_trucks4_dict[i]['variant_sliv_50']
            volume_92 = temp_azs_trucks4_dict[i]['volume_92']
            volume_95 = temp_azs_trucks4_dict[i]['volume_95']
            volume_50 = temp_azs_trucks4_dict[i]['volume_50']
            if len(variant_sliv_92) != 0 and len(variant_sliv_95) != 0 and len(variant_sliv_50) != 0:
                for index_a, a in enumerate(variant_sliv_92):
                    for index_b, b in enumerate(variant_sliv_95):
                        for index_c, c in enumerate(variant_sliv_50):
                            azs_trucks_4_dict_final = {'variant': i,
                                                       'sum_92': volume_92[index_a],
                                                       'sum_95': volume_95[index_b],
                                                       'sum_50': volume_50[index_c],
                                                       'min_rez1': 0,
                                                       'min_rez2': 0,
                                                       'min_rez3': 0,
                                                       'variant_sliv_50': c,
                                                       'variant_sliv_92': a,
                                                       'variant_sliv_95': b,
                                                       'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                                       'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                                       }
                            azs_trucks_4_list_final.append(azs_trucks_4_dict_final)

            if len(variant_sliv_92) != 0 and len(variant_sliv_95) != 0 and len(variant_sliv_50) == 0:
                for index_a, a in enumerate(variant_sliv_92):
                    for index_b, b in enumerate(variant_sliv_95):
                        azs_trucks_4_dict_final = {'variant': i,
                                                   'sum_92': volume_92[index_a],
                                                   'sum_95': volume_95[index_b],
                                                   'sum_50': 0,
                                                   'min_rez1': 0,
                                                   'min_rez2': 0,
                                                   'min_rez3': 0,
                                                   'variant_sliv_50': 0,
                                                   'variant_sliv_92': a,
                                                   'variant_sliv_95': b,
                                                   'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                                   'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                                   }
                        azs_trucks_4_list_final.append(azs_trucks_4_dict_final)

            if len(variant_sliv_92) != 0 and len(variant_sliv_95) == 0 and len(variant_sliv_50) != 0:
                for index_a, a in enumerate(variant_sliv_92):
                    for index_c, c in enumerate(variant_sliv_50):
                        azs_trucks_4_dict_final = {'variant': i,
                                                   'sum_92': volume_92[index_a],
                                                   'sum_95': 0,
                                                   'sum_50': volume_50[index_c],
                                                   'min_rez1': 0,
                                                   'min_rez2': 0,
                                                   'min_rez3': 0,
                                                   'variant_sliv_50': c,
                                                   'variant_sliv_92': a,
                                                   'variant_sliv_95': 0,
                                                   'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                                   'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                                   }
                        azs_trucks_4_list_final.append(azs_trucks_4_dict_final)
            if len(variant_sliv_92) == 0 and len(variant_sliv_95) != 0 and len(variant_sliv_50) != 0:
                for index_b, b in enumerate(variant_sliv_95):
                    for index_c, c in enumerate(variant_sliv_50):
                        azs_trucks_4_dict_final = {'variant': i,
                                                   'sum_92': 0,
                                                   'sum_95': volume_95[index_b],
                                                   'sum_50': volume_50[index_c],
                                                   'min_rez1': 0,
                                                   'min_rez2': 0,
                                                   'min_rez3': 0,
                                                   'variant_sliv_50': c,
                                                   'variant_sliv_92': 0,
                                                   'variant_sliv_95': b,
                                                   'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                                   'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                                   }
                        azs_trucks_4_list_final.append(azs_trucks_4_dict_final)
            if len(variant_sliv_92) == 0 and len(variant_sliv_95) == 0 and len(variant_sliv_50) != 0:
                for index_c, c in enumerate(variant_sliv_50):
                    azs_trucks_4_dict_final = {'variant': i,
                                               'sum_92': 0,
                                               'sum_95': 0,
                                               'sum_50': volume_50[index_c],
                                               'min_rez1': 0,
                                               'min_rez2': 0,
                                               'min_rez3': 0,
                                               'variant_sliv_50': c,
                                               'variant_sliv_92': 0,
                                               'variant_sliv_95': 0,
                                               'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                               'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                               }
                    azs_trucks_4_list_final.append(azs_trucks_4_dict_final)
            if len(variant_sliv_92) == 0 and len(variant_sliv_95) != 0 and len(variant_sliv_50) == 0:
                for index_b, b in enumerate(variant_sliv_95):
                    azs_trucks_4_dict_final = {'variant': i,
                                               'sum_92': 0,
                                               'sum_95': volume_95[index_b],
                                               'sum_50': 0,
                                               'min_rez1': 0,
                                               'min_rez2': 0,
                                               'min_rez3': 0,
                                               'variant_sliv_50': 0,
                                               'variant_sliv_92': 0,
                                               'variant_sliv_95': b,
                                               'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                               'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                               }
                    azs_trucks_4_list_final.append(azs_trucks_4_dict_final)
            if len(variant_sliv_92) != 0 and len(variant_sliv_95) == 0 and len(variant_sliv_50) == 0:
                for index_a, a in enumerate(variant_sliv_92):
                    azs_trucks_4_dict_final = {'variant': i,
                                               'sum_92': volume_92[index_a],
                                               'sum_95': 0,
                                               'sum_50': 0,
                                               'min_rez1': 0,
                                               'min_rez2': 0,
                                               'min_rez3': 0,
                                               'variant_sliv_50': 0,
                                               'variant_sliv_92': a,
                                               'variant_sliv_95': 0,
                                               'azs_id': temp_azs_trucks4_dict[i]['azs_id'],
                                               'truck_id': temp_azs_trucks4_dict[i]['truck_id']
                                               }
                    azs_trucks_4_list_final.append(azs_trucks_4_dict_final)

        db.engine.execute(TempAzsTrucks3SecondTrip.__table__.insert(), temp_azs_trucks3_list)
        return azs_trucks_4_list_final, temp_azs_trucks3_list

    # определение худшего запаса суток среди всех резервуаров АЗС ДЛЯ ВТОРОГО РЕЙСА
    def preparation_six_second_trip():
        # берем таблицу 4
        table_azs_trucks_4, table_azs_trucks_3 = preparation_four_second_trip()
        table_azs_trucks_4_list = list()
        fuel_realisation = FuelRealisation.query.all()
        days_stock_old_dict = dict()
        for i in fuel_realisation:
            days_stock_old_dict[i.azs_id] = {'days_stock': [],
                                             'tank_id': []}

        for i in fuel_realisation:
            days_stock_min_old_list = [i.days_stock_min]
            tank_ids_list = [i.tank_id]
            days_stock_old_dict[i.azs_id] = {
                'days_stock': days_stock_old_dict[i.azs_id]['days_stock'] + days_stock_min_old_list,
                'tank_id': days_stock_old_dict[i.azs_id]['tank_id'] + tank_ids_list}
        variants_dict = dict()
        for i in table_azs_trucks_3:
            variants_dict[(i['variant'], i['variant_sliv'])] = {'days_stock': [],
                                                                'tank_id': []}
        for i in table_azs_trucks_3:
            days_stock_min_new_list = [i['new_days_stock']]
            tank_ids_list = [i['tank_id']]
            variants_dict[(i['variant'], i['variant_sliv'])] = {
                'days_stock': variants_dict[(i['variant'], i['variant_sliv'])][
                                  'days_stock'] + days_stock_min_new_list,
                'tank_id': variants_dict[(i['variant'], i['variant_sliv'])]['tank_id'] + tank_ids_list
            }
        for row in table_azs_trucks_4:

            variant = row['variant']
            variants_list = list()
            if row['variant_sliv_92']:
                variants_list.append(row['variant_sliv_92'])
            if row['variant_sliv_95']:
                variants_list.append(row['variant_sliv_95'])
            if row['variant_sliv_50']:
                variants_list.append(row['variant_sliv_50'])
            new_days_stock_dict = dict()
            tank_list = days_stock_old_dict[row['azs_id']]['tank_id']
            stock_list = days_stock_old_dict[row['azs_id']]['days_stock']
            for index, tank in enumerate(tank_list):
                new_days_stock_dict[tank] = stock_list[index]
            for variant_sliv in variants_list:
                tank_list_new = variants_dict[(int(row['variant']), variant_sliv)]['tank_id']
                stock_list_new = variants_dict[(int(row['variant']), variant_sliv)]['days_stock']
                for index, tank in enumerate(tank_list_new):
                    new_days_stock_dict[tank] = stock_list_new[index]

            sorted_new_days_stock_dict = sorted(new_days_stock_dict.items(), key=lambda x: x[1])
            if len(sorted_new_days_stock_dict) < 3:
                temp_azs_trucks_4_dict = {'truck_id': row['truck_id'],
                                          'azs_id': row['azs_id'],
                                          'variant': row['variant'],
                                          'sum_92': row['sum_92'],
                                          'sum_95': row['sum_95'],
                                          'sum_50': row['sum_50'],
                                          'min_rez1': round(sorted_new_days_stock_dict[0][1], 1),
                                          'min_rez2': round(sorted_new_days_stock_dict[1][1], 1),
                                          'min_rez3': 1000,
                                          'variant_sliv_92': row['variant_sliv_92'],
                                          'variant_sliv_95': row['variant_sliv_95'],
                                          'variant_sliv_50': row['variant_sliv_50']
                                          }

            else:
                temp_azs_trucks_4_dict = {'truck_id': row['truck_id'],
                                          'azs_id': row['azs_id'],
                                          'variant': row['variant'],
                                          'sum_92': row['sum_92'],
                                          'sum_95': row['sum_95'],
                                          'sum_50': row['sum_50'],
                                          'min_rez1': round(sorted_new_days_stock_dict[0][1], 1),
                                          'min_rez2': round(sorted_new_days_stock_dict[1][1], 1),
                                          'min_rez3': round(sorted_new_days_stock_dict[2][1], 1),
                                          'variant_sliv_92': row['variant_sliv_92'],
                                          'variant_sliv_95': row['variant_sliv_95'],
                                          'variant_sliv_50': row['variant_sliv_50']
                                          }
            table_azs_trucks_4_list.append(temp_azs_trucks_4_dict)

        db.engine.execute(TempAzsTrucks4SecondTrip.__table__.insert(), table_azs_trucks_4_list)

    preparation_six_second_trip()
    return redirect(url_for('main.wait_10'))


@bp.route('/wait_10', methods=['POST', 'GET'])
@login_required
def wait_10():
    time.sleep(10)
    return redirect(url_for('main.start_second_trip'))


@bp.route('/start_second_trip', methods=['POST', 'GET'])
@login_required
def start_second_trip():
    work_type = WorkType.query.filter_by(active=True).first_or_404()
    if work_type.id == 2 or work_type.id == 3:
        fuel_type = int(work_type.fuel_type)
        min_days_stock_global = float(work_type.days_stock_limit)
    else:
        fuel_type = 0

    # получаем данные из 4 таблицы второго рейса
    table_azs_trucks_4 = TempAzsTrucks4SecondTrip.query.all()
    # создаем словарь для хранения информации о том, какие бензовозы могут отправиться на данную АЗС
    trucks_for_azs_dict = dict()
    azs_trucks_best_days_stock = dict()
    # берем все активные азс расставленые по приоритету (от самой критичной к менее критичной)
    azs_list = Priority.query.order_by("priority").all()
    # берем все активные бензовозы из таблицы
    active_trucks = Trucks.query.filter_by(active=True).all()
    # получаем информацию о первом рейсе
    first_trip = Trips.query.filter_by(trip_number=1).order_by(desc("calculate_id")).first()
    # получаем информацию о расстановке первого рейса
    first_trip_list = Result.query.filter_by(calculate_id=first_trip.calculate_id).all()
    # создаем список азс на которые отправлены бензовозы первым рейсом
    first_trip_azs_list = list()
    # создаем список бензовозов из первого рейса
    first_trip_truck_list = list()

    # заполняем список с азс первого рейса
    for i in first_trip_list:
        # заполняем список с азс первого рейса
        first_trip_azs_list.append(i.azs_id)

    # for i in active_trucks:
    # заполняем список бензовозов из первого рейса
    # first_trip_truck_list.append(i.truck_id)

    # создаем список АЗС на которые необходимо отправить бензовозы во второй рейс(с сортировкой по критичности)
    second_trip_azs_list = list()
    # заполняем список с азс для второго рейса
    for i in azs_list:
        if i.azs_id not in first_trip_azs_list:
            second_trip_azs_list.append(i.azs_id)

    # словарь для хранения информации о том, на какие АЗС может отправиться бензовоз
    azs_for_trucks_dict = dict()
    for i in table_azs_trucks_4:
        if i.azs_id in second_trip_azs_list:
            # инициализируем словарь для хранения информации о том, на какие АЗС может отправиться бензовоз
            azs_for_trucks_dict[i.truck_id] = {'azs_ids': [],
                                               'priority': []}

    for i in table_azs_trucks_4:
        # если на АЗС первым рейсом уже был отправлен бензовоз, то мы исключаем её из списка для второго рейса
        if i.azs_id in second_trip_azs_list:
            # словарь для хранения информации о том, какие бензовозы могут отправиться на данную АЗС
            trucks_for_azs_dict[i.azs_id] = {'azs_trucks': [0]}
            # Словарь для первого режима работы
            # Информация о лучщем заполнении для пары АЗС:БЕНЗОВОЗ
            azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_rez1': -1,
                                                                                 'min_rez2': -1,
                                                                                 'min_rez3': -1,
                                                                                 'variant': 0,
                                                                                 'variant_sliv_92': 0,
                                                                                 'variant_sliv_95': 0,
                                                                                 'variant_sliv_50': 0}

    # Перебираем таблицу table_azs_truck_4 для заполнения словарей для всех трех режимов работы
    for i in table_azs_trucks_4:
        # заполняем словарь для хранения информации о том, на какие азс может отправиться бензовоз
        # если на АЗС первым рейсом уже был отправлен бензовоз, то мы исключаем её из списка для второго рейса
        if i.azs_id in second_trip_azs_list:
            azs_list_second_trip = list()
            azs_list_second_trip.append(i.azs_id)
            if i.azs_id not in azs_for_trucks_dict[i.truck_id]['azs_ids']:
                azs_for_trucks_dict[i.truck_id] = {
                    'azs_ids': azs_for_trucks_dict[i.truck_id]['azs_ids'] + azs_list_second_trip,
                    'priority': 0}

            # заполняем словарь для хранения информации о том, какие бензовозы могут отправиться на данную АЗС
            trucks_list = list()
            trucks_list.append(i.truck_id)
            if i.truck_id not in trucks_for_azs_dict[i.azs_id]['azs_trucks']:
                trucks_for_azs_dict[i.azs_id] = {
                    'azs_trucks': trucks_for_azs_dict[i.azs_id]['azs_trucks'] + trucks_list
                }

            # Заполняем словари информацией о лучщем заполнении для пары АЗС:БЕНЗОВОЗ (для первого режима работы)
            min_rez1 = azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez1']
            min_rez2 = azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez2']
            min_rez3 = azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)]['min_rez3']

            if i.min_rez1 > min_rez1 \
                    or (i.min_rez1 == min_rez1 and i.min_rez2 > min_rez2) \
                    or (i.min_rez1 == min_rez1 and i.min_rez2 == min_rez2 and i.min_rez3 > min_rez3):
                azs_trucks_best_days_stock[str(i.azs_id) + ':' + str(i.truck_id)] = {'min_rez1': i.min_rez1,
                                                                                     'min_rez2': i.min_rez2,
                                                                                     'min_rez3': i.min_rez3,
                                                                                     'variant': i.variant,
                                                                                     'variant_sliv_92': i.variant_sliv_92,
                                                                                     'variant_sliv_95': i.variant_sliv_95,
                                                                                     'variant_sliv_50': i.variant_sliv_50}

    # ПОДГОТОВКА К РАССТАНОВКЕ БЕНЗОВОЗОВ
    global_settings = GlobalSettings.query.filter_by(name="algorithm_2").first()

    variant_send_truck = global_settings.algorithm_id
    active_azs_count = Priority.query.count()

    active_trucks = Trucks.query.filter_by(active=True).count()  # получаем количество активных бензовозов
    active_azs = Priority.query.order_by(
        "priority").all()  # получаем список активных АЗС из таблицы Priority
    # с сортировкой по важности (чем меньше число (стобец priority), тем важнее отправить бензовоз на эту АЗС
    choices_dict_work_type_1 = dict()  # храним итоговые варианты расстановки с оценкой каждого для 1 режима работы
    choices_dict_work_type_2 = dict()  # храним итоговые варианты расстановки с оценкой каждого для 2 режима работы
    choices_dict_work_type_3 = dict()  # храним итоговые варианты расстановки с оценкой каждого для 3 режима работы
    azs_queue_dict = dict()  # создаем словарь для хранения id АЗС в порядке важности отправки бензовоза на АЗС
    # (нужна для последующего анализа итоговых расстановок)

    for i in azs_list:  # заполняем словарь
        azs_queue_dict[i.azs_id] = {'queue': i.priority}

    # таймаут для принудительной остановки расстановки бензовозов через
    # указанное количество времени (сейчас минута)
    timeout = time.time() + 20 * 1

    # создаем словарь для хранения всех удачных расстановок
    good_choices_dict = dict()

    # количество успешных расстановок
    number_of_success_loops = 0
    # вводим переменную на случай, если расстановка бензовозов не удастся.
    # Изначально переменной присвоена 1, что означает неудачную расстановку.
    alarm = 1

    # Главный цикл расстановки бензовозов (количество попыток от 0 до 10 млн)

    for x in trucks_for_azs_dict:
        print("АЗС:", x, "БЕЗОВОЗЫ:", trucks_for_azs_dict[x])

    # АЛГОРИТМ №1 - СЛУЧАНАЯ РАССТАНОВКА С ОГРОМНЫМ КОЛИЧЕСТВОМ ВАРИАНТОВ
    if variant_send_truck == 1:
        if active_trucks <= active_azs_count:
            max_trucks_good = 0
            trucks_for_azs_dict_sorted = dict()
            for i in active_azs:
                azs_id = i.azs_id
                if azs_id in trucks_for_azs_dict:
                    trucks_for_azs_dict_sorted[azs_id] = {
                        'azs_trucks': trucks_for_azs_dict[azs_id]['azs_trucks']}

            trucks_for_azs_list_sorted = sorted(trucks_for_azs_dict_sorted,
                                                key=lambda k: len(
                                                    trucks_for_azs_dict_sorted[k]['azs_trucks']))
            print('Sorted table')
            for k in sorted(trucks_for_azs_dict_sorted,
                            key=lambda k: len(trucks_for_azs_dict_sorted[k]['azs_trucks'])):
                print(k, trucks_for_azs_dict_sorted[k]['azs_trucks'])

            print('Start!')

            choise_good = 0  # переменная-триггер, по которой определяем, что расстановка удалась (или нет!)
            for choice in range(0, 1000000000):
                choice_azs_truck_dict = dict()
                used_trucks = list()
                temp_truck_count = 0
                for i in trucks_for_azs_list_sorted:  # перебираем все активные АЗС
                    azs_id = i
                    if azs_id in trucks_for_azs_dict:  # если есть хотябы один бензовоз, который можно отправить на АЗС
                        azs_trucks = trucks_for_azs_dict[azs_id][
                            'azs_trucks']  # получаем список всех бензовозов,
                        # которые можно отправить на эту АЗС (включая 0 - т.е. АЗС на которую не будет отправлен бензовоз)
                        truck_id = random.choice(azs_trucks)  # функцией RANDOM из списка azs_trucks
                        # выбираем бензовоз для этой АЗС
                        if truck_id in used_trucks and truck_id != 0:  # если данный безовоз уже был в данном варианте
                            # расстановки и он не равен 0, то считаем вариант не удачным, и досрочно прерыываем цикл
                            # good = 0
                            break
                        # если все хорошо, то
                        else:
                            # добавляем этот бензовоз к списку использованных
                            # в данном варианте бензовозов
                            used_trucks.append(truck_id)

                            # добавляем параметр azs_id-truck_id
                            # в словарь с расстановкой
                            choice_azs_truck_dict[azs_id] = {'truck_id': truck_id}
                            # если безовоз не нулевой, то уменьшаем количество бензовозов которые
                            if truck_id != 0:
                                # требуется расставить на 1
                                temp_truck_count = temp_truck_count + 1
                                # если все бензовозы расставлены (счетчик равен 0)

                if temp_truck_count == active_trucks or (
                        temp_truck_count >= max_trucks_good and temp_truck_count > 0):
                    max_trucks_good = temp_truck_count
                    # то помечаем вариант хорошим
                    choise_good = 1
                    number_of_success_loops = number_of_success_loops + 1
                    good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict,
                                                                  'max_trucks_good': max_trucks_good}

                if time.time() > timeout:
                    # то цикл принудительно прерывается
                    break

            if choise_good == 1:
                print("OK! Нашли ", number_of_success_loops, "вариантов")
            else:
                print("NOT OK!!!")
        else:
            print(
                "Расстановка бензовозов невозможна! Количество активных бензовозов больше числа активных АЗС!")

    # АЛГОРИТМ №2 - СЛУЧАЙНАЯ РАССТАНОВКА ОТ ВЕРХА К НИЗУ
    if variant_send_truck == 2:
        if active_trucks <= active_azs_count:
            max_trucks_good = 0
            choise_good = 0
            for choice in range(0, 1000000000):
                choice_azs_truck_dict = dict()
                used_trucks = list()
                temp_truck_count = 0
                good = 0
                for i in active_azs:  # перебираем все активные АЗС
                    azs_id = i.azs_id
                    if azs_id in trucks_for_azs_dict:  # если есть хотябы один бензовоз, который можно отправить на АЗС
                        trig = 0
                        for x in range(1, 1000):
                            azs_trucks = trucks_for_azs_dict[azs_id][
                                'azs_trucks']  # получаем список всех бензовозов,
                            # которые можно отправить на эту АЗС (включая 0 - т.е. АЗС на которую не будет отправлен бензовоз)
                            truck_id = random.choice(azs_trucks)  # функцией RANDOM из списка azs_trucks
                            # выбираем бензовоз для этой АЗС
                            if truck_id in used_trucks and truck_id != 0:  # если данный безовоз уже был в данном варианте
                                d = 1  # заглушка
                            else:  # если все хорошо, то
                                # добавляем этот бензовоз к списку использованных
                                # в данном варианте бензовозов
                                used_trucks.append(truck_id)
                                # добавляем параметр azs_id-truck_id
                                # в словарь с расстановкой
                                choice_azs_truck_dict[azs_id] = {'truck_id': truck_id}
                                trig = 1
                                if truck_id != 0:
                                    temp_truck_count = temp_truck_count + 1
                                break
                        if trig == 0:
                            used_trucks.append(0)
                            choice_azs_truck_dict[azs_id] = {'truck_id': 0}

                        if temp_truck_count == active_trucks:
                            good = 1
                            break

                if good == 1 or (temp_truck_count >= max_trucks_good and temp_truck_count > 0):
                    max_trucks_good = temp_truck_count
                    number_of_success_loops = number_of_success_loops + 1
                    good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict,
                                                                  'max_trucks_good': max_trucks_good
                                                                  }
                    choise_good = 1

                if time.time() > timeout:
                    # то цикл принудительно прерывается
                    break

            if choise_good == 1:
                print("OK! Нашли ", number_of_success_loops, "вариантов")
            else:
                print("NOT OK!!!")
        else:
            print(
                "Расстановка бензовозов невозможна! Количество активных бензовозов больше числа активных АЗС!")

    # АЛГОРИТМ №3 - ПОЛНЫЙ ПЕРЕБОР
    if variant_send_truck == 3:
        choice_azs_truck_dict = dict()
        azs_for_trucks_with_priority = dict()
        for i in azs_for_trucks_dict:
            temp_azs_list = azs_for_trucks_dict[i]['azs_ids']
            temp_priority = list()
            for azs in temp_azs_list:
                temp_priority.append(azs_queue_dict[azs]['queue'])
            azs_for_trucks_with_priority[i] = {'azs_ids': azs_for_trucks_dict[i]['azs_ids'],
                                               'priority': temp_priority}

        bubble_azs_for_trucks = dict()
        # делаем сортировку пузырьком

        for i in azs_for_trucks_with_priority:
            temp_azs_list = azs_for_trucks_with_priority[i]['azs_ids']
            temp_priority = azs_for_trucks_with_priority[i]['priority']
            for n2 in range(1, len(temp_priority)):
                for n in range(1, len(temp_priority)):
                    if temp_priority[n] < temp_priority[n - 1]:
                        temp = temp_priority[n - 1]
                        temp_priority[n - 1] = temp_priority[n]
                        temp_priority[n] = temp

                        temp2 = temp_azs_list[n - 1]
                        temp_azs_list[n - 1] = temp_azs_list[n]
                        temp_azs_list[n] = temp2
            bubble_azs_for_trucks[i] = {'azs_ids': temp_azs_list[:active_trucks - 1],
                                        'priority:': temp_priority[:active_trucks - 1]}

        k = 1
        final_azs_for_trucks = dict()
        truck_id_number = dict()
        for i in bubble_azs_for_trucks:
            truck_id_number[k] = {'truck_id': i}
            final_azs_for_trucks[k] = {"azs_ids": bubble_azs_for_trucks[i]['azs_ids']}
            k = k + 1

        # проверяем, что количество активных бензовозов не меньше, чем количество доступных для отправки на АЗС
        available_active_trucks = len(azs_for_trucks_dict)
        # если количество доступных бензовозов меньше, то только это количество считаем активными
        if active_trucks >= len(azs_for_trucks_dict):
            active_trucks = available_active_trucks
        print('---------------LENA POLENO-------------------')
        print(azs_for_trucks_dict)
        print(active_trucks)
        if active_trucks == 11:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0

            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        for n3 in final_azs_for_trucks[3]["azs_ids"]:
                            if n3 != n2 and n3 != n1:
                                for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                    if n4 != n3 and n4 != n2 and n4 != n1:
                                        for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                            if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                    if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                        for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                            if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                for n8 in final_azs_for_trucks[8]["azs_ids"]:
                                                                    if n8 != n7 and n8 != n6 and n8 != n5 and n8 != n4 and n8 != n3 and n8 != n2 and n8 != n1:
                                                                        for n9 in final_azs_for_trucks[9]["azs_ids"]:
                                                                            if n9 != n8 and n9 != n7 and n9 != n6 and n9 != n5 and n9 != n4 and n9 != n3 and n9 != n2 and n9 != n1:
                                                                                for n10 in final_azs_for_trucks[10][
                                                                                    "azs_ids"]:
                                                                                    if n10 != n9 and n10 != n8 and n10 != n7 and n10 != n6 and n10 != n5 and n10 != n4 and n10 != n3 and n10 != n2 and n10 != n1:
                                                                                        for n11 in \
                                                                                                final_azs_for_trucks[
                                                                                                    11][
                                                                                                    "azs_ids"]:
                                                                                            if n11 != n10 and n11 != n9 and n11 != n8 and n11 != n7 and n11 != n6 and n11 != n5 and n11 != n4 and n11 != n3 and n11 != n2 and n11 != n1:

                                                                                                choice_azs_truck_dict = dict()
                                                                                                choice_azs_truck_dict[
                                                                                                    n1] = {'truck_id':
                                                                                                               truck_id_number[
                                                                                                                   1][
                                                                                                                   'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n2] = {'truck_id':
                                                                                                               truck_id_number[
                                                                                                                   2][
                                                                                                                   'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n3] = {'truck_id':
                                                                                                               truck_id_number[
                                                                                                                   3][
                                                                                                                   'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n4] = {'truck_id':
                                                                                                               truck_id_number[
                                                                                                                   4][
                                                                                                                   'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n5] = {'truck_id':
                                                                                                               truck_id_number[
                                                                                                                   5][
                                                                                                                   'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n6] = {'truck_id':
                                                                                                               truck_id_number[
                                                                                                                   6][
                                                                                                                   'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n7] = {'truck_id':
                                                                                                               truck_id_number[
                                                                                                                   7][
                                                                                                                   'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n8] = {'truck_id':
                                                                                                               truck_id_number[
                                                                                                                   8][
                                                                                                                   'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n9] = {'truck_id':
                                                                                                               truck_id_number[
                                                                                                                   9][
                                                                                                                   'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n10] = {'truck_id':
                                                                                                                truck_id_number[
                                                                                                                    10][
                                                                                                                    'truck_id']}
                                                                                                choice_azs_truck_dict[
                                                                                                    n11] = {'truck_id':
                                                                                                                truck_id_number[
                                                                                                                    11][
                                                                                                                    'truck_id']}
                                                                                                number_variant = number_variant + 1
                                                                                                if number_variant & 1000 == 0:
                                                                                                    number_of_success_loops = number_of_success_loops + 1
                                                                                                    good_choices_dict[
                                                                                                        number_of_success_loops] = {
                                                                                                        'variants': choice_azs_truck_dict}
        if active_trucks == 10:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0

            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        for n3 in final_azs_for_trucks[3]["azs_ids"]:
                            if n3 != n2 and n3 != n1:
                                for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                    if n4 != n3 and n4 != n2 and n4 != n1:
                                        for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                            if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                    if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                        for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                            if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                for n8 in final_azs_for_trucks[8]["azs_ids"]:
                                                                    if n8 != n7 and n8 != n6 and n8 != n5 and n8 != n4 and n8 != n3 and n8 != n2 and n8 != n1:
                                                                        for n9 in final_azs_for_trucks[9]["azs_ids"]:
                                                                            if n9 != n8 and n9 != n7 and n9 != n6 and n9 != n5 and n9 != n4 and n9 != n3 and n9 != n2 and n9 != n1:
                                                                                for n10 in final_azs_for_trucks[10][
                                                                                    "azs_ids"]:
                                                                                    if n10 != n9 and n10 != n8 and n10 != n7 and n10 != n6 and n10 != n5 and n10 != n4 and n10 != n3 and n10 != n2 and n10 != n1:
                                                                                        choice_azs_truck_dict = dict()
                                                                                        choice_azs_truck_dict[n1] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[1][
                                                                                                    'truck_id']}
                                                                                        choice_azs_truck_dict[n2] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[2][
                                                                                                    'truck_id']}
                                                                                        choice_azs_truck_dict[n3] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[3][
                                                                                                    'truck_id']}
                                                                                        choice_azs_truck_dict[n4] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[4][
                                                                                                    'truck_id']}
                                                                                        choice_azs_truck_dict[n5] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[5][
                                                                                                    'truck_id']}
                                                                                        choice_azs_truck_dict[n6] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[6][
                                                                                                    'truck_id']}
                                                                                        choice_azs_truck_dict[n7] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[7][
                                                                                                    'truck_id']}
                                                                                        choice_azs_truck_dict[n8] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[8][
                                                                                                    'truck_id']}
                                                                                        choice_azs_truck_dict[n9] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[9][
                                                                                                    'truck_id']}
                                                                                        choice_azs_truck_dict[n10] = {
                                                                                            'truck_id':
                                                                                                truck_id_number[10][
                                                                                                    'truck_id']}
                                                                                        number_variant = number_variant + 1
                                                                                        if number_variant & 1000 == 0:
                                                                                            number_of_success_loops = number_of_success_loops + 1
                                                                                            good_choices_dict[
                                                                                                number_of_success_loops] = {
                                                                                                'variants': choice_azs_truck_dict}
        if active_trucks == 9:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0

            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        for n3 in final_azs_for_trucks[3]["azs_ids"]:
                            if n3 != n2 and n3 != n1:
                                for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                    if n4 != n3 and n4 != n2 and n4 != n1:
                                        for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                            if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                    if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                        for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                            if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                for n8 in final_azs_for_trucks[8]["azs_ids"]:
                                                                    if n8 != n7 and n8 != n6 and n8 != n5 and n8 != n4 and n8 != n3 and n8 != n2 and n8 != n1:
                                                                        for n9 in final_azs_for_trucks[9]["azs_ids"]:
                                                                            if n9 != n8 and n9 != n7 and n9 != n6 and n9 != n5 and n9 != n4 and n9 != n3 and n9 != n2 and n9 != n1:
                                                                                choice_azs_truck_dict = dict()
                                                                                choice_azs_truck_dict[n1] = {
                                                                                    'truck_id': truck_id_number[1][
                                                                                        'truck_id']}
                                                                                choice_azs_truck_dict[n2] = {
                                                                                    'truck_id': truck_id_number[2][
                                                                                        'truck_id']}
                                                                                choice_azs_truck_dict[n3] = {
                                                                                    'truck_id': truck_id_number[3][
                                                                                        'truck_id']}
                                                                                choice_azs_truck_dict[n4] = {
                                                                                    'truck_id': truck_id_number[4][
                                                                                        'truck_id']}
                                                                                choice_azs_truck_dict[n5] = {
                                                                                    'truck_id': truck_id_number[5][
                                                                                        'truck_id']}
                                                                                choice_azs_truck_dict[n6] = {
                                                                                    'truck_id': truck_id_number[6][
                                                                                        'truck_id']}
                                                                                choice_azs_truck_dict[n7] = {
                                                                                    'truck_id': truck_id_number[7][
                                                                                        'truck_id']}
                                                                                choice_azs_truck_dict[n8] = {
                                                                                    'truck_id': truck_id_number[8][
                                                                                        'truck_id']}
                                                                                choice_azs_truck_dict[n9] = {
                                                                                    'truck_id': truck_id_number[9][
                                                                                        'truck_id']}

                                                                                number_variant = number_variant + 1
                                                                                if number_variant & 1000 == 0:
                                                                                    number_of_success_loops = number_of_success_loops + 1
                                                                                    good_choices_dict[
                                                                                        number_of_success_loops] = {
                                                                                        'variants': choice_azs_truck_dict}

        if active_trucks == 8:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0

            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        for n3 in final_azs_for_trucks[3]["azs_ids"]:
                            if n3 != n2 and n3 != n1:
                                for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                    if n4 != n3 and n4 != n2 and n4 != n1:
                                        for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                            if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                    if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                        for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                            if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                for n8 in final_azs_for_trucks[8]["azs_ids"]:
                                                                    if n8 != n7 and n8 != n6 and n8 != n5 and n8 != n4 and n8 != n3 and n8 != n2 and n8 != n1:
                                                                        choice_azs_truck_dict = dict()
                                                                        choice_azs_truck_dict[n1] = {
                                                                            'truck_id': truck_id_number[1]['truck_id']}
                                                                        choice_azs_truck_dict[n2] = {
                                                                            'truck_id': truck_id_number[2]['truck_id']}
                                                                        choice_azs_truck_dict[n3] = {
                                                                            'truck_id': truck_id_number[3]['truck_id']}
                                                                        choice_azs_truck_dict[n4] = {
                                                                            'truck_id': truck_id_number[4]['truck_id']}
                                                                        choice_azs_truck_dict[n5] = {
                                                                            'truck_id': truck_id_number[5]['truck_id']}
                                                                        choice_azs_truck_dict[n6] = {
                                                                            'truck_id': truck_id_number[6]['truck_id']}
                                                                        choice_azs_truck_dict[n7] = {
                                                                            'truck_id': truck_id_number[7]['truck_id']}
                                                                        choice_azs_truck_dict[n8] = {
                                                                            'truck_id': truck_id_number[8]['truck_id']}

                                                                        number_variant = number_variant + 1
                                                                        number_of_success_loops = number_of_success_loops + 1
                                                                        good_choices_dict[number_of_success_loops] = {
                                                                            'variants': choice_azs_truck_dict}
        if active_trucks == 7:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0
            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        for n3 in final_azs_for_trucks[3]["azs_ids"]:
                            if n3 != n2 and n3 != n1:
                                for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                    if n4 != n3 and n4 != n2 and n4 != n1:
                                        for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                            if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                    if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                        for n7 in final_azs_for_trucks[7]["azs_ids"]:
                                                            if n7 != n6 and n7 != n5 and n7 != n4 and n7 != n3 and n7 != n2 and n7 != n1:
                                                                choice_azs_truck_dict = dict()
                                                                choice_azs_truck_dict[n1] = {
                                                                    'truck_id': truck_id_number[1]['truck_id']}
                                                                choice_azs_truck_dict[n2] = {
                                                                    'truck_id': truck_id_number[2]['truck_id']}
                                                                choice_azs_truck_dict[n3] = {
                                                                    'truck_id': truck_id_number[3]['truck_id']}
                                                                choice_azs_truck_dict[n4] = {
                                                                    'truck_id': truck_id_number[4]['truck_id']}
                                                                choice_azs_truck_dict[n5] = {
                                                                    'truck_id': truck_id_number[5]['truck_id']}
                                                                choice_azs_truck_dict[n6] = {
                                                                    'truck_id': truck_id_number[6]['truck_id']}
                                                                choice_azs_truck_dict[n7] = {
                                                                    'truck_id': truck_id_number[7]['truck_id']}

                                                                number_variant = number_variant + 1

                                                                number_of_success_loops = number_of_success_loops + 1
                                                                good_choices_dict[number_of_success_loops] = {
                                                                    'variants': choice_azs_truck_dict}

        if active_trucks == 6:
            print('6 бензовозов')
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0

            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        for n3 in final_azs_for_trucks[3]["azs_ids"]:
                            if n3 != n2 and n3 != n1:
                                for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                    if n4 != n3 and n4 != n2 and n4 != n1:
                                        for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                            if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                for n6 in final_azs_for_trucks[6]["azs_ids"]:
                                                    if n6 != n5 and n6 != n4 and n6 != n3 and n6 != n2 and n6 != n1:
                                                        choice_azs_truck_dict = dict()
                                                        choice_azs_truck_dict[n1] = {
                                                            'truck_id': truck_id_number[1]['truck_id']}
                                                        choice_azs_truck_dict[n2] = {
                                                            'truck_id': truck_id_number[2]['truck_id']}
                                                        choice_azs_truck_dict[n3] = {
                                                            'truck_id': truck_id_number[3]['truck_id']}
                                                        choice_azs_truck_dict[n4] = {
                                                            'truck_id': truck_id_number[4]['truck_id']}
                                                        choice_azs_truck_dict[n5] = {
                                                            'truck_id': truck_id_number[5]['truck_id']}
                                                        choice_azs_truck_dict[n6] = {
                                                            'truck_id': truck_id_number[6]['truck_id']}

                                                        number_variant = number_variant + 1
                                                        number_of_success_loops = number_of_success_loops + 1
                                                        good_choices_dict[number_of_success_loops] = {
                                                            'variants': choice_azs_truck_dict}
            # if number_variant == 0:
            #     active_trucks = active_trucks - 1
        if active_trucks == 5:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0

            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        for n3 in final_azs_for_trucks[3]["azs_ids"]:
                            if n3 != n2 and n3 != n1:
                                for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                    if n4 != n3 and n4 != n2 and n4 != n1:
                                        for n5 in final_azs_for_trucks[5]["azs_ids"]:
                                            if n5 != n4 and n5 != n3 and n5 != n2 and n5 != n1:
                                                choice_azs_truck_dict = dict()
                                                choice_azs_truck_dict[n1] = {'truck_id': truck_id_number[1]['truck_id']}
                                                choice_azs_truck_dict[n2] = {'truck_id': truck_id_number[2]['truck_id']}
                                                choice_azs_truck_dict[n3] = {'truck_id': truck_id_number[3]['truck_id']}
                                                choice_azs_truck_dict[n4] = {'truck_id': truck_id_number[4]['truck_id']}
                                                choice_azs_truck_dict[n5] = {'truck_id': truck_id_number[5]['truck_id']}

                                                number_variant = number_variant + 1
                                                number_of_success_loops = number_of_success_loops + 1
                                                good_choices_dict[number_of_success_loops] = {
                                                    'variants': choice_azs_truck_dict}

        if active_trucks == 4:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0
            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        for n3 in final_azs_for_trucks[3]["azs_ids"]:
                            if n3 != n2 and n3 != n1:
                                for n4 in final_azs_for_trucks[4]["azs_ids"]:
                                    if n4 != n3 and n4 != n2 and n4 != n1:
                                        choice_azs_truck_dict = dict()
                                        choice_azs_truck_dict[n1] = {'truck_id': truck_id_number[1]['truck_id']}
                                        choice_azs_truck_dict[n2] = {'truck_id': truck_id_number[2]['truck_id']}
                                        choice_azs_truck_dict[n3] = {'truck_id': truck_id_number[3]['truck_id']}
                                        choice_azs_truck_dict[n4] = {'truck_id': truck_id_number[4]['truck_id']}
                                        number_variant = number_variant + 1
                                        number_of_success_loops = number_of_success_loops + 1
                                        good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict}
        if active_trucks == 3:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0
            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        for n3 in final_azs_for_trucks[3]["azs_ids"]:
                            if n3 != n2 and n3 != n1:
                                choice_azs_truck_dict = dict()
                                choice_azs_truck_dict[n1] = {'truck_id': truck_id_number[1]['truck_id']}
                                choice_azs_truck_dict[n2] = {'truck_id': truck_id_number[2]['truck_id']}
                                choice_azs_truck_dict[n3] = {'truck_id': truck_id_number[3]['truck_id']}
                                number_variant = number_variant + 1
                                number_of_success_loops = number_of_success_loops + 1
                                good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict}

        if active_trucks == 2:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0
            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                for n2 in final_azs_for_trucks[2]["azs_ids"]:
                    if n2 != n1:
                        choice_azs_truck_dict = dict()
                        choice_azs_truck_dict[n1] = {'truck_id': truck_id_number[1]['truck_id']}
                        choice_azs_truck_dict[n2] = {'truck_id': truck_id_number[2]['truck_id']}
                        number_variant = number_variant + 1
                        number_of_success_loops = number_of_success_loops + 1
                        good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict}

        if active_trucks == 1:
            # поскольку при полном переборе нет нерабочих вариантов расстановки, то расстановка всегда успешна
            choise_good = 1
            number_of_success_loops = 0
            number_variant = 0
            for n1 in final_azs_for_trucks[1]["azs_ids"]:
                choice_azs_truck_dict = dict()
                choice_azs_truck_dict[n1] = {'truck_id': truck_id_number[1]['truck_id']}
                number_variant = number_variant + 1
                number_of_success_loops = number_of_success_loops + 1
                good_choices_dict[number_of_success_loops] = {'variants': choice_azs_truck_dict}
    
    if choise_good == 1:
        for choice in good_choices_dict:
            '''**************************************************************************************************'''
            choice_azs_truck_dict = good_choices_dict[choice]['variants']
            # Оцениваем вариант расстановки на предмет не отправки бензовоза на критичные АЗС
            # переменная для хранения оценки текущей расстановки бензовозов (т.е. чем большее количество
            points = 0
            # критичных АЗС пропущено, тем меньше оценка (расстановка хуже)
            for i in choice_azs_truck_dict:  #
                if choice_azs_truck_dict[i]['truck_id'] != 0:
                    points = points + (1 / (azs_queue_dict[i]['queue'])) * 1000
            # округляем оценку до целого числа
            points = int(points)
            '''**************************************************************************************************'''

            '''**************************************************************************************************'''
            # минимальный запас суток среди всех АЗС
            min_days_stock1_work_type_1 = 1234
            min_days_stock2_work_type_1 = 1234
            # перебираем список расстановки
            for i in choice_azs_truck_dict:
                if choice_azs_truck_dict[i]['truck_id'] != 0 \
                        and (azs_trucks_best_days_stock[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                                 'min_rez1'] < min_days_stock1_work_type_1):
                    min_days_stock2_work_type_1 = min_days_stock1_work_type_1
                    min_days_stock1_work_type_1 = \
                        azs_trucks_best_days_stock[str(i) + ':' + str(choice_azs_truck_dict[i]['truck_id'])][
                            'min_rez1']
                else:
                    min_days_stock1_work_type_1 = 0
                    min_days_stock2_work_type_1 = 0
            '''**************************************************************************************************'''

            '''**************************************************************************************************'''
            # собираем все оцененные варианты расстановки в словарь
            choices_dict_work_type_1[choice] = {'variants': choice_azs_truck_dict,
                                                'points': points,
                                                'days_stock_min1': min_days_stock1_work_type_1,
                                                'days_stock_min2': min_days_stock2_work_type_1
                                                }
            '''**************************************************************************************************'''
            '''**************************************************************************************************'''
            # собираем все оцененные варианты расстановки в словарь
            choices_dict_work_type_2[choice] = {'variants': choice_azs_truck_dict,
                                                'points': points,
                                                'days_stock_min1': min_days_stock1_work_type_1,
                                                'days_stock_min2': min_days_stock2_work_type_1
                                                }
            '''**************************************************************************************************'''
            '''**************************************************************************************************'''
            # собираем все оцененные варианты расстановки в словарь
            choices_dict_work_type_3[choice] = {'variants': choice_azs_truck_dict,
                                                'points': points,
                                                'days_stock_min1': min_days_stock1_work_type_1,
                                                'days_stock_min2': min_days_stock2_work_type_1
                                                }
            '''**************************************************************************************************'''

        # сортируем полученные результаты по трем параметрам
        # На выходе получим отсортированный список ключей словаря choices_dict
        print('ASFASFUOALSFJOAISJFPAVJAOSFJIAOGIHJWPEGJOICSJ:GHSDLIHJS:DTJWEOGILASDOGJ')
        sort_choices_dict = sorted(choices_dict_work_type_1,
                                   key=lambda k: (choices_dict_work_type_1[k]['points'],
                                                  choices_dict_work_type_1[k]['days_stock_min1'],
                                                  choices_dict_work_type_1[k]['days_stock_min2']
                                                  ))

        # elif work_type.id == 2:
        #     print('Режим работы № 2 . Сортировка')
        #     if fuel_type == 92:
        #         sort_choices_dict = sorted(choices_dict_work_type_2,
        #
        #                                    key=lambda k: (choices_dict_work_type_1[k]['points'],
        #                                                   choices_dict_work_type_1[k]['days_stock_min1'],
        #                                                   choices_dict_work_type_1[k]['days_stock_min2']
        #                                                   ))
        #
        #     elif fuel_type == 95:
        #
        #         sort_choices_dict = sorted(choices_dict_work_type_2,
        #                                    key=lambda k: (choices_dict_work_type_1[k]['points'],
        #                                                   choices_dict_work_type_1[k]['days_stock_min1'],
        #                                                   choices_dict_work_type_1[k]['days_stock_min2']
        #                                                   ))
        #
        #     elif fuel_type == 50:
        #         sort_choices_dict = sorted(choices_dict_work_type_2,
        #
        #                                    key=lambda k: (choices_dict_work_type_1[k]['points'],
        #                                                   choices_dict_work_type_1[k]['days_stock_min1'],
        #                                                   choices_dict_work_type_1[k]['days_stock_min1']
        #                                                   ))
        #
        #     elif work_type.id == 3:
        #         print('Режим работы № 3 . Сортировка')
        #         if fuel_type == 92:
        #             sort_choices_dict = sorted(choices_dict_work_type_3,
        #                                        key=lambda k: (choices_dict_work_type_1[k]['points'],
        #                                                       choices_dict_work_type_1[k]['days_stock_min1'],
        #                                                       choices_dict_work_type_1[k]['days_stock_min2']
        #                                                       ))
        #
        #         elif fuel_type == 95:
        #             sort_choices_dict = sorted(choices_dict_work_type_3,
        #                                        key=lambda k: (choices_dict_work_type_1[k]['points'],
        #                                                       choices_dict_work_type_1[k]['days_stock_min1'],
        #                                                       choices_dict_work_type_1[k]['days_stock_min2']
        #                                                       ))
        #
        #         elif fuel_type == 50:
        #             sort_choices_dict = sorted(choices_dict_work_type_3,
        #                                        key=lambda k: (choices_dict_work_type_1[k]['points'],
        #                                                       choices_dict_work_type_1[k]['days_stock_min1'],
        #                                                       choices_dict_work_type_1[k]['days_stock_min1']
        #                                                       ))

        if work_type.id == 1:
            # создаем список, в котором будем хранить лучшие варианты расстановки
            best_choices = list()
            # берем айди предыдущего варианта расстановки
            trips_last = Trips.query.order_by(desc("calculate_id")).first()

            if not trips_last:
                previous_variant_id = 0
            else:
                previous_variant_id = trips_last.calculate_id

            # перебираем отсортированные от худшего к лучшему варианты расстановки
            for i in sort_choices_dict:
                # каждый вариант добавляем в список
                best_choices.append(i)
                # сокращаем список до 1
                best_choices = sort_choices_dict[-1:]
            # перебираем список из 10 лучших вариантов
            for i in best_choices:
                trips = Trips(trip_number=2, day=date.today(), date=datetime.today(), work_type_id=work_type.id,
                              calculate_id=previous_variant_id)
                db.session.add(trips)
                print(i, choices_dict_work_type_1[i]['points'], choices_dict_work_type_1[i]['days_stock_min1'],
                      choices_dict_work_type_1[i]['days_stock_min2'])

                for z in trucks_for_azs_dict:
                    trucks_for_azs = TrucksForAzs(azs_id=z,
                                                  number_of_trucks=len(trucks_for_azs_dict[z]['azs_trucks']) - 1,
                                                  calculate_id=previous_variant_id, trip_number=2)

                    db.session.add(trucks_for_azs)
                variants_sliva_for_trip = list()
                for x in choices_dict_work_type_1[i]['variants']:
                    if choices_dict_work_type_1[i]['variants'][x]['truck_id'] != 0:
                        variant_sliv_92 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])][
                            'variant_sliv_92']
                        variant = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])]['variant']
                        truck_id = choices_dict_work_type_1[i]['variants'][x]['truck_id']
                        azs_id = x
                        variant_sliv_95 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])][
                            'variant_sliv_95']
                        variant_sliv_50 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])][
                            'variant_sliv_50']
                        min_rez1 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])]['min_rez1']
                        min_rez2 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])]['min_rez2']
                        min_rez3 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_1[i]['variants'][x]['truck_id'])]['min_rez3']
                        query_variant = TempAzsTrucks4.query.filter_by(azs_id=azs_id, truck_id=truck_id,
                                                                       variant=variant,
                                                                       variant_sliv_92=variant_sliv_92).first()
                        calculate_id = previous_variant_id
                        result = Result(azs_id=azs_id, truck_id=truck_id,
                                        variant=variant,
                                        variant_sliv_92=variant_sliv_92,
                                        variant_sliv_95=variant_sliv_95,
                                        variant_sliv_50=variant_sliv_50,
                                        min_rez1=min_rez1,
                                        min_rez2=min_rez2,
                                        min_rez3=min_rez3,
                                        volume_92=query_variant.sum_92,
                                        volume_95=query_variant.sum_95,
                                        volume_50=query_variant.sum_50,
                                        calculate_id=calculate_id,
                                        trip_number=2)
                        db.session.add(result)

                        print("АЗС:", azs_id,
                              "Бензовоз:", truck_id,
                              "Вариант налива:", variant,
                              "Вариант слива 92:", variant_sliv_92,
                              "Вариант слива 95:", variant_sliv_95,
                              "Вариант слива 50:", variant_sliv_50)

                        variant_sliva = dict()

                        table_azs_trucks1 = TempAzsTrucks.query.filter_by(variant_id=int(variant)).all()
                        table_azs_trucks3_92 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                              variant_sliv=variant_sliv_92).all()
                        table_azs_trucks3_95 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                              variant_sliv=variant_sliv_95).all()
                        table_azs_trucks3_50 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                              variant_sliv=variant_sliv_50).all()

                        for table_variant in table_azs_trucks1:
                            variant_naliva = VariantNalivaForTrip(variant_from_table=int(variant),
                                                                  calculate_id=calculate_id,
                                                                  truck_tank_id=table_variant.truck_tank_id,
                                                                  truck_id=truck_id,
                                                                  azs_id=azs_id,
                                                                  fuel_type=table_variant.fuel_type,
                                                                  capacity=table_variant.capacity,
                                                                  trip_number=2)
                            db.session.add(variant_naliva)

                        for row in table_azs_trucks3_92:
                            variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                calculate_id=calculate_id,
                                                                azs_id=row.azs_id,
                                                                tank_id=row.tank_id,
                                                                truck_id=row.truck_id,
                                                                truck_tank_id=row.truck_tank_id_string,
                                                                fuel_type=row.fuel_type,
                                                                capacity=row.sum_sliv,
                                                                trip_number=2)
                            db.session.add(variant_sliva)

                        for row in table_azs_trucks3_95:
                            variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                calculate_id=calculate_id,
                                                                azs_id=row.azs_id,
                                                                tank_id=row.tank_id,
                                                                truck_id=row.truck_id,
                                                                truck_tank_id=row.truck_tank_id_string,
                                                                fuel_type=row.fuel_type,
                                                                capacity=row.sum_sliv,
                                                                trip_number=2)
                            db.session.add(variant_sliva)

                        for row in table_azs_trucks3_50:
                            variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                calculate_id=calculate_id,
                                                                azs_id=row.azs_id,
                                                                tank_id=row.tank_id,
                                                                truck_id=row.truck_id,
                                                                truck_tank_id=row.truck_tank_id_string,
                                                                fuel_type=row.fuel_type,
                                                                capacity=row.sum_sliv,
                                                                trip_number=2)
                            db.session.add(variant_sliva)
            db.session.commit()
        elif work_type.id == 2:
            # создаем список, в котором будем хранить лучшие варианты расстановки
            best_choices = list()
            # берем айди предыдущего варианта расстановки
            trips_last = Trips.query.order_by(desc("calculate_id")).first()

            if not trips_last:
                previous_variant_id = 0
            else:
                previous_variant_id = trips_last.calculate_id

            # перебираем отсортированные от худшего к лучшему варианты расстановки
            for i in sort_choices_dict:
                # каждый вариант добавляем в список
                best_choices.append(i)
                # сокращаем список до 1
                best_choices = sort_choices_dict[-1:]
            # перебираем список из 10 лучших вариантов
            for i in best_choices:
                trips = Trips(trip_number=2, day=date.today(), date=datetime.today(), work_type_id=work_type.id,
                              calculate_id=previous_variant_id)
                db.session.add(trips)
                print(i, choices_dict_work_type_2[i]['points'], choices_dict_work_type_2[i]['days_stock_min1'],
                      choices_dict_work_type_2[i]['days_stock_min2'])

                for z in trucks_for_azs_dict:
                    trucks_for_azs = TrucksForAzs(azs_id=z,
                                                  number_of_trucks=len(trucks_for_azs_dict[z]['azs_trucks']) - 1,
                                                  calculate_id=previous_variant_id, trip_number=2)

                    db.session.add(trucks_for_azs)
                variants_sliva_for_trip = list()
                for x in choices_dict_work_type_2[i]['variants']:
                    if choices_dict_work_type_2[i]['variants'][x]['truck_id'] != 0:
                        variant_sliv_92 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                            'variant_sliv_92']
                        variant = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])]['variant']
                        truck_id = choices_dict_work_type_2[i]['variants'][x]['truck_id']
                        azs_id = x
                        variant_sliv_95 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                            'variant_sliv_95']
                        variant_sliv_50 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])][
                            'variant_sliv_50']
                        min_rez1 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])]['min_rez1']
                        min_rez2 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])]['min_rez2']
                        min_rez3 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_2[i]['variants'][x]['truck_id'])]['min_rez3']
                        query_variant = TempAzsTrucks4.query.filter_by(azs_id=azs_id, truck_id=truck_id,
                                                                       variant=variant,
                                                                       variant_sliv_92=variant_sliv_92).first()
                        calculate_id = previous_variant_id
                        result = Result(azs_id=azs_id, truck_id=truck_id,
                                        variant=variant,
                                        variant_sliv_92=variant_sliv_92,
                                        variant_sliv_95=variant_sliv_95,
                                        variant_sliv_50=variant_sliv_50,
                                        min_rez1=min_rez1,
                                        min_rez2=min_rez2,
                                        min_rez3=min_rez3,
                                        volume_92=query_variant.sum_92,
                                        volume_95=query_variant.sum_95,
                                        volume_50=query_variant.sum_50,
                                        calculate_id=calculate_id,
                                        trip_number=2)
                        db.session.add(result)

                        print("АЗС:", azs_id,
                              "Бензовоз:", truck_id,
                              "Вариант налива:", variant,
                              "Вариант слива 92:", variant_sliv_92,
                              "Вариант слива 95:", variant_sliv_95,
                              "Вариант слива 50:", variant_sliv_50)

                        variant_sliva = dict()

                        table_azs_trucks1 = TempAzsTrucks.query.filter_by(variant_id=int(variant)).all()
                        table_azs_trucks3_92 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                              variant_sliv=variant_sliv_92).all()
                        table_azs_trucks3_95 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                              variant_sliv=variant_sliv_95).all()
                        table_azs_trucks3_50 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                              variant_sliv=variant_sliv_50).all()

                        for table_variant in table_azs_trucks1:
                            variant_naliva = VariantNalivaForTrip(variant_from_table=int(variant),
                                                                  calculate_id=calculate_id,
                                                                  truck_tank_id=table_variant.truck_tank_id,
                                                                  truck_id=truck_id,
                                                                  azs_id=azs_id,
                                                                  fuel_type=table_variant.fuel_type,
                                                                  capacity=table_variant.capacity,
                                                                  trip_number=2)
                            db.session.add(variant_naliva)

                        for row in table_azs_trucks3_92:
                            variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                calculate_id=calculate_id,
                                                                azs_id=row.azs_id,
                                                                tank_id=row.tank_id,
                                                                truck_id=row.truck_id,
                                                                truck_tank_id=row.truck_tank_id_string,
                                                                fuel_type=row.fuel_type,
                                                                capacity=row.sum_sliv,
                                                                trip_number=2)
                            db.session.add(variant_sliva)

                        for row in table_azs_trucks3_95:
                            variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                calculate_id=calculate_id,
                                                                azs_id=row.azs_id,
                                                                tank_id=row.tank_id,
                                                                truck_id=row.truck_id,
                                                                truck_tank_id=row.truck_tank_id_string,
                                                                fuel_type=row.fuel_type,
                                                                capacity=row.sum_sliv,
                                                                trip_number=2)
                            db.session.add(variant_sliva)

                        for row in table_azs_trucks3_50:
                            variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                calculate_id=calculate_id,
                                                                azs_id=row.azs_id,
                                                                tank_id=row.tank_id,
                                                                truck_id=row.truck_id,
                                                                truck_tank_id=row.truck_tank_id_string,
                                                                fuel_type=row.fuel_type,
                                                                capacity=row.sum_sliv,
                                                                trip_number=2)
                            db.session.add(variant_sliva)
            db.session.commit()
        elif work_type.id == 3:
            # создаем список, в котором будем хранить лучшие варианты расстановки
            best_choices = list()
            # берем айди предыдущего варианта расстановки
            trips_last = Trips.query.order_by(desc("calculate_id")).first()

            if not trips_last:
                previous_variant_id = 0
            else:
                previous_variant_id = trips_last.calculate_id

            # перебираем отсортированные от худшего к лучшему варианты расстановки
            for i in sort_choices_dict:
                # каждый вариант добавляем в список
                best_choices.append(i)
                # сокращаем список до 1
                best_choices = sort_choices_dict[-1:]
            # перебираем список из 10 лучших вариантов
            for i in best_choices:
                trips = Trips(trip_number=2, day=date.today(), date=datetime.today(), work_type_id=work_type.id,
                              calculate_id=previous_variant_id)
                db.session.add(trips)
                print(i, choices_dict_work_type_3[i]['points'], choices_dict_work_type_3[i]['days_stock_min1'],
                      choices_dict_work_type_3[i]['days_stock_min2'])

                for z in trucks_for_azs_dict:
                    trucks_for_azs = TrucksForAzs(azs_id=z,
                                                  number_of_trucks=len(trucks_for_azs_dict[z]['azs_trucks']) - 1,
                                                  calculate_id=previous_variant_id, trip_number=2)

                    db.session.add(trucks_for_azs)
                variants_sliva_for_trip = list()
                for x in choices_dict_work_type_3[i]['variants']:
                    if choices_dict_work_type_3[i]['variants'][x]['truck_id'] != 0:
                        variant_sliv_92 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])][
                            'variant_sliv_92']
                        variant = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])]['variant']
                        truck_id = choices_dict_work_type_3[i]['variants'][x]['truck_id']
                        azs_id = x
                        variant_sliv_95 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])][
                            'variant_sliv_95']
                        variant_sliv_50 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])][
                            'variant_sliv_50']
                        min_rez1 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])]['min_rez1']
                        min_rez2 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])]['min_rez2']
                        min_rez3 = azs_trucks_best_days_stock[
                            str(x) + ':' + str(choices_dict_work_type_3[i]['variants'][x]['truck_id'])]['min_rez3']
                        query_variant = TempAzsTrucks4.query.filter_by(azs_id=azs_id, truck_id=truck_id,
                                                                       variant=variant,
                                                                       variant_sliv_92=variant_sliv_92).first()
                        calculate_id = previous_variant_id
                        result = Result(azs_id=azs_id, truck_id=truck_id,
                                        variant=variant,
                                        variant_sliv_92=variant_sliv_92,
                                        variant_sliv_95=variant_sliv_95,
                                        variant_sliv_50=variant_sliv_50,
                                        min_rez1=min_rez1,
                                        min_rez2=min_rez2,
                                        min_rez3=min_rez3,
                                        volume_92=query_variant.sum_92,
                                        volume_95=query_variant.sum_95,
                                        volume_50=query_variant.sum_50,
                                        calculate_id=calculate_id,
                                        trip_number=2)
                        db.session.add(result)

                        print("АЗС:", azs_id,
                              "Бензовоз:", truck_id,
                              "Вариант налива:", variant,
                              "Вариант слива 92:", variant_sliv_92,
                              "Вариант слива 95:", variant_sliv_95,
                              "Вариант слива 50:", variant_sliv_50)

                        variant_sliva = dict()

                        table_azs_trucks1 = TempAzsTrucks.query.filter_by(variant_id=int(variant)).all()
                        table_azs_trucks3_92 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                              variant_sliv=variant_sliv_92).all()
                        table_azs_trucks3_95 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                              variant_sliv=variant_sliv_95).all()
                        table_azs_trucks3_50 = TempAzsTrucks3.query.filter_by(variant=int(variant),
                                                                              variant_sliv=variant_sliv_50).all()

                        for table_variant in table_azs_trucks1:
                            variant_naliva = VariantNalivaForTrip(variant_from_table=int(variant),
                                                                  calculate_id=calculate_id,
                                                                  truck_tank_id=table_variant.truck_tank_id,
                                                                  truck_id=truck_id,
                                                                  azs_id=azs_id,
                                                                  fuel_type=table_variant.fuel_type,
                                                                  capacity=table_variant.capacity,
                                                                  trip_number=2)
                            db.session.add(variant_naliva)

                        for row in table_azs_trucks3_92:
                            variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                calculate_id=calculate_id,
                                                                azs_id=row.azs_id,
                                                                tank_id=row.tank_id,
                                                                truck_id=row.truck_id,
                                                                truck_tank_id=row.truck_tank_id_string,
                                                                fuel_type=row.fuel_type,
                                                                capacity=row.sum_sliv,
                                                                trip_number=2)
                            db.session.add(variant_sliva)

                        for row in table_azs_trucks3_95:
                            variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                calculate_id=calculate_id,
                                                                azs_id=row.azs_id,
                                                                tank_id=row.tank_id,
                                                                truck_id=row.truck_id,
                                                                truck_tank_id=row.truck_tank_id_string,
                                                                fuel_type=row.fuel_type,
                                                                capacity=row.sum_sliv,
                                                                trip_number=2)
                            db.session.add(variant_sliva)

                        for row in table_azs_trucks3_50:
                            variant_sliva = VariantSlivaForTrip(variant_from_table=row.variant,
                                                                calculate_id=calculate_id,
                                                                azs_id=row.azs_id,
                                                                tank_id=row.tank_id,
                                                                truck_id=row.truck_id,
                                                                truck_tank_id=row.truck_tank_id_string,
                                                                fuel_type=row.fuel_type,
                                                                capacity=row.sum_sliv,
                                                                trip_number=2)
                            db.session.add(variant_sliva)
            db.session.commit()

    def time_to_return_second():
        # получаем информацию о втором рейсе
        second_trip = Trips.query.filter_by(trip_number=2).order_by(desc("calculate_id")).first()
        # получаем информацию о расстановке второго рейса
        second_trip_list = Result.query.filter_by(calculate_id=second_trip.calculate_id, trip_number=2).all()
        trip = Trip.query.all()
        trip_dict = dict()
        for i in trip:
            trip_dict[i.azs_id] = {'time_to': i.time_to,
                                   'time_from': i.time_from}
        # заполняем словарь с сопоставлением БЕНЗОВОЗ-АЗС из первого рейса
        for i in second_trip_list:
            full_time = trip_dict[i.azs_id]['time_to'] + trip_dict[i.azs_id]['time_from'] + 60
            trip_start_time = Result.query.filter_by(calculate_id=second_trip.calculate_id, truck_id=i.truck_id,
                                                     trip_number=1).first()
            t = trip_start_time.trip_end_time
            delta = timedelta(minutes=full_time)
            trip_end = (datetime.combine(date(1, 1, 1), t) + delta).time()
            result = Result.query.filter_by(calculate_id=second_trip.calculate_id, truck_id=i.truck_id,
                                            trip_number=2).first_or_404()
            result.time_to_return = full_time
            result.trip_end_time = trip_end
            db.session.commit()

    time_to_return_second()

    return redirect(url_for('main.trip_creation'))


@bp.route('/start_trip', methods=['POST', 'GET'])
@login_required
def start_trip():
    if current_user.get_task_in_progress('prepare_tables'):
        flash(_('Пересчет таблиц уже выполняется данных уже выполняется!'))
    else:
        current_user.launch_task('prepare_tables', _('Начат пересчет подготовительных таблиц...'))
        db.session.commit()
    return redirect(url_for('main.trip_creation'))


@bp.route('/restart_trip', methods=['POST', 'GET'])
@login_required
def restart_trip():
    last_trip = Trips.query.order_by(desc("calculate_id")).first()
    last_trip.incorrect = True
    db.session.commit()
    if current_user.get_task_in_progress('prepare_tables'):
        flash(_('Пересчет таблиц уже выполняется данных уже выполняется!'))
    else:
        current_user.launch_task('prepare_tables', _('Начат пересчет подготовительных таблиц...'))
        db.session.commit()
    return redirect(url_for('main.trip_creation'))


@bp.route('/trip_creation', methods=['POST', 'GET'])
@login_required
def trip_creation():
    print("STOP AT" + str(datetime.now()))
    trips = Trips.query.order_by(desc("calculate_id")).first()
    if trips.date.strftime("%d.%m.%Y") == datetime.today().strftime("%d.%m.%Y") and trips.incorrect != True:
        trips = True
    else:
        trips = False
    return render_template('trip_creation.html', title='Отправка бензовозов', trip_creation=True, trips=trips)


@bp.route('/load', methods=['POST', 'GET'])
@login_required
def load():
    # flash(_('Выполняется расстановка безовозов. Ожидайте завершения, страница будет перезагружена автоматически'))
    return render_template('load.html', title='Выполняется расстановка бензовозов')


@bp.route('/trips.json', methods=['POST', 'GET'])
@login_required
def trips_json():
    rows = list()
    priority = Priority.query.all()
    trips = Trips.query.order_by(desc("calculate_id")).first()
    for i in priority:
        tank = Tanks.query.filter_by(id=i.tank_id).first()
        azs = AzsList.query.filter_by(id=i.azs_id).first()
        result = Result.query.filter_by(calculate_id=trips.calculate_id, azs_id=i.azs_id).first()
        trucks_for_azs_first = TrucksForAzs.query.filter_by(azs_id=i.azs_id, calculate_id=trips.calculate_id,
                                                            trip_number=1).first()
        trucks_for_azs_second = TrucksForAzs.query.filter_by(azs_id=i.azs_id, calculate_id=trips.calculate_id,
                                                             trip_number=2).first()

        url_name = "АЗС № " + str(azs.number)
        url = '<a href="' + str(url_for('main.page_azs', id=i.azs_id)) + '">' + url_name + '</a>'

        if result:
            if result.trip_number == 1:
                trip_end = result.trip_end_time.strftime("%H:%M")
                trucks = Trucks.query.filter_by(id=result.truck_id).first()
                reg_number_first = trucks.reg_number
                new_day_stock_first = result.min_rez1
                new_day_stock_second = "-"
                reg_number_second = "-"
                number_of_trucks = trucks_for_azs_first.number_of_trucks
            elif result.trip_number == 2:
                trip_end = result.trip_end_time.strftime("%H:%M")
                trucks = Trucks.query.filter_by(id=result.truck_id).first()
                reg_number_second = trucks.reg_number
                new_day_stock_second = result.min_rez1
                number_of_trucks = trucks_for_azs_second.number_of_trucks
                new_day_stock_first = "-"
                reg_number_first = "-"
            else:
                reg_number_second = "-"
                trip_end = "-"
                reg_number_first = "-"
                number_of_trucks = trucks_for_azs_second.number_of_trucks
        else:
            reg_number_first = "-"
            trip_end = "-"
            reg_number_second = "-"
            new_day_stock_first = "-"
            new_day_stock_second = "-"
            if trucks_for_azs_second:
                number_of_trucks = trucks_for_azs_second.number_of_trucks
            else:
                number_of_trucks = "0"

        trip = Trip.query.filter_by(azs_id=i.azs_id).first()

        if trip.weigher == True:
            weighter_icon = ' <i alt="Весы" title="На пути к этой АЗС есть зона весового контроля" ' \
                            'class="fas fa-weight text-info"></i>'
        else:
            weighter_icon = ''

        row = {'priority': i.priority,
               'azs_number': url + weighter_icon,
               'tank_number': tank.tank_number,
               'day_stock': i.day_stock,
               'first_trip': reg_number_first,
               'second_trip': reg_number_second,
               'new_day_stock_first': new_day_stock_first,
               'new_day_stock_second': new_day_stock_second,
               'number_of_trucks': number_of_trucks,
               'datetime': trip_end,
               }
        rows.append(row)
    return Response(json.dumps(rows), mimetype='application/json')


@bp.route('/trip/<trip_number>/trip_naliv.json', methods=['POST', 'GET'])
@login_required
def trips_naliv_json(trip_number):
    trip_number = int(trip_number)
    rows = list()
    naliv_cell_1 = "Отсек отсутствует"
    naliv_cell_2 = "Отсек отсутствует"
    naliv_cell_3 = "Отсек отсутствует"
    naliv_cell_4 = "Отсек отсутствует"
    naliv_cell_5 = "Отсек отсутствует"
    naliv_cell_6 = "Отсек отсутствует"
    sliv_cell_1 = "Отсек отсутствует"
    sliv_cell_2 = "Отсек отсутствует"
    sliv_cell_3 = "Отсек отсутствует"
    sliv_cell_4 = "Отсек отсутствует"
    sliv_cell_5 = "Отсек отсутствует"
    sliv_cell_6 = "Отсек отсутствует"
    trips = Trips.query.order_by(desc("calculate_id")).first()
    result = Result.query.filter_by(calculate_id=trips.calculate_id).all()

    for i in result:
        naliv_cell_1 = "Отсек отсутствует"
        naliv_cell_2 = "Отсек отсутствует"
        naliv_cell_3 = "Отсек отсутствует"
        naliv_cell_4 = "Отсек отсутствует"
        naliv_cell_5 = "Отсек отсутствует"
        naliv_cell_6 = "Отсек отсутствует"
        sliv_cell_1 = "Отсек отсутствует"
        sliv_cell_2 = "Отсек отсутствует"
        sliv_cell_3 = "Отсек отсутствует"
        sliv_cell_4 = "Отсек отсутствует"
        sliv_cell_5 = "Отсек отсутствует"
        sliv_cell_6 = "Отсек отсутствует"
        trip = Trip.query.filter_by(azs_id=i.azs_id).first()
        if trip.weigher == True:
            weighter_icon = ' <i alt="Весы" title="На пути к этой АЗС есть зона весового контроля" ' \
                            'class="fas fa-weight text-info"></i>'
        else:
            weighter_icon = ''
        azs = AzsList.query.filter_by(id=i.azs_id).first()
        url_name = "АЗС № " + str(azs.number)
        url = '<a href="' + str(url_for('main.page_azs', id=i.azs_id)) + '">' + url_name + '</a>'
        if result:
            trucks = Trucks.query.filter_by(id=i.truck_id).first()
            reg_number = trucks.reg_number
            naliv = VariantNalivaForTrip.query.filter_by(azs_id=i.azs_id,
                                                         calculate_id=trips.calculate_id,
                                                         truck_id=i.truck_id,
                                                         variant_from_table=i.variant).all()

            sliv = VariantSlivaForTrip.query.filter_by(azs_id=i.azs_id,
                                                       calculate_id=trips.calculate_id,
                                                       truck_id=i.truck_id,
                                                       variant_from_table=i.variant).all()
            for var in naliv:
                truck_tanks = TruckTanks.query.filter_by(id=var.truck_tank_id).first()
                if truck_tanks.number == 1:
                    naliv_cell_1 = var.fuel_type
                if truck_tanks.number == 2:
                    naliv_cell_2 = var.fuel_type
                if truck_tanks.number == 3:
                    naliv_cell_3 = var.fuel_type
                if truck_tanks.number == 4:
                    naliv_cell_4 = var.fuel_type
                if truck_tanks.number == 5:
                    naliv_cell_5 = var.fuel_type
                if truck_tanks.number == 6:
                    naliv_cell_6 = var.fuel_type

                for var in sliv:
                    truck_tanks_list = var.truck_tank_id.split("+")
                    for truck_tank in truck_tanks_list:
                        truck_tank_list = TruckTanks.query.filter_by(id=truck_tank).first()
                        tank = Tanks.query.filter_by(id=var.tank_id).first()
                        if truck_tank_list.number == 1:
                            sliv_cell_1 = "Резервуар №" + str(tank.tank_number) + " (" + str(var.fuel_type) + ")"
                        if truck_tank_list.number == 2:
                            sliv_cell_2 = "Резервуар №" + str(tank.tank_number) + " (" + str(var.fuel_type) + ")"
                        if truck_tank_list.number == 3:
                            sliv_cell_3 = "Резервуар №" + str(tank.tank_number) + " (" + str(var.fuel_type) + ")"
                        if truck_tank_list.number == 4:
                            sliv_cell_4 = "Резервуар №" + str(tank.tank_number) + " (" + str(var.fuel_type) + ")"
                        if truck_tank_list.number == 5:
                            sliv_cell_5 = "Резервуар №" + str(tank.tank_number) + " (" + str(var.fuel_type) + ")"
                        if truck_tank_list.number == 6:
                            sliv_cell_6 = "Резервуар №" + str(tank.tank_number) + " (" + str(var.fuel_type) + ")"
        else:
            reg_number = "-"
        if trip_number == 1 and i.trip_number == 1:
            row = {'azs_number': url + weighter_icon,
                   'truck_number': reg_number,
                   'naliv_cell_1': naliv_cell_1,
                   'naliv_cell_2': naliv_cell_2,
                   'naliv_cell_3': naliv_cell_3,
                   'naliv_cell_4': naliv_cell_4,
                   'naliv_cell_5': naliv_cell_5,
                   'naliv_cell_6': naliv_cell_6,
                   'sliv_cell_1': sliv_cell_1,
                   'sliv_cell_2': sliv_cell_2,
                   'sliv_cell_3': sliv_cell_3,
                   'sliv_cell_4': sliv_cell_4,
                   'sliv_cell_5': sliv_cell_5,
                   'sliv_cell_6': sliv_cell_6,
                   }
            rows.append(row)
        elif i.trip_number == 2 and trip_number == 2:
            row = {'azs_number': url + weighter_icon,
                   'truck_number': reg_number,
                   'naliv_cell_1': naliv_cell_1,
                   'naliv_cell_2': naliv_cell_2,
                   'naliv_cell_3': naliv_cell_3,
                   'naliv_cell_4': naliv_cell_4,
                   'naliv_cell_5': naliv_cell_5,
                   'naliv_cell_6': naliv_cell_6,
                   'sliv_cell_1': sliv_cell_1,
                   'sliv_cell_2': sliv_cell_2,
                   'sliv_cell_3': sliv_cell_3,
                   'sliv_cell_4': sliv_cell_4,
                   'sliv_cell_5': sliv_cell_5,
                   'sliv_cell_6': sliv_cell_6,
                   }
            rows.append(row)

    return Response(json.dumps(rows), mimetype='application/json')


@bp.route('/testform/<id>', methods=['POST', 'GET'])
@login_required
def test_form(id):
    cells_list = list()
    cells = TruckTanks.query.filter_by(truck_id=id).all()
    cells_count = TruckTanks.query.filter_by(truck_id=id).count()
    CellsForm.cell = FieldList(FormField(FuelForm), min_entries=cells_count)
    form = CellsForm()
    add_data = TruckTanksVariations.query.order_by(desc('variant_good')).first()
    if add_data:
        variant_good = add_data.variant_good + 1
    else:
        variant_good = 1
    for i in cells:
        cells_list.append(i.number)

    if form.validate_on_submit():
        for entry in form.cell.entries:
            truck_cell_id = TruckTanks.query.filter_by(truck_id=id, number=entry.data['id']).first()
            add_data = TruckTanksVariations(variant_good=variant_good, truck_tank_id=truck_cell_id.id, truck_id=id,
                                            diesel=entry.data['fuel_type'])
            db.session.add(add_data)

            print(entry.data['id'])
            print(entry.data['fuel_type'])

        db.session.commit()
        return redirect(url_for('admin.truck', id=id))
    return render_template('admin/add_cell.html', form=form)


@bp.route('/recreate_trip', methods=['POST', 'GET'])
@login_required
def recreate_trip():
    last_trip = Trips.query.order_by(desc("calculate_id")).first()
    last_trips = Trips.query.filter_by(calculate_id=last_trip.calculate_id).all()
    for i in last_trips:
        i.incorrect = True
        db.session.commit()
    return render_template('recreate_trip.html')


@bp.route('/creation_failed', methods=['POST', 'GET'])
@login_required
def creation_failed():
    return render_template('creation_failed.html', title="Расстановка не удалась")


@bp.route('/trip_history', methods=['POST', 'GET'])
@login_required
def trip_history():
    page = request.args.get('page', 1, type=int)
    posts = Trips.query.filter_by(incorrect=None).order_by(desc("date")).group_by(sa.func.year(Trips.day),
                                                                                  sa.func.month(Trips.day),
                                                                                  sa.func.day(Trips.day)).paginate(page,
                                                                                                                   10,
                                                                                                                   False)
    next_url = url_for('main.trip_history', page=posts.next_num) \
        if posts.has_next else None
    prev_url = url_for('main.trip_history', page=posts.prev_num) \
        if posts.has_prev else None
    results = Result.query.all()

    return render_template('trip_history.html', title="История расстановок", posts=posts.items, next_url=next_url,
                           prev_url=prev_url, result=results)


@bp.route('/trip_xlsx', methods=['POST', 'GET'])
@login_required
def trip_xlsx_maker():
    trips = Trips.query.order_by(desc("calculate_id")).first()
    result = Result.query.filter_by(calculate_id=trips.calculate_id).all()
    zadanie_first = dict()
    zadanie_second = dict()
    for truck in result:
        if truck.trip_number == 1:
            zadanie_first[truck.truck_id] = {'id': 0,
                                   'reg_number': 0,
                                   }
        elif truck.trip_number == 2:
            zadanie_second[truck.truck_id] = {'id': 0,
                                             'reg_number': 0,
                                             }
    if result:
        for i in result:
            if i.trip_number == 1:
                azs = AzsList.query.filter_by(id=i.azs_id).first()
                truck = Trucks.query.filter_by(id=i.truck_id).first()

                zadanie_first[truck.id] = {'id': i.id,
                                           'reg_number': truck.reg_number,
                                           'azs_number': azs.number,
                                           'day': trips.day
                                           }

            elif i.trip_number == 2:
                azs = AzsList.query.filter_by(id=i.azs_id).first()
                truck = Trucks.query.filter_by(id=i.truck_id).first()
                reg_number_first = truck.reg_number
                new_day_stock_first = i.min_rez1

                zadanie_second[truck.id] = {'id': i.id,
                                           'reg_number': truck.reg_number,
                                           'azs_number': azs.number,
                                           'day': trips.day
                                           }
    path = pathlib.Path().absolute()
    file_path = '/app/static/xls/zadanie.xlsx'
    wb = load_workbook(str(path) + file_path)
    for i in zadanie_first:
        ws = wb.create_sheet(zadanie_first[i]['reg_number'])
        sheet = wb.active
    wb.save(filename=str(path) + '/app/static/xls/test.xlsx')

    wb = load_workbook(str(path) + '/app/static/xls/test.xlsx')
